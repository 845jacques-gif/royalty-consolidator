import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\jacques\Documents\202601_PLYGRND Model_v3.xlsx"
print("Loading workbook (145MB, this may take several minutes)...", flush=True)

# Load with read_only mode for speed on large files
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

print("Workbook loaded. Reading B1_Model...", flush=True)
ws = wb['B1_Model']

print("=== B1_Model: Period Headers (rows 47-51, cols 25-60) ===", flush=True)
row_num = 0
for row in ws.iter_rows(min_row=47, max_row=51, min_col=25, max_col=60, values_only=False):
    r = 47 + row_num
    row_data = []
    for cell in row:
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            row_data.append(f"{col_letter}{r}(col{cell.column}): {cell.value}")
    if row_data:
        print(f"Row {r}: {' | '.join(row_data)}", flush=True)
    row_num += 1

print("\n=== B1_Model: First data row (52), cols 25-60 ===", flush=True)
row_data = []
for row in ws.iter_rows(min_row=52, max_row=52, min_col=25, max_col=60, values_only=False):
    for cell in row:
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            row_data.append(f"{col_letter}52(col{cell.column}): {cell.value}")
if row_data:
    print(' | '.join(row_data), flush=True)

print("\n=== B1_Model: Data row 52, cols 50-100 (where Jan 2022 should be) ===", flush=True)
row_data = []
for row in ws.iter_rows(min_row=52, max_row=52, min_col=50, max_col=100, values_only=False):
    for cell in row:
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            row_data.append(f"{col_letter}52(col{cell.column}): {cell.value}")
if row_data:
    print(' | '.join(row_data), flush=True)

print("\nReading RJ_Model...", flush=True)
ws2 = wb['RJ_Model']

print("=== RJ_Model: Period Headers (rows 47-51, cols 25-55) ===", flush=True)
row_num = 0
for row in ws2.iter_rows(min_row=47, max_row=51, min_col=25, max_col=55, values_only=False):
    r = 47 + row_num
    row_data = []
    for cell in row:
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            row_data.append(f"{col_letter}{r}(col{cell.column}): {cell.value}")
    if row_data:
        print(f"Row {r}: {' | '.join(row_data)}", flush=True)
    row_num += 1

print("\n=== RJ_Model: First data row (52), cols 25-55 ===", flush=True)
row_data = []
for row in ws2.iter_rows(min_row=52, max_row=52, min_col=25, max_col=56, values_only=False):
    for cell in row:
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            row_data.append(f"{col_letter}52(col{cell.column}): {cell.value}")
if row_data:
    print(' | '.join(row_data), flush=True)

print("\n=== B1_Model: Config cells (rows 23-42, cols 2-7) ===", flush=True)
# Need to re-read B1_Model for different rows - read_only mode iterates sequentially
# So we read the whole needed range
ws_b1 = wb['B1_Model']
row_num = 0
for row in ws_b1.iter_rows(min_row=23, max_row=42, min_col=2, max_col=7, values_only=False):
    r = 23 + row_num
    row_data = []
    for cell in row:
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            row_data.append(f"{col_letter}{r}: {cell.value}")
    if row_data:
        print(' | '.join(row_data), flush=True)
    row_num += 1

print("\n=== Metadata: Header row and first 3 data rows ===", flush=True)
ws3 = wb['Metadata']
row_num = 0
for row in ws3.iter_rows(min_row=3, max_row=8, min_col=1, max_col=11, values_only=False):
    r = 3 + row_num
    row_data = []
    for cell in row:
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            row_data.append(f"{col_letter}{r}: {cell.value}")
    if row_data:
        print(' | '.join(row_data), flush=True)
    row_num += 1

wb.close()
print("\nDone.", flush=True)
