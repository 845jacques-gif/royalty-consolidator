"""
Tests for Forecast UX changes:
  - parse_sofr_from_excel() utility
  - /api/parse-sofr-excel API route
  - Beta gate password protection
  - Per-payor config table rendering
  - Form submission with hidden JSON fields
"""

import json
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from forecast import parse_sofr_from_excel


# ---------------------------------------------------------------------------
# Helpers: create test Excel workbooks
# ---------------------------------------------------------------------------

def _make_sofr_workbook(path, sheet_name='SOFR', rows=None, rate_as_pct=False):
    """Create a minimal Excel workbook with a SOFR sheet for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Headers in row 7 (data starts at row 8)
    ws.cell(row=7, column=2, value='Date Lookup')
    ws.cell(row=7, column=4, value='SOFR Rate')

    if rows is None:
        from datetime import date
        rows = [
            (date(2025, 12, 31), 0.036517),
            (date(2026, 1, 31), 0.036532),
            (date(2026, 2, 28), 0.036065),
            (date(2026, 3, 31), 0.035200),
            (date(2026, 4, 30), 0.034500),
        ]

    for i, (dt, rate) in enumerate(rows):
        ws.cell(row=8 + i, column=2, value=dt)
        ws.cell(row=8 + i, column=4, value=rate * 100 if rate_as_pct else rate)

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Unit Tests: parse_sofr_from_excel()
# ---------------------------------------------------------------------------

class TestParseSofrFromExcel:

    def test_basic_parse(self, tmp_path):
        """Parse a normal SOFR workbook with 5 data points."""
        path = str(tmp_path / 'test.xlsx')
        _make_sofr_workbook(path)

        curve = parse_sofr_from_excel(path)

        assert len(curve) == 5
        assert curve[0]['date'] == '2025-12-31'
        assert abs(curve[0]['rate'] - 0.036517) < 1e-5
        assert curve[4]['date'] == '2026-04-30'

    def test_no_sofr_sheet(self, tmp_path):
        """Raise ValueError when SOFR sheet is missing."""
        path = str(tmp_path / 'no_sofr.xlsx')
        _make_sofr_workbook(path, sheet_name='NotSOFR')

        with pytest.raises(ValueError, match="No 'SOFR' sheet"):
            parse_sofr_from_excel(path)

    def test_case_insensitive_sheet_name(self, tmp_path):
        """Find SOFR sheet regardless of case."""
        path = str(tmp_path / 'sofr_case.xlsx')
        _make_sofr_workbook(path, sheet_name='Sofr')

        curve = parse_sofr_from_excel(path)
        assert len(curve) == 5

    def test_percentage_rates_auto_convert(self, tmp_path):
        """Rates > 1 are auto-divided by 100."""
        path = str(tmp_path / 'pct.xlsx')
        _make_sofr_workbook(path, rate_as_pct=True)

        curve = parse_sofr_from_excel(path)
        # All rates should be small decimals (< 0.1)
        for pt in curve:
            assert pt['rate'] < 0.1, f"Rate {pt['rate']} was not converted from percentage"

    def test_skip_none_cells(self, tmp_path):
        """Rows with None date or rate are skipped."""
        from datetime import date
        path = str(tmp_path / 'gaps.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SOFR'
        # Row 8: valid
        ws.cell(row=8, column=2, value=date(2025, 12, 31))
        ws.cell(row=8, column=4, value=0.035)
        # Row 9: missing rate
        ws.cell(row=9, column=2, value=date(2026, 1, 31))
        ws.cell(row=9, column=4, value=None)
        # Row 10: missing date
        ws.cell(row=10, column=2, value=None)
        ws.cell(row=10, column=4, value=0.034)
        # Row 11: both valid
        ws.cell(row=11, column=2, value=date(2026, 2, 28))
        ws.cell(row=11, column=4, value=0.033)
        wb.save(path)
        wb.close()

        curve = parse_sofr_from_excel(path)
        assert len(curve) == 2  # Only rows 8 and 11
        assert curve[0]['date'] == '2025-12-31'
        assert curve[1]['date'] == '2026-02-28'

    def test_empty_sheet(self, tmp_path):
        """Empty SOFR sheet returns empty list."""
        path = str(tmp_path / 'empty.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SOFR'
        wb.save(path)
        wb.close()

        curve = parse_sofr_from_excel(path)
        assert curve == []

    def test_real_lucki_model(self):
        """Parse SOFR from the actual Lucki model (if available)."""
        model_path = r'C:\Users\jacques\Documents\202602_Lucki_Model_v1.xlsx'
        if not os.path.isfile(model_path):
            pytest.skip('Lucki model not found')

        curve = parse_sofr_from_excel(model_path)
        assert len(curve) == 121, f"Expected 121 data points, got {len(curve)}"
        # First point: Dec 2025
        assert curve[0]['date'] == '2025-12-31'
        # Last point: Dec 2035
        assert curve[-1]['date'] == '2035-12-31'
        # Rates in reasonable range
        for pt in curve:
            assert 0.01 < pt['rate'] < 0.10, f"Rate {pt['rate']} out of range for {pt['date']}"


# ---------------------------------------------------------------------------
# Flask App Tests
# ---------------------------------------------------------------------------

@pytest.fixture(scope='module')
def client():
    """Create a Flask test client. Requires deal data to exist on disk."""
    # Import app but don't start the server
    os.environ.setdefault('FORECAST_BETA_PASSWORD', 'testpass123')
    from app import app
    app.config['TESTING'] = True
    app.config['FORECAST_BETA_PASSWORD'] = 'testpass123'

    with app.test_client() as c:
        yield c


def _find_test_deal_slug(client):
    """Find an available deal slug for testing."""
    r = client.get('/deals')
    if r.status_code != 200:
        return None
    # Extract a deal slug from the page
    import re
    m = re.search(r'href="/deals/([^/]+)/forecast"', r.data.decode())
    return m.group(1) if m else None


class TestSofrApiRoute:

    def test_no_file_returns_400(self, client):
        """POST without file returns 400."""
        r = client.post('/api/parse-sofr-excel')
        assert r.status_code == 400
        data = r.get_json()
        assert 'error' in data

    def test_valid_file_returns_curve(self, client, tmp_path):
        """POST with valid SOFR workbook returns curve data."""
        path = str(tmp_path / 'test_api.xlsx')
        _make_sofr_workbook(path)

        with open(path, 'rb') as f:
            r = client.post('/api/parse-sofr-excel',
                            data={'sofr_file': (f, 'test.xlsx')},
                            content_type='multipart/form-data')

        assert r.status_code == 200
        data = r.get_json()
        assert data['count'] == 5
        assert len(data['curve']) == 5
        assert data['curve'][0]['date'] == '2025-12-31'

    def test_no_sofr_sheet_returns_error(self, client, tmp_path):
        """POST with workbook missing SOFR sheet returns error."""
        path = str(tmp_path / 'no_sofr_api.xlsx')
        _make_sofr_workbook(path, sheet_name='NotSOFR')

        with open(path, 'rb') as f:
            r = client.post('/api/parse-sofr-excel',
                            data={'sofr_file': (f, 'bad.xlsx')},
                            content_type='multipart/form-data')

        assert r.status_code == 400
        data = r.get_json()
        assert 'error' in data
        assert 'SOFR' in data['error']


class TestBetaGate:

    def test_forecast_page_shows_gate_when_locked(self, client):
        """GET forecast page without session shows beta gate."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess.pop('forecast_unlocked', None)

        r = client.get(f'/deals/{slug}/forecast')
        assert r.status_code == 200
        html = r.data.decode()
        assert 'Forecast Beta' in html
        assert 'beta_password' in html

    def test_wrong_password_shows_error(self, client):
        """POST with wrong password shows error."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess.pop('forecast_unlocked', None)

        r = client.post(f'/deals/{slug}/forecast',
                        data={'beta_password': 'wrong'})
        assert r.status_code == 200
        html = r.data.decode()
        assert 'Incorrect password' in html

    def test_correct_password_unlocks(self, client):
        """POST with correct password sets session and redirects."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess.pop('forecast_unlocked', None)

        r = client.post(f'/deals/{slug}/forecast',
                        data={'beta_password': 'testpass123'},
                        follow_redirects=False)
        assert r.status_code == 302
        assert f'/deals/{slug}/forecast' in r.headers.get('Location', '')

    def test_unlocked_session_shows_forecast_form(self, client):
        """GET with unlocked session shows full forecast form."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess['forecast_unlocked'] = True

        r = client.get(f'/deals/{slug}/forecast')
        assert r.status_code == 200
        html = r.data.decode()
        assert 'Run Forecast' in html
        assert 'genre_default' in html

    def test_download_blocked_when_locked(self, client):
        """Download route redirects when not unlocked."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess.pop('forecast_unlocked', None)

        r = client.get(f'/deals/{slug}/forecast/download', follow_redirects=False)
        assert r.status_code == 302


class TestPayorConfigTable:

    def test_table_renders_with_payor_data(self, client):
        """Forecast form shows per-payor table with data-payor attributes."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess['forecast_unlocked'] = True

        r = client.get(f'/deals/{slug}/forecast')
        html = r.data.decode()
        assert 'payorConfigTable' in html
        assert 'data-payor=' in html
        assert 'pc-rights' in html
        assert 'pc-fee' in html
        assert 'pc-source-ccy' in html

    def test_sofr_import_ui_present(self, client):
        """Forecast form shows SOFR import UI elements."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess['forecast_unlocked'] = True

        r = client.get(f'/deals/{slug}/forecast')
        html = r.data.decode()
        assert 'sofrFileInput' in html
        assert 'importSofrExcel' in html
        assert 'sofrCurveJson' in html
        assert 'sofrPreview' in html

    def test_hidden_fields_present(self, client):
        """Hidden JSON fields exist for form submission."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess['forecast_unlocked'] = True

        r = client.get(f'/deals/{slug}/forecast')
        html = r.data.decode()
        assert 'name="sofr_curve_json"' in html
        assert 'name="payor_configs_json"' in html
        assert 'name="fx_rates_json"' in html

    def test_js_functions_present(self, client):
        """JavaScript functions for SOFR import and payor config are present."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess['forecast_unlocked'] = True

        r = client.get(f'/deals/{slug}/forecast')
        html = r.data.decode()
        assert 'function importSofrExcel' in html
        assert 'function collectPayorConfigs' in html
        assert 'function renderSofrPreview' in html


class TestForecastFormSubmission:

    def test_submit_with_json_fields(self, client):
        """Submit forecast with SOFR curve and payor config JSON hidden fields."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess['forecast_unlocked'] = True

        sofr_curve = json.dumps([
            {'date': '2025-12-31', 'rate': 0.036517},
            {'date': '2026-01-31', 'rate': 0.036532},
        ])
        payor_configs = json.dumps({
            'EM': {'income_rights': 'Masters', 'fee_rate': 0.0,
                   'fx_currency': 'USD', 'synergy': False}
        })

        r = client.post(f'/deals/{slug}/forecast', data={
            'genre_default': 'default',
            'horizon_years': '5',
            'discount_rate': '9.375',
            'exit_multiple': '15',
            'purchase_price': '5000000',
            'ltv': '55',
            'sofr_rate': '4.5',
            'sofr_floor': '2.0',
            'sofr_spread': '275',
            'cash_flow_sweep': '100',
            'synergy_ramp_months': '12',
            'sofr_curve_json': sofr_curve,
            'payor_configs_json': payor_configs,
            'fx_rates_json': '{}',
        })

        assert r.status_code == 200
        html = r.data.decode()
        assert 'Projected Revenue by Year' in html or 'forecast_result' in html.lower()

    def test_submit_without_optional_json_fields(self, client):
        """Submit forecast without SOFR/payor JSON — should still work."""
        slug = _find_test_deal_slug(client)
        if not slug:
            pytest.skip('No deals available')

        with client.session_transaction() as sess:
            sess['forecast_unlocked'] = True

        r = client.post(f'/deals/{slug}/forecast', data={
            'genre_default': 'default',
            'horizon_years': '5',
            'discount_rate': '9.375',
            'exit_multiple': '15',
            'purchase_price': '0',
            'ltv': '55',
            'sofr_rate': '4.5',
            'sofr_floor': '2.0',
            'sofr_spread': '275',
            'cash_flow_sweep': '100',
            'synergy_ramp_months': '12',
            'sofr_curve_json': '',
            'payor_configs_json': '',
            'fx_rates_json': '',
        })

        assert r.status_code == 200


class TestDealsPageBetaLabel:

    def test_shows_beta_label_when_locked(self, client):
        """Deals page shows 'Forecast (Beta)' when session is locked."""
        with client.session_transaction() as sess:
            sess.pop('forecast_unlocked', None)

        r = client.get('/deals')
        if r.status_code != 200:
            pytest.skip('Deals page not accessible')
        html = r.data.decode()
        if 'forecast' in html.lower():
            assert 'Forecast (Beta)' in html

    def test_shows_forecast_label_when_unlocked(self, client):
        """Deals page shows 'Forecast' (without Beta) when unlocked."""
        with client.session_transaction() as sess:
            sess['forecast_unlocked'] = True

        r = client.get('/deals')
        if r.status_code != 200:
            pytest.skip('Deals page not accessible')
        html = r.data.decode()
        if 'forecast' in html.lower():
            # Should have "Forecast" but not "Forecast (Beta)"
            assert '>Forecast<' in html or '>Forecast </a>' in html.replace('\n', '')
