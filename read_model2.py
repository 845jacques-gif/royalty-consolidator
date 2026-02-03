import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\jacques\Documents\202601_PLYGRND Model_v3.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

focus_sheets = ['RJ_Model', 'B2_Model', 'B1_Model', 'Song Listing_By RecordJet', 'Song Listing_By B1', 'Song Listing_By B2', 'Artist_Splits', 'Decay Curve']

for sheet_name in focus_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"MISSING: {sheet_name}")
        continue
    
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    
    if '_Model' in sheet_name:
        row_num = 0
        last_rows = []
        for row in ws.rows:
            row_num += 1
            if row_num <= 55:
                row_data = []
                for cell in row[:35]:
                    if cell.value is not None:
                        val = str(cell.value)
                        if len(val) > 60:
                            val = val[:60] + "..."
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        row_data.append(f"{col_letter}{row_num}: {val}")
                if row_data:
                    print(" | ".join(row_data[:12]))
            elif row_num <= 60:
                if row_num == 56:
                    print("\n--- Data rows 56-60 ---")
                row_data = []
                for cell in row[:100]:
                    if cell.value is not None:
                        val = str(cell.value)
                        if len(val) > 40:
                            val = val[:40] + "..."
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        row_data.append(f"{col_letter}{row_num}: {val}")
                if row_data:
                    print(" | ".join(row_data[:15]))
            else:
                row_data = []
                for cell in row[:100]:
                    if cell.value is not None:
                        val = str(cell.value)
                        if len(val) > 40:
                            val = val[:40] + "..."
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        row_data.append(f"{col_letter}{row_num}: {val}")
                if row_data:
                    last_rows.append(" | ".join(row_data[:15]))
                    if len(last_rows) > 4:
                        last_rows.pop(0)
        
        print(f"\nTotal rows: {row_num}")
        if last_rows:
            print(f"\n--- Last few data rows ---")
            for lr in last_rows:
                print(lr)
    
    else:
        row_num = 0
        total = 0
        for row in ws.rows:
            row_num += 1
            total += 1
            if row_num <= 20:
                row_data = []
                for cell in row[:30]:
                    if cell.value is not None:
                        val = str(cell.value)
                        if len(val) > 50:
                            val = val[:50] + "..."
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        row_data.append(f"{col_letter}{row_num}: {val}")
                if row_data:
                    print(" | ".join(row_data[:12]))
        print(f"Total rows: {total}")

    sys.stdout.flush()

wb.close()
print("\nDone.")
