import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\jacques\Documents\202601_PLYGRND Model_v3.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

# Check B1_Model header rows (47-51) for the monthly column labels
ws = wb['B1_Model']

print("=== B1_Model: Period Headers (rows 47-51, cols 25-60) ===")
for r in range(47, 52):
    row_data = []
    for c in range(25, 61):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_data.append(f"{col_letter}{r}(col{c}): {val}")
    if row_data:
        print(f"Row {r}: {' | '.join(row_data)}")

print("\n=== B1_Model: First data row (52), cols 25-60 ===")
row_data = []
for c in range(25, 61):
    val = ws.cell(row=52, column=c).value
    if val is not None:
        col_letter = openpyxl.utils.get_column_letter(c)
        row_data.append(f"{col_letter}52(col{c}): {val}")
if row_data:
    print(' | '.join(row_data))

print("\n=== B1_Model: Data row 52, cols 50-100 (where Jan 2022 should be) ===")
row_data = []
for c in range(50, 101):
    val = ws.cell(row=52, column=c).value
    if val is not None:
        col_letter = openpyxl.utils.get_column_letter(c)
        row_data.append(f"{col_letter}52(col{c}): {val}")
if row_data:
    print(' | '.join(row_data))

# Also check RJ_Model which has data from 2021
print("\n=== RJ_Model: Period Headers (rows 47-51, cols 25-45) ===")
ws2 = wb['RJ_Model']
for r in range(47, 52):
    row_data = []
    for c in range(25, 55):
        val = ws2.cell(row=r, column=c).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_data.append(f"{col_letter}{r}(col{c}): {val}")
    if row_data:
        print(f"Row {r}: {' | '.join(row_data)}")

print("\n=== RJ_Model: First data row (52), cols 25-55 ===")
row_data = []
for c in range(25, 56):
    val = ws2.cell(row=52, column=c).value
    if val is not None:
        col_letter = openpyxl.utils.get_column_letter(c)
        row_data.append(f"{col_letter}52(col{c}): {val}")
if row_data:
    print(' | '.join(row_data))

# Check what config cells look like
print("\n=== B1_Model: Config cells ===")
for r in range(23, 43):
    row_data = []
    for c in range(2, 8):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_data.append(f"{col_letter}{r}: {val}")
    if row_data:
        print(' | '.join(row_data))

# Check Metadata columns
print("\n=== Metadata: Header row and first 3 data rows ===")
ws3 = wb['Metadata']
for r in range(3, 9):
    row_data = []
    for c in range(1, 12):
        val = ws3.cell(row=r, column=c).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_data.append(f"{col_letter}{r}: {val}")
    if row_data:
        print(' | '.join(row_data))

wb.close()
