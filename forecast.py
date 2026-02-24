"""
Forecast Projection Engine — DCF-style decay curve projections for music catalog valuation.

Uses FTI Report decay curves by genre to project per-ISRC revenue over a configurable horizon.
Supports dual terminal value methods (exit multiple + perpetuity growth), excl/incl synergy tracks,
sensitivity tables, unlevered + levered returns (IRR/MOIC), a full debt schedule, and Excel export.
"""

import json
import logging
import math
import os
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

log = logging.getLogger('royalty')

# ---------------------------------------------------------------------------
# Decay Curves (FTI Report)
# ---------------------------------------------------------------------------

DECAY_CURVES = {
    'r&b': {
        'label': 'R&B / Soul',
        'rates': {
            1: -0.55, 2: -0.40, 3: -0.28, 4: -0.20, 5: -0.15,
            6: -0.10, 7: -0.10, 8: -0.08, 9: -0.08, 10: -0.07,
        },
        'terminal': 0.01,
    },
    'rap': {
        'label': 'Rap / Hip-Hop',
        'rates': {
            1: -0.65, 2: -0.41, 3: -0.29, 4: -0.22, 5: -0.17,
            6: -0.10, 7: -0.10, 8: -0.07, 9: -0.07, 10: -0.07,
        },
        'terminal': 0.01,
    },
    'edm': {
        'label': 'EDM / Latin / Rock',
        'rates': {
            1: -0.49, 2: -0.35, 3: -0.27, 4: -0.20, 5: -0.17,
            6: -0.11, 7: -0.10, 8: -0.08, 9: -0.07, 10: -0.07,
        },
        'terminal': 0.01,
    },
    'pop': {
        'label': 'Pop',
        'rates': {
            1: -0.49, 2: -0.35, 3: -0.27, 4: -0.20, 5: -0.17,
            6: -0.11, 7: -0.10, 8: -0.08, 9: -0.07, 10: -0.07,
        },
        'terminal': 0.01,
    },
    'default': {
        'label': 'Default (EDM/Latin/Rock)',
        'rates': {
            1: -0.49, 2: -0.35, 3: -0.27, 4: -0.20, 5: -0.17,
            6: -0.11, 7: -0.10, 8: -0.08, 9: -0.07, 10: -0.07,
        },
        'terminal': 0.01,
    },
}

# Virtu-style aliases — match model naming conventions
DECAY_CURVE_ALIASES = {
    'virtu - r&b': 'r&b',
    'virtu - rap': 'rap',
    'virtu - electronic': 'edm',
    'virtu - pop': 'pop',
    'virtu - default': 'default',
}


def resolve_curve(name: str) -> dict:
    """Resolve a curve name (or Virtu alias) to a DECAY_CURVES entry."""
    key = name.lower().strip()
    key = DECAY_CURVE_ALIASES.get(key, key)
    return DECAY_CURVES.get(key, DECAY_CURVES['default'])


GENRE_CHOICES = [
    ('default', 'Default (EDM/Latin/Rock)'),
    ('r&b', 'R&B / Soul'),
    ('rap', 'Rap / Hip-Hop'),
    ('edm', 'EDM / Latin / Rock'),
    ('pop', 'Pop'),
    ('virtu - r&b', 'Virtu - R&B'),
    ('virtu - rap', 'Virtu - Rap'),
    ('virtu - electronic', 'Virtu - Electronic'),
]


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

@dataclass
class ForecastConfig:
    """Configuration for a forecast projection."""
    # --- Core projection ---
    genre_default: str = 'default'
    genre_overrides: Dict[str, str] = field(default_factory=dict)  # {isrc: genre}
    base_period: str = 'ltm'
    horizon_years: int = 5
    discount_rate: float = 0.09375  # CMG WACC — 9.375% default
    new_fee_rate: Optional[float] = None
    synergy_start_year: int = 1
    reversions: Dict[str, str] = field(default_factory=dict)  # {isrc: 'YYYY-MM-DD'}
    terminal_growth: Optional[float] = None
    # --- Transaction ---
    purchase_price: float = 0.0
    exit_multiple: float = 15.0
    # --- Debt structure ---
    ltv: float = 0.55
    sofr_rate: float = 0.045
    sofr_floor: float = 0.02
    sofr_spread: float = 0.0275
    cash_flow_sweep: float = 1.0
    # --- Synergy ramp ---
    synergy_ramp_months: int = 12
    third_party_synergy_rate: Optional[float] = None

    # --- NEW: Dual WACC ---
    virtu_wacc: Optional[float] = None  # e.g. 0.09; None = skip Virtu calcs

    # --- NEW: Purchase Structure ---
    holdback: float = 0.0
    pcdpcdr: float = 0.0  # Post-Close Date Pre-Cash Date Royalties
    cash_date: Optional[str] = None   # YYYY-MM-DD
    close_date: Optional[str] = None  # YYYY-MM-DD

    # --- NEW: Deal Metadata ---
    opportunity_name: str = ''
    opportunity_details: str = 'Catalog Acquisition'
    rights_included: str = 'Masters'  # Masters / Publishing / NR / combo
    deal_type: str = 'Catalog'        # Catalog / Corporate / JV

    # --- NEW: SOFR Forward Curve ---
    sofr_curve: List[dict] = field(default_factory=list)
    # [{date: 'YYYY-MM-DD', rate: 0.0365}, ...] — if non-empty, overrides flat sofr_rate

    # --- NEW: Per-ISRC overrides ---
    isrc_fee_overrides: Dict[str, float] = field(default_factory=dict)   # {isrc: fee_rate}
    isrc_label_shares: Dict[str, float] = field(default_factory=dict)    # {isrc: label_share_pct}
    isrc_3p_shares: Dict[str, float] = field(default_factory=dict)       # {isrc: 3p_share_pct}
    isrc_decay_curves: Dict[str, str] = field(default_factory=dict)      # {isrc: curve_name}

    # --- NEW: Per-Payor config ---
    payor_configs: Dict[str, dict] = field(default_factory=dict)
    # {payor_code: {income_rights, fx_currency, fx_rate, fee_rate, synergy, synergy_date, label_engine_fee}}

    # --- NEW: FX Rates ---
    fx_rates: Dict[str, float] = field(default_factory=dict)  # {currency_code: rate_to_usd}
    base_currency: str = 'USD'

    # --- NEW: Sensitivity config ---
    irr_purchase_prices: List[float] = field(default_factory=list)  # IRR/MOIC grid rows
    irr_exit_multiples: List[float] = field(default_factory=list)   # IRR/MOIC grid cols

    @property
    def closing_amount(self) -> float:
        """Net cash at closing = purchase_price - holdback - pcdpcdr."""
        return self.purchase_price - self.holdback - self.pcdpcdr

    def to_dict(self) -> dict:
        return {
            'genre_default': self.genre_default,
            'genre_overrides': self.genre_overrides,
            'base_period': self.base_period,
            'horizon_years': self.horizon_years,
            'discount_rate': self.discount_rate,
            'new_fee_rate': self.new_fee_rate,
            'synergy_start_year': self.synergy_start_year,
            'reversions': self.reversions,
            'terminal_growth': self.terminal_growth,
            'purchase_price': self.purchase_price,
            'exit_multiple': self.exit_multiple,
            'ltv': self.ltv,
            'sofr_rate': self.sofr_rate,
            'sofr_floor': self.sofr_floor,
            'sofr_spread': self.sofr_spread,
            'cash_flow_sweep': self.cash_flow_sweep,
            'synergy_ramp_months': self.synergy_ramp_months,
            'third_party_synergy_rate': self.third_party_synergy_rate,
            'virtu_wacc': self.virtu_wacc,
            'holdback': self.holdback,
            'pcdpcdr': self.pcdpcdr,
            'cash_date': self.cash_date,
            'close_date': self.close_date,
            'opportunity_name': self.opportunity_name,
            'opportunity_details': self.opportunity_details,
            'rights_included': self.rights_included,
            'deal_type': self.deal_type,
            'sofr_curve': self.sofr_curve,
            'isrc_fee_overrides': self.isrc_fee_overrides,
            'isrc_label_shares': self.isrc_label_shares,
            'isrc_3p_shares': self.isrc_3p_shares,
            'isrc_decay_curves': self.isrc_decay_curves,
            'payor_configs': self.payor_configs,
            'fx_rates': self.fx_rates,
            'base_currency': self.base_currency,
            'irr_purchase_prices': self.irr_purchase_prices,
            'irr_exit_multiples': self.irr_exit_multiples,
        }

    @staticmethod
    def from_dict(d: dict) -> 'ForecastConfig':
        return ForecastConfig(
            genre_default=d.get('genre_default', 'default'),
            genre_overrides=d.get('genre_overrides', {}),
            base_period=d.get('base_period', 'ltm'),
            horizon_years=d.get('horizon_years', 5),
            discount_rate=d.get('discount_rate', 0.09375),
            new_fee_rate=d.get('new_fee_rate'),
            synergy_start_year=d.get('synergy_start_year', 1),
            reversions=d.get('reversions', {}),
            terminal_growth=d.get('terminal_growth'),
            purchase_price=d.get('purchase_price', 0.0),
            exit_multiple=d.get('exit_multiple', 15.0),
            ltv=d.get('ltv', 0.55),
            sofr_rate=d.get('sofr_rate', 0.045),
            sofr_floor=d.get('sofr_floor', 0.02),
            sofr_spread=d.get('sofr_spread', 0.0275),
            cash_flow_sweep=d.get('cash_flow_sweep', 1.0),
            synergy_ramp_months=d.get('synergy_ramp_months', 12),
            third_party_synergy_rate=d.get('third_party_synergy_rate'),
            virtu_wacc=d.get('virtu_wacc'),
            holdback=d.get('holdback', 0.0),
            pcdpcdr=d.get('pcdpcdr', 0.0),
            cash_date=d.get('cash_date'),
            close_date=d.get('close_date'),
            opportunity_name=d.get('opportunity_name', ''),
            opportunity_details=d.get('opportunity_details', 'Catalog Acquisition'),
            rights_included=d.get('rights_included', 'Masters'),
            deal_type=d.get('deal_type', 'Catalog'),
            sofr_curve=d.get('sofr_curve', []),
            isrc_fee_overrides=d.get('isrc_fee_overrides', {}),
            isrc_label_shares=d.get('isrc_label_shares', {}),
            isrc_3p_shares=d.get('isrc_3p_shares', {}),
            isrc_decay_curves=d.get('isrc_decay_curves', {}),
            payor_configs=d.get('payor_configs', {}),
            fx_rates=d.get('fx_rates', {}),
            base_currency=d.get('base_currency', 'USD'),
            irr_purchase_prices=d.get('irr_purchase_prices', []),
            irr_exit_multiples=d.get('irr_exit_multiples', []),
        )


# ---------------------------------------------------------------------------
# SOFR Excel Import
# ---------------------------------------------------------------------------

def parse_sofr_from_excel(filepath: str) -> List[dict]:
    """Parse SOFR forward curve from a Chatham Financial-style Excel sheet.

    Expects a sheet named 'SOFR' (case-insensitive) with:
      - Column B: Date (end-of-month dates), rows 8-128
      - Column D: SOFR rate as decimal (e.g. 0.0365)

    Returns list of {"date": "YYYY-MM-DD", "rate": float} dicts,
    matching ForecastConfig.sofr_curve format.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # Find SOFR sheet (case-insensitive)
    sofr_sheet = None
    for name in wb.sheetnames:
        if name.lower().strip() == 'sofr':
            sofr_sheet = wb[name]
            break
    if sofr_sheet is None:
        wb.close()
        raise ValueError("No 'SOFR' sheet found in workbook")

    curve = []
    for row in sofr_sheet.iter_rows(min_row=8, max_row=128, min_col=2, max_col=4):
        date_cell = row[0]   # Column B
        rate_cell = row[2]   # Column D

        if date_cell.value is None or rate_cell.value is None:
            continue

        # Parse date
        dt_val = date_cell.value
        if isinstance(dt_val, datetime):
            dt_str = dt_val.strftime('%Y-%m-%d')
        elif isinstance(dt_val, date):
            dt_str = dt_val.strftime('%Y-%m-%d')
        elif isinstance(dt_val, str):
            dt_str = dt_val.strip()
        else:
            continue

        # Parse rate
        try:
            rate = float(rate_cell.value)
        except (ValueError, TypeError):
            continue

        # Sanity check: rate should be a small decimal (0-1 range typical)
        if rate > 1:
            rate = rate / 100  # Convert from percentage if needed

        curve.append({'date': dt_str, 'rate': round(rate, 6)})

    wb.close()
    return curve


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _xirr(cashflows: List[Tuple[date, float]], guess: float = 0.10) -> Optional[float]:
    """Compute XIRR using Newton-Raphson. Returns None if no convergence.

    cashflows: list of (date, amount) tuples.
    """
    if not cashflows or len(cashflows) < 2:
        return None

    d0 = cashflows[0][0]
    # Convert dates to year fractions
    years = [(cf[0] - d0).days / 365.25 for cf in cashflows]
    amounts = [cf[1] for cf in cashflows]

    rate = guess
    for _ in range(200):
        npv = sum(a / ((1 + rate) ** t) if (1 + rate) > 0 else 0
                  for a, t in zip(amounts, years))
        dnpv = sum(-t * a / ((1 + rate) ** (t + 1)) if (1 + rate) > 0 else 0
                   for a, t in zip(amounts, years))
        if abs(dnpv) < 1e-12:
            break
        new_rate = rate - npv / dnpv
        if abs(new_rate - rate) < 1e-8:
            return round(new_rate, 6)
        rate = new_rate
        # Guard against divergence
        if rate < -0.99 or rate > 100:
            break

    return None


def compute_synergy_ramp(year: int, config: ForecastConfig) -> float:
    """Compute synergy ramp factor (0.0 to 1.0) for a given projection year.

    Linear interpolation over synergy_ramp_months starting at synergy_start_year.
    Year is 1-indexed projection year.
    """
    if config.new_fee_rate is None and config.third_party_synergy_rate is None:
        return 0.0

    start_year = config.synergy_start_year
    ramp_months = max(1, config.synergy_ramp_months)

    # Months into projection at end of this year
    months_into_projection = year * 12
    # Months since synergy start
    synergy_start_month = (start_year - 1) * 12
    months_since_start = months_into_projection - synergy_start_month

    if months_since_start <= 0:
        return 0.0
    if months_since_start >= ramp_months:
        return 1.0

    return months_since_start / ramp_months


def compute_synergy_ramp_for_payor(year: int, config: ForecastConfig,
                                    payor_code: Optional[str] = None) -> tuple:
    """Compute per-payor synergy ramp, fee rate, and TP synergy rate.

    Checks payor_configs[code] for per-payor overrides, falls back to global config.

    Returns:
        (ramp, effective_new_fee_rate, effective_tp_synergy_rate)
    """
    pc = config.payor_configs.get(payor_code, {}) if payor_code else {}

    # Per-payor synergy flag — if explicitly False, no synergy for this payor
    payor_synergy = pc.get('synergy')
    if payor_synergy is False:
        return (0.0, None, 0.0)

    # Resolve synergy parameters: per-payor overrides > global config
    new_fee_rate = pc.get('synergy_new_fee_rate', config.new_fee_rate)
    tp_synergy_rate = pc.get('synergy_tp_rate', config.third_party_synergy_rate) or 0.0
    start_year = pc.get('synergy_start_year', config.synergy_start_year)
    ramp_months = pc.get('synergy_ramp_months', config.synergy_ramp_months)

    if new_fee_rate is None and tp_synergy_rate == 0:
        return (0.0, None, 0.0)

    ramp_months = max(1, ramp_months)
    months_into_projection = year * 12
    synergy_start_month = (start_year - 1) * 12
    months_since_start = months_into_projection - synergy_start_month

    if months_since_start <= 0:
        return (0.0, new_fee_rate, tp_synergy_rate)
    if months_since_start >= ramp_months:
        return (1.0, new_fee_rate, tp_synergy_rate)

    ramp = months_since_start / ramp_months
    return (ramp, new_fee_rate, tp_synergy_rate)


def lookup_sofr_rate(sofr_curve: List[dict], target_date: date,
                     floor: float = 0.02) -> float:
    """Find the SOFR rate for a given date from the forward curve.

    Uses nearest-date lookup (picks entry with closest date <= target, or first entry).
    Returns max(rate, floor).
    """
    if not sofr_curve:
        return floor

    # Parse and sort curve entries
    parsed = []
    for entry in sofr_curve:
        try:
            d_str = str(entry.get('date', ''))
            rate = float(entry.get('rate', 0))
            d = date.fromisoformat(d_str[:10])
            parsed.append((d, rate))
        except (ValueError, TypeError):
            continue

    if not parsed:
        return floor

    parsed.sort(key=lambda x: x[0])

    # Find closest date <= target_date
    best_rate = parsed[0][1]
    for d, r in parsed:
        if d <= target_date:
            best_rate = r
        else:
            break

    return max(best_rate, floor)


def get_interest_rate_for_year(year: int, config: ForecastConfig,
                               forecast_start: date) -> float:
    """Get the applicable interest rate for a projection year.

    If sofr_curve is populated, uses forward curve lookup.
    Otherwise uses flat sofr_rate.
    """
    if config.sofr_curve:
        target = date(forecast_start.year + year, 7, 1)
        sofr = lookup_sofr_rate(config.sofr_curve, target, config.sofr_floor)
    else:
        sofr = max(config.sofr_rate, config.sofr_floor)
    return sofr + config.sofr_spread


# ---------------------------------------------------------------------------
# Core Engine
# ---------------------------------------------------------------------------

def estimate_age_from_trajectory(yearly_earnings: Dict[int, float],
                                  curve: dict,
                                  reference_year: int,
                                  max_age: float = 10.0) -> Optional[float]:
    """Estimate an ISRC's catalog age by matching its observed YoY decay to the curve.

    For ISRCs without release dates, we look at how their earnings change
    year-over-year and find where on the decay curve that rate corresponds.
    This lets old songs (which have flat trajectories) get assigned appropriate
    high ages, rather than a uniform default.

    Important: individual songs should NOT get terminal growth rates.  The
    terminal rate models aggregate catalog behavior, not individual tracks.
    Songs with flat or growing trajectories are capped at max_age (default 10),
    which is the last explicit data point on the curve.

    Args:
        yearly_earnings: {year: gross_earnings} for at least 2 consecutive years
        curve: DECAY_CURVES entry with 'rates' and 'terminal'
        reference_year: The year at which we want to know the age (typically forecast start year)
        max_age: Maximum catalog age to assign (default 10.0 — the last explicit curve point)

    Returns:
        Estimated catalog age at reference_year, or None if insufficient data.
    """
    # Need at least 2 years of data to compute YoY changes
    sorted_years = sorted(y for y, g in yearly_earnings.items() if g > 0)
    if len(sorted_years) < 2:
        return None

    # Compute YoY growth rates for consecutive year pairs
    yoy_rates = []
    for i in range(len(sorted_years) - 1):
        y1, y2 = sorted_years[i], sorted_years[i + 1]
        if y2 - y1 != 1:  # Only consecutive years
            continue
        g1 = yearly_earnings[y1]
        g2 = yearly_earnings[y2]
        if g1 > 0:
            rate = (g2 - g1) / g1
            # Clamp to reasonable range (-0.80 to +0.50) to avoid outliers
            rate = max(-0.80, min(0.50, rate))
            yoy_rates.append((rate, y2))  # (rate, in which year this rate was observed)

    if not yoy_rates:
        return None

    # Use median rate (robust to outliers like viral spikes)
    rates_only = sorted(r for r, _ in yoy_rates)
    median_rate = rates_only[len(rates_only) // 2]

    # Find the age on the curve that produces this rate (inverse lookup).
    # Only search up to max_age — do NOT use terminal growth for individual songs.
    curve_rates = curve['rates']
    max_explicit_age = max(curve_rates.keys())  # typically 10
    min_curve_rate = curve_rates.get(max_explicit_age, -0.07)  # flattest explicit rate

    # Songs with rates flatter than the curve's flattest explicit rate are
    # outperforming the curve.  Two sub-cases:
    # (a) Mild outperformance (rate < +5%): likely old catalog at the flat
    #     end of the curve.  Assign age ~(max_explicit - 2) at observation
    #     time (e.g. age 8 for a 10-year curve).  Ages 8-10 all have the
    #     same rate on most curves, so the exact age barely matters for Y1.
    # (b) Strong growth (rate >= +5%): viral/playlist boost — can't
    #     reliably estimate.  Fall through to WAA back-solve.
    if median_rate >= min_curve_rate:
        if median_rate < 0.05:
            # Flat-trajectory old catalog → assign to flat end of curve
            estimated_age = max(max_explicit_age - 2, 1.0)  # typically 8
            most_recent_year = max(y for _, y in yoy_rates)
            years_to_ref = reference_year - most_recent_year
            estimated_age += years_to_ref
            max_base_age = max_age - 1
            return max(min(estimated_age, max_base_age), 1.0)
        else:
            return None  # Viral — fall through to WAA
    else:
        # Search ages 1.0 to max_age in 0.1 steps
        best_age = max_age
        best_diff = float('inf')
        for age_tenths in range(10, int(max_age * 10) + 1):
            age = age_tenths / 10.0
            lower = int(age)
            upper = min(max_explicit_age, lower + 1)
            lr = curve_rates.get(lower, min_curve_rate)
            ur = curve_rates.get(upper, min_curve_rate)
            frac = age - lower
            cr = lr + (ur - lr) * frac

            diff = abs(cr - median_rate)
            if diff < best_diff:
                best_diff = diff
                best_age = age

    # best_age is the age at which the curve rate best matches the observed rate.
    # The median rate was observed in the most recent pair of years.
    # Adjust to reference_year.
    most_recent_year = max(y for _, y in yoy_rates)
    years_to_ref = reference_year - most_recent_year
    estimated_age = best_age + years_to_ref

    # Cap so that Y1 catalog_age (= estimated_age + 1) stays within the
    # explicit curve range.  This prevents individual songs from hitting
    # terminal growth.
    max_base_age = max_age - 1  # Y1 age = base + 1 → stays ≤ max_age
    return max(min(estimated_age, max_base_age), 1.0)


def interpolate_decay_rate(catalog_age_years: float, curve: dict,
                           terminal_blend: float = 1.0) -> float:
    """Interpolate the decay rate for a given catalog age using curve data.

    For ages 0-1, interpolates between 0 (no decay for brand-new) and year-1 rate.
    For ages 1-10, interpolates between bracketing year rates.
    For ages 10+, uses terminal growth rate blended with the flattest explicit rate.

    Args:
        terminal_blend: 0.0 = use flattest explicit rate, 1.0 = full terminal growth,
            0.5 = midpoint. ISRCs with uncertain estimated ages should use < 1.0
            because their true age distribution spans both sides of the terminal cutoff.
    """
    rates = curve['rates']
    terminal = curve.get('terminal', 0.01)

    if catalog_age_years <= 0:
        return 0.0  # Brand new release, no decay yet

    if catalog_age_years >= 10:
        if terminal_blend >= 1.0:
            return terminal
        max_explicit = max(rates.keys())
        flat_rate = rates.get(max_explicit, terminal)
        return flat_rate + (terminal - flat_rate) * terminal_blend

    # Ages 0-1: linear ramp from 0 (no decay) to year-1 rate
    if catalog_age_years < 1.0:
        year1_rate = rates.get(1, -0.50)
        return year1_rate * catalog_age_years

    lower_year = int(catalog_age_years)
    upper_year = min(10, lower_year + 1)

    if lower_year == upper_year or lower_year >= 10:
        return rates.get(min(lower_year, 10), terminal)

    lower_rate = rates.get(lower_year, terminal)
    upper_rate = rates.get(upper_year, terminal)

    fraction = catalog_age_years - lower_year
    return lower_rate + (upper_rate - lower_rate) * fraction


def project_isrc(baseline: dict, release_date_str: Optional[str],
                 config: ForecastConfig, genre: str,
                 forecast_start_date: date = None,
                 fee_ratio: float = 0.15,
                 tp_ratio: float = 0.0,
                 label_share: Optional[float] = None,
                 tp_share: Optional[float] = None,
                 curve_name: Optional[str] = None,
                 artist_share: float = 0.0,
                 terminal_blend: float = 1.0,
                 payor_code: Optional[str] = None) -> List[dict]:
    """Project revenue for a single ISRC over the forecast horizon.

    Args:
        label_share: Per-ISRC label share override (0-1). If set, used instead of fee_ratio.
        tp_share: Per-ISRC third-party share override (0-1). If set, used instead of tp_ratio.
        curve_name: Explicit decay curve name (supports Virtu aliases). Overrides genre param.
        artist_share: Fraction of net receipts paid to artist/label (0-1). Deducted after
                      fees and before third-party share. Default 0.
        terminal_blend: How much terminal growth to apply when catalog_age > 10.
                        1.0 = full terminal (known age), 0.5 = blend (uncertain age),
                        0.0 = stay at flattest explicit rate (highly uncertain).

    Returns list of dicts per year with BOTH excl and incl synergy tracks:
      {year, gross, fees_original, net_receipts_excl, artist_share_excl, third_party_excl,
       net_earnings_excl, fee_savings, tp_savings, net_earnings_incl, decay_rate, catalog_age}
    """
    if forecast_start_date is None:
        forecast_start_date = date.today()

    # Per-ISRC overrides take precedence
    if label_share is not None:
        fee_ratio = 1.0 - label_share  # label_share = 1 - fee_ratio
    if tp_share is not None:
        tp_ratio = tp_share

    curve = resolve_curve(curve_name) if curve_name else resolve_curve(genre)
    horizon = config.horizon_years
    base_gross = baseline.get('gross', 0)

    if base_gross <= 0:
        return [{'year': y + 1, 'gross': 0, 'fees_original': 0,
                 'net_receipts_excl': 0, 'artist_share_excl': 0,
                 'third_party_excl': 0, 'net_earnings_excl': 0,
                 'fee_savings': 0, 'tp_savings': 0, 'net_earnings_incl': 0,
                 'decay_rate': 0, 'catalog_age': 0} for y in range(horizon)]

    # Parse release date for age calculation
    release_date = None
    if release_date_str:
        try:
            if len(release_date_str) >= 10:
                release_date = date.fromisoformat(release_date_str[:10])
            elif len(release_date_str) == 4:
                release_date = date(int(release_date_str), 7, 1)
        except (ValueError, TypeError):
            pass

    if release_date:
        base_age = (forecast_start_date - release_date).days / 365.25
    else:
        base_age = 3.0

    projections = []
    prev_gross = base_gross

    for y in range(1, horizon + 1):
        catalog_age = base_age + y
        decay_rate = interpolate_decay_rate(catalog_age, curve,
                                            terminal_blend=terminal_blend)

        year_gross = prev_gross * (1 + decay_rate)

        # --- Excl synergies track (original fee structure) ---
        fees_original = year_gross * fee_ratio
        net_receipts_excl = year_gross - fees_original
        artist_share_excl = net_receipts_excl * artist_share
        after_artist = net_receipts_excl - artist_share_excl
        third_party_excl = after_artist * tp_ratio
        net_earnings_excl = after_artist - third_party_excl

        # --- Synergy savings (per-payor aware) ---
        ramp, eff_new_fee, eff_tp_syn = compute_synergy_ramp_for_payor(y, config, payor_code)

        # Fee savings
        fee_savings = 0.0
        if eff_new_fee is not None and ramp > 0:
            blended_fee_rate = fee_ratio * (1 - ramp) + eff_new_fee * ramp
            fees_synergy = year_gross * blended_fee_rate
            fee_savings = fees_original - fees_synergy
        else:
            fee_savings = 0.0

        # Third-party savings
        tp_savings = 0.0
        if eff_tp_syn > 0 and ramp > 0:
            tp_savings = third_party_excl * eff_tp_syn * ramp

        # --- Incl synergies track ---
        # Fee savings reduce fees → more NR → more after-artist → more NE
        # We need to recompute the full waterfall with synergy fees
        if fee_savings > 0:
            synergy_nr = net_receipts_excl + fee_savings
            synergy_artist = synergy_nr * artist_share
            synergy_after_artist = synergy_nr - synergy_artist
            synergy_tp = synergy_after_artist * tp_ratio
            net_earnings_incl = synergy_after_artist - synergy_tp + tp_savings
        else:
            net_earnings_incl = net_earnings_excl + tp_savings

        # Apply reversion: zero out after reversion date
        isrc = baseline.get('isrc', '')
        reversion_date_str = config.reversions.get(isrc)
        if reversion_date_str:
            try:
                rev_date = date.fromisoformat(reversion_date_str[:10])
                forecast_year_date = date(forecast_start_date.year + y, 1, 1)
                if forecast_year_date >= rev_date:
                    year_gross = 0
                    fees_original = 0
                    net_receipts_excl = 0
                    artist_share_excl = 0
                    third_party_excl = 0
                    net_earnings_excl = 0
                    fee_savings = 0
                    tp_savings = 0
                    net_earnings_incl = 0
            except (ValueError, TypeError):
                pass

        projections.append({
            'year': y,
            'gross': round(year_gross, 2),
            'fees_original': round(fees_original, 2),
            'net_receipts_excl': round(net_receipts_excl, 2),
            'artist_share_excl': round(artist_share_excl, 2),
            'third_party_excl': round(third_party_excl, 2),
            'net_earnings_excl': round(net_earnings_excl, 2),
            'fee_savings': round(fee_savings, 2),
            'tp_savings': round(tp_savings, 2),
            'net_earnings_incl': round(net_earnings_incl, 2),
            'decay_rate': round(decay_rate, 4),
            'catalog_age': round(catalog_age, 1),
        })

        prev_gross = year_gross if year_gross > 0 else prev_gross * 0.5

    return projections


# ---------------------------------------------------------------------------
# Valuation Functions
# ---------------------------------------------------------------------------

def compute_dcf_analysis(year_totals: List[dict], config: ForecastConfig,
                         ltm_net: float = 0) -> dict:
    """Compute DCF analysis with both terminal value methods × both synergy tracks.

    Returns {terminal_multiple: {excl: {...}, incl: {...}},
             perpetuity_growth: {excl: {...}, incl: {...}}}
    """
    wacc = config.discount_rate
    tgr = config.terminal_growth if config.terminal_growth is not None else 0.01
    exit_mult = config.exit_multiple
    n = config.horizon_years

    result = {
        'terminal_multiple': {'excl': {}, 'incl': {}},
        'perpetuity_growth': {'excl': {}, 'incl': {}},
    }

    for track in ('excl', 'incl'):
        ne_key = f'net_earnings_{track}'

        # PV of cash flows
        pv_cfs = 0.0
        for yt in year_totals:
            y = yt['year']
            ne = yt.get(ne_key, 0)
            df = 1 / ((1 + wacc) ** y)
            pv_cfs += ne * df

        last_ne = year_totals[-1].get(ne_key, 0) if year_totals else 0
        exit_df = 1 / ((1 + wacc) ** n) if n > 0 else 1

        # Terminal Multiple Method
        tv_mult = last_ne * exit_mult
        pv_tv_mult = tv_mult * exit_df
        implied_mult = pv_cfs + pv_tv_mult

        result['terminal_multiple'][track] = {
            'pv_cash_flows': round(pv_cfs, 2),
            'exit_year_ne': round(last_ne, 2),
            'exit_multiple': exit_mult,
            'terminal_value': round(tv_mult, 2),
            'pv_terminal_value': round(pv_tv_mult, 2),
            'implied_valuation': round(implied_mult, 2),
            'implied_ltm_multiple': round(implied_mult / ltm_net, 2) if ltm_net > 0 else 0,
        }

        # Perpetuity Growth Method
        if wacc > tgr:
            tv_perp = last_ne * (1 + tgr) / (wacc - tgr)
        else:
            tv_perp = last_ne * 20  # Cap
        pv_tv_perp = tv_perp * exit_df
        implied_perp = pv_cfs + pv_tv_perp

        result['perpetuity_growth'][track] = {
            'pv_cash_flows': round(pv_cfs, 2),
            'exit_year_ne': round(last_ne, 2),
            'terminal_growth_rate': tgr,
            'terminal_value': round(tv_perp, 2),
            'pv_terminal_value': round(pv_tv_perp, 2),
            'implied_valuation': round(implied_perp, 2),
            'implied_ltm_multiple': round(implied_perp / ltm_net, 2) if ltm_net > 0 else 0,
        }

    return result


def compute_sensitivity_tables(year_totals: List[dict], config: ForecastConfig,
                               ltm_net: float = 0) -> dict:
    """Compute 3x3 sensitivity matrices for both methods.

    Terminal Multiple: WACC rows (±1%) × Exit Multiple cols (±2x)
    Perpetuity Growth: WACC rows (±1%) × TGR cols (±0.5%)
    """
    base_wacc = config.discount_rate
    base_exit_mult = config.exit_multiple
    base_tgr = config.terminal_growth if config.terminal_growth is not None else 0.01
    n = config.horizon_years

    # Track: use incl synergies for sensitivity
    ne_key = 'net_earnings_incl'
    last_ne = year_totals[-1].get(ne_key, 0) if year_totals else 0

    # WACC grid: base -1%, base, base +1%
    wacc_values = [round(base_wacc - 0.01, 4), round(base_wacc, 4), round(base_wacc + 0.01, 4)]

    # --- Terminal Multiple sensitivity ---
    exit_mult_values = [base_exit_mult - 2.0, base_exit_mult, base_exit_mult + 2.0]
    tm_matrix_excl = []
    tm_matrix_incl = []

    for w in wacc_values:
        row_excl = []
        row_incl = []
        for em in exit_mult_values:
            for track, row in [('excl', row_excl), ('incl', row_incl)]:
                tk = f'net_earnings_{track}'
                pv_cfs = sum(yt.get(tk, 0) / ((1 + w) ** yt['year']) for yt in year_totals)
                ln = year_totals[-1].get(tk, 0) if year_totals else 0
                tv = ln * em
                pv_tv = tv / ((1 + w) ** n) if n > 0 else tv
                row.append(round(pv_cfs + pv_tv, 2))
        tm_matrix_excl.append(row_excl)
        tm_matrix_incl.append(row_incl)

    # --- Perpetuity Growth sensitivity ---
    tgr_values = [round(base_tgr - 0.005, 4), round(base_tgr, 4), round(base_tgr + 0.005, 4)]
    pg_matrix_excl = []
    pg_matrix_incl = []

    for w in wacc_values:
        row_excl = []
        row_incl = []
        for g in tgr_values:
            for track, row in [('excl', row_excl), ('incl', row_incl)]:
                tk = f'net_earnings_{track}'
                pv_cfs = sum(yt.get(tk, 0) / ((1 + w) ** yt['year']) for yt in year_totals)
                ln = year_totals[-1].get(tk, 0) if year_totals else 0
                if w > g:
                    tv = ln * (1 + g) / (w - g)
                else:
                    tv = ln * 20
                pv_tv = tv / ((1 + w) ** n) if n > 0 else tv
                row.append(round(pv_cfs + pv_tv, 2))
        pg_matrix_excl.append(row_excl)
        pg_matrix_incl.append(row_incl)

    return {
        'terminal_multiple': {
            'wacc_values': [round(w * 100, 2) for w in wacc_values],
            'exit_mult_values': exit_mult_values,
            'matrix_excl': tm_matrix_excl,
            'matrix_incl': tm_matrix_incl,
        },
        'perpetuity_growth': {
            'wacc_values': [round(w * 100, 2) for w in wacc_values],
            'tgr_values': [round(g * 100, 2) for g in tgr_values],
            'matrix_excl': pg_matrix_excl,
            'matrix_incl': pg_matrix_incl,
        },
    }


def compute_unlevered_returns(year_totals: List[dict], config: ForecastConfig) -> dict:
    """Compute unlevered IRR and MOIC.

    Year 0: -purchase_price
    Years 1-N: UFCF = net_earnings_incl
    Exit year: + exit_year_NE × exit_multiple
    """
    if config.purchase_price <= 0:
        return {}

    ne_key = 'net_earnings_incl'
    purchase = config.purchase_price
    exit_mult = config.exit_multiple
    forecast_start = date.today()  # Will be overridden by caller context

    schedule = []
    total_cf = 0.0

    for yt in year_totals:
        ufcf = yt.get(ne_key, 0)
        is_exit = (yt['year'] == config.horizon_years)
        exit_proceeds = yt.get(ne_key, 0) * exit_mult if is_exit else 0
        total_flow = ufcf + exit_proceeds

        schedule.append({
            'year': yt['year'],
            'calendar_year': yt.get('calendar_year', 0),
            'ufcf': round(ufcf, 2),
            'exit_proceeds': round(exit_proceeds, 2),
            'total': round(total_flow, 2),
        })
        total_cf += total_flow

    # MOIC
    moic = total_cf / purchase if purchase > 0 else 0

    # IRR via XIRR
    xirr_flows = [(date(2026, 1, 1), -purchase)]  # Placeholder start date
    for s in schedule:
        cy = s.get('calendar_year', 2027)
        xirr_flows.append((date(cy, 7, 1), s['total']))

    irr = _xirr(xirr_flows)

    exit_ev = year_totals[-1].get(ne_key, 0) * exit_mult if year_totals else 0

    return {
        'irr': round(irr, 4) if irr is not None else None,
        'moic': round(moic, 2),
        'purchase_price': round(purchase, 2),
        'exit_ev': round(exit_ev, 2),
        'total_cash_flows': round(total_cf, 2),
        'schedule': schedule,
    }


def compute_levered_returns(year_totals: List[dict], config: ForecastConfig) -> dict:
    """Compute levered returns with debt schedule.

    Debt = purchase_price × LTV, Equity = purchase_price - debt
    Interest = avg(opening, closing) × rate  (iterative solve)
    Principal = min(balance, (UFCF - interest) × sweep)
    LFCF = UFCF - interest - principal
    """
    if config.purchase_price <= 0:
        return {}

    ne_key = 'net_earnings_incl'
    purchase = config.purchase_price
    debt_initial = purchase * config.ltv
    equity = purchase - debt_initial
    interest_rate = max(config.sofr_rate, config.sofr_floor) + config.sofr_spread
    sweep = config.cash_flow_sweep

    if equity <= 0:
        return {}

    balance = debt_initial
    debt_schedule = []
    total_lfcf = 0.0

    for yt in year_totals:
        ufcf = yt.get(ne_key, 0)
        opening = balance

        # Iterative interest calc: interest = avg(opening, closing) × rate
        # closing = opening - principal, principal = min(opening, (ufcf - interest) * sweep)
        # Solve iteratively (converges fast)
        est_interest = opening * interest_rate
        for _ in range(10):
            available = max(0, ufcf - est_interest) * sweep
            principal = min(opening, available)
            closing = opening - principal
            new_interest = ((opening + closing) / 2) * interest_rate
            if abs(new_interest - est_interest) < 0.01:
                est_interest = new_interest
                break
            est_interest = new_interest

        interest = est_interest
        available = max(0, ufcf - interest) * sweep
        principal = min(opening, available)
        closing = opening - principal
        lfcf = ufcf - interest - principal

        debt_schedule.append({
            'year': yt['year'],
            'calendar_year': yt.get('calendar_year', 0),
            'ufcf': round(ufcf, 2),
            'opening_balance': round(opening, 2),
            'interest': round(interest, 2),
            'principal': round(principal, 2),
            'closing_balance': round(closing, 2),
            'lfcf': round(lfcf, 2),
        })

        total_lfcf += lfcf
        balance = closing

    # Exit
    last_ne = year_totals[-1].get(ne_key, 0) if year_totals else 0
    exit_ev = last_ne * config.exit_multiple
    remaining_debt = balance
    exit_equity = exit_ev - remaining_debt

    # MOIC
    total_distributions = total_lfcf + exit_equity
    moic = total_distributions / equity if equity > 0 else 0

    # IRR via XIRR
    xirr_flows = [(date(2026, 1, 1), -equity)]
    for i, ds in enumerate(debt_schedule):
        cy = ds.get('calendar_year', 2027)
        cf = ds['lfcf']
        if i == len(debt_schedule) - 1:
            cf += exit_equity
        xirr_flows.append((date(cy, 7, 1), cf))

    irr = _xirr(xirr_flows)

    return {
        'irr': round(irr, 4) if irr is not None else None,
        'moic': round(moic, 2),
        'equity': round(equity, 2),
        'debt_initial': round(debt_initial, 2),
        'interest_rate': round(interest_rate, 4),
        'exit_ev': round(exit_ev, 2),
        'remaining_debt': round(remaining_debt, 2),
        'exit_equity': round(exit_equity, 2),
        'total_lfcf': round(total_lfcf, 2),
        'debt_schedule': debt_schedule,
    }


# ---------------------------------------------------------------------------
# Monthly Projection
# ---------------------------------------------------------------------------

def project_isrc_monthly(baseline_annual_gross: float,
                          release_date_str: Optional[str],
                          config: ForecastConfig,
                          curve_name: str,
                          fee_ratio: float,
                          label_share: float,
                          tp_share: float,
                          forecast_start: date,
                          seasonality: Optional[Dict[int, float]] = None,
                          payor_code: Optional[str] = None) -> List[dict]:
    """Project per-ISRC at monthly granularity.

    Distributes annual projected gross across months using LTM seasonality pattern
    (if provided) or uniform 1/12 split. Returns list of monthly dicts.
    """
    curve = resolve_curve(curve_name)
    horizon = config.horizon_years

    # Parse release date
    release_date = None
    if release_date_str:
        try:
            if len(release_date_str) >= 10:
                release_date = date.fromisoformat(release_date_str[:10])
            elif len(release_date_str) == 4:
                release_date = date(int(release_date_str), 7, 1)
        except (ValueError, TypeError):
            pass

    base_age = ((forecast_start - release_date).days / 365.25) if release_date else 3.0

    # Default seasonality: uniform
    if not seasonality:
        seasonality = {m: 1.0 / 12 for m in range(1, 13)}
    else:
        total_s = sum(seasonality.values())
        if total_s > 0:
            seasonality = {m: v / total_s for m, v in seasonality.items()}

    monthly_results = []
    prev_annual_gross = baseline_annual_gross

    for y in range(1, horizon + 1):
        catalog_age = base_age + y
        decay_rate = interpolate_decay_rate(catalog_age, curve)
        year_gross = prev_annual_gross * (1 + decay_rate)

        for m in range(1, 13):
            month_gross = year_gross * seasonality.get(m, 1.0 / 12)
            month_fees = month_gross * fee_ratio
            month_nr = month_gross - month_fees
            month_tp = month_nr * tp_share
            month_ne_excl = month_nr - month_tp

            ramp, eff_new_fee, eff_tp_syn = compute_synergy_ramp_for_payor(y, config, payor_code)
            fee_savings = 0.0
            if eff_new_fee is not None and ramp > 0:
                blended = fee_ratio * (1 - ramp) + eff_new_fee * ramp
                fee_savings = (fee_ratio - blended) * month_gross

            tp_savings = 0.0
            if eff_tp_syn > 0 and ramp > 0:
                tp_savings = month_tp * eff_tp_syn * ramp

            month_ne_incl = month_ne_excl + fee_savings + tp_savings

            month_date = date(forecast_start.year + y,
                              min(m, 12), 1)

            monthly_results.append({
                'month': month_date.strftime('%Y-%m'),
                'year': y,
                'gross': round(month_gross, 2),
                'fees': round(month_fees, 2),
                'net_receipts': round(month_nr, 2),
                'third_party': round(month_tp, 2),
                'net_earnings_excl': round(month_ne_excl, 2),
                'fee_savings': round(fee_savings, 2),
                'tp_savings': round(tp_savings, 2),
                'net_earnings_incl': round(month_ne_incl, 2),
            })

        prev_annual_gross = year_gross if year_gross > 0 else prev_annual_gross * 0.5

    return monthly_results


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

def _aggregate_projections(isrc_projections: dict, horizon: int,
                           forecast_start_year: int) -> List[dict]:
    """Sum both excl/incl fields across all ISRCs per year."""
    agg_fields = [
        'gross', 'fees_original', 'net_receipts_excl', 'artist_share_excl',
        'third_party_excl', 'net_earnings_excl', 'fee_savings', 'tp_savings',
        'net_earnings_incl',
    ]

    year_totals = []
    for y in range(1, horizon + 1):
        yt = {'year': y, 'calendar_year': forecast_start_year + y}
        for fld in agg_fields:
            yt[fld] = 0.0

        for ip in isrc_projections.values():
            projs = ip['projections']
            if y - 1 < len(projs):
                p = projs[y - 1]
                for fld in agg_fields:
                    yt[fld] += p.get(fld, 0)

        # Round all
        for fld in agg_fields:
            yt[fld] = round(yt[fld], 2)

        year_totals.append(yt)

    return year_totals


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def run_forecast(payor_results: dict, analytics: dict,
                 config: ForecastConfig) -> dict:
    """Orchestrator: run full forecast projection with per-payor and per-ISRC granularity.

    Returns {
        config, forecast_start, isrc_count,
        aggregate: {year_totals, total_gross, total_net_excl, total_net_incl},
        ntm: {gross, net_excl, net_incl},
        per_payor: {code: {income_rights, year_totals, isrc_projections, wa_release_date}},
        by_income_type: {type: {year_totals}},
        dcf, sensitivity, irr_sensitivity,
        unlevered_returns, levered_returns, virtu_levered_returns,
        cohorts, wa_release_dates, isrc_projections, isrc_metadata,
        summary, top_isrcs,
    }
    """
    import pandas as pd

    # Build LTM monthly data
    all_monthly = []
    payor_code_map = {}  # code -> payor name
    for code, pr in payor_results.items():
        m = pr.monthly.copy()
        m['payor'] = pr.config.name
        m['payor_code'] = code
        all_monthly.append(m)
        payor_code_map[code] = pr.config.name
    monthly = pd.concat(all_monthly, ignore_index=True)

    periods = sorted(monthly['period'].unique().tolist())
    if not periods:
        return _empty_forecast(config)

    # --- Per-payor LTM windows ---
    # Each payor may have a different most-recent statement period.
    # Compute trailing-12-month window per payor, then stitch together.
    payor_max_periods = {}   # {code: int}  e.g. {'DITTO': 202502, 'PRS': 202512}
    payor_ltm_starts = {}    # {code: int}  trailing 12-month start per payor
    for code in payor_code_map:
        payor_periods = monthly[monthly['payor_code'] == code]['period']
        if payor_periods.empty:
            continue
        p_max = int(payor_periods.max())
        payor_max_periods[code] = p_max
        p_yr = p_max // 100
        p_mn = p_max % 100
        if p_mn == 12:
            payor_ltm_starts[code] = p_yr * 100 + 1
        else:
            payor_ltm_starts[code] = (p_yr - 1) * 100 + (p_mn + 1)

    # Global max period (for display, forecast_start, seasonality)
    max_period = max(periods)
    max_year = int(str(max_period)[:4])
    max_month = int(str(max_period)[4:6])
    if max_month == 12:
        ltm_start_period = max_year * 100 + 1
    else:
        ltm_start_period = (max_year - 1) * 100 + (max_month + 1)

    # Build ltm_monthly from per-payor windows: each payor's data filtered to
    # its own trailing 12 months.  This prevents stale payors from having their
    # ISRC baselines computed from a partial (non-trailing-12) window.
    ltm_chunks = []
    for code, ltm_s in payor_ltm_starts.items():
        chunk = monthly[(monthly['payor_code'] == code) & (monthly['period'] >= ltm_s)]
        ltm_chunks.append(chunk)
    ltm_monthly = pd.concat(ltm_chunks, ignore_index=True) if ltm_chunks else monthly[monthly['period'] >= ltm_start_period]

    # Stale payor warnings: payors >6 months behind the newest
    payor_ltm_warnings = []
    if payor_max_periods:
        global_max = max(payor_max_periods.values())
        for code, p_max in payor_max_periods.items():
            # Compute month gap
            gap_months = (global_max // 100 - p_max // 100) * 12 + (global_max % 100 - p_max % 100)
            if gap_months > 6:
                payor_ltm_warnings.append({
                    'code': code,
                    'name': payor_code_map.get(code, code),
                    'max_period': p_max,
                    'gap_months': gap_months,
                })
    if payor_ltm_warnings:
        log.warning(f"Stale payors detected: {[(w['name'], w['gap_months']) for w in payor_ltm_warnings]}")

    # Get all unique ISRCs with LTM revenue
    ltm_by_isrc = ltm_monthly.groupby('identifier')['gross'].sum()
    active_isrcs = ltm_by_isrc[ltm_by_isrc > 0].index.tolist()

    if not active_isrcs:
        return _empty_forecast(config)

    # Per-payor ISRC breakdown for per-payor models
    ltm_by_payor_isrc = ltm_monthly.groupby(['payor_code', 'identifier'])['gross'].sum()

    # LTM seasonality pattern — use GLOBAL window for seasonality so that
    # seasonal weights reflect the newest data's month distribution, not
    # stale payors' partial windows.
    ltm_monthly_global = monthly[monthly['period'] >= ltm_start_period]
    if 'period' in ltm_monthly_global.columns:
        ltm_monthly_sums = ltm_monthly_global.copy()
        ltm_monthly_sums['month_num'] = ltm_monthly_sums['period'].astype(str).str[4:6].astype(int)
        month_totals = ltm_monthly_sums.groupby('month_num')['gross'].sum()
        total_for_seasonality = month_totals.sum()
        if total_for_seasonality > 0:
            seasonality = {int(m): float(v / total_for_seasonality) for m, v in month_totals.items()}
        else:
            seasonality = None
    else:
        seasonality = None

    # ISRC coverage
    isrc_gross_total = float(ltm_by_isrc[ltm_by_isrc > 0].sum())
    analytics_ltm_gross = analytics.get('ltm_gross_total', 0)
    if analytics_ltm_gross > 0 and isrc_gross_total > 0:
        coverage_ratio = isrc_gross_total / analytics_ltm_gross
    else:
        coverage_ratio = 1.0

    # Gather release dates + metadata
    release_dates = _gather_release_dates(payor_results)
    meta_lookup = _gather_metadata(payor_results)

    # Infer release dates from first appearance in historical data for ISRCs without dates.
    # Each ISRC gets an age based on when it first appeared, producing varied per-ISRC decay.
    # BUT: if a song first appears in the payor's earliest period, it was already released
    # before data collection began, so the inferred date is unreliable — use default age.
    first_appearance = monthly.groupby('identifier')['period'].min()

    # Find each payor's earliest data period
    payor_earliest = monthly.groupby('payor_code')['period'].min().to_dict()
    # ISRC-to-payor mapping for earliest period check
    isrc_primary_payor = {}
    for (pcode, isrc_id), g in ltm_by_payor_isrc.items():
        if g > 0:
            if isrc_id not in isrc_primary_payor or g > ltm_by_payor_isrc.get(
                    (isrc_primary_payor[isrc_id], isrc_id), 0):
                isrc_primary_payor[isrc_id] = pcode

    inferred_count = 0
    for isrc_id, first_period in first_appearance.items():
        if isrc_id not in release_dates:
            fp = int(first_period)
            yr = fp // 100
            mn = fp % 100
            if not (1 <= mn <= 12 and yr >= 2000):
                continue

            # Check if this ISRC first appeared at the payor's data start
            # If so, the song predates the data — don't use this as release date
            primary_payor = isrc_primary_payor.get(isrc_id)
            if primary_payor:
                payor_start = int(payor_earliest.get(primary_payor, 0))
                # If first appearance is within 2 months of payor's earliest data,
                # the song likely predates the data collection period
                if fp <= payor_start + 1:  # +1 for 1-month tolerance
                    continue

            release_dates[isrc_id] = f"{yr}-{mn:02d}-01"
            inferred_count += 1
    if inferred_count > 0:
        log.info(f"Inferred release dates from first appearance for {inferred_count} ISRCs")

    forecast_start = date(max_year, max_month, 28)

    # --- Trajectory-based age estimation for ISRCs without release dates ---
    # Build per-ISRC annual earnings from historical data for age estimation.
    # ISRCs that predated the data (skipped by first-appearance inference) are
    # typically the OLDEST songs.  Using their observed YoY earnings pattern to
    # estimate where they sit on the decay curve gives much more accurate ages
    # than a uniform default.
    #
    # Track which ISRCs got estimated ages (trajectory or WAA default) so we
    # can suppress terminal growth for them — terminal growth models aggregate
    # catalog behavior, not individual tracks with uncertain ages.
    estimated_age_isrcs = set()
    default_curve = resolve_curve(config.genre_default)
    trajectory_count = 0
    undated_isrcs = [i for i in active_isrcs if i not in release_dates]

    if undated_isrcs:
        # Annual earnings per ISRC (calendar year -> gross)
        monthly_copy = monthly.copy()
        monthly_copy['cal_year'] = monthly_copy['period'].astype(str).str[:4].astype(int)
        isrc_yearly = monthly_copy.groupby(['identifier', 'cal_year'])['gross'].sum()

        for isrc_id in undated_isrcs:
            try:
                isrc_years = isrc_yearly.loc[isrc_id]
            except KeyError:
                continue

            yearly_dict = {int(yr): float(g) for yr, g in isrc_years.items() if g > 0}
            if len(yearly_dict) < 2:
                continue

            # Select the appropriate decay curve for this ISRC
            curve_key = config.isrc_decay_curves.get(isrc_id, config.genre_default)
            curve_for_isrc = resolve_curve(curve_key)

            est_age = estimate_age_from_trajectory(yearly_dict, curve_for_isrc, max_year)
            if est_age is not None:
                pseudo_date = date(
                    max(forecast_start.year - int(est_age), 1990),
                    max(1, min(12, int((est_age % 1) * 12) + 1)),
                    1,
                )
                release_dates[isrc_id] = pseudo_date.isoformat()
                estimated_age_isrcs.add(isrc_id)
                trajectory_count += 1

    if trajectory_count > 0:
        log.info(f"Estimated age from YoY trajectory for {trajectory_count} ISRCs "
                 f"(out of {len(undated_isrcs)} undated)")

    # Weighted average age fallback — compute smarter default for remaining undated ISRCs
    waa = analytics.get('weighted_avg_age', {})
    analytics_waa = waa.get('waa_years', 0) if isinstance(waa, dict) else 0

    known_age_weighted = 0.0
    known_gross = 0.0
    unknown_gross = 0.0
    for isrc in active_isrcs:
        ltm_g = float(ltm_by_isrc.get(isrc, 0))
        if isrc in release_dates:
            rd_str = release_dates[isrc]
            try:
                rd = date.fromisoformat(rd_str[:10])
                age = (forecast_start - rd).days / 365.25
                known_age_weighted += age * ltm_g
                known_gross += ltm_g
            except (ValueError, TypeError):
                unknown_gross += ltm_g
        else:
            unknown_gross += ltm_g

    total_gross_for_age = known_gross + unknown_gross

    # Determine global WAA for back-solve.
    # If analytics provides a WAA, use it.  Otherwise, compute an empirical one:
    # undated ISRCs predate the data, so they're at least data_span old.
    # Assume their average age ≈ data_span + 2 (released ~2 years before data started).
    if analytics_waa > 0:
        global_waa = analytics_waa
    elif unknown_gross > 0 and known_gross > 0 and total_gross_for_age > 0:
        # Compute data span (years from earliest to latest period)
        earliest_period = int(monthly['period'].min())
        latest_period = int(monthly['period'].max())
        data_span = (latest_period // 100 - earliest_period // 100) + \
                     (latest_period % 100 - earliest_period % 100) / 12.0
        # Undated ISRCs predate the data — assume avg release ~3yr before data start.
        # (Most flat-trajectory ISRCs are now handled by trajectory estimation;
        # the remaining ones here are viral/growing songs, which also tend to be
        # older catalog that got a second-life boost.)
        assumed_undated_age = data_span + 3.0
        known_avg_age = known_age_weighted / known_gross if known_gross > 0 else 0
        global_waa = (known_avg_age * known_gross +
                      assumed_undated_age * unknown_gross) / total_gross_for_age
        log.info(f"Computed empirical WAA={global_waa:.1f}yr (known avg={known_avg_age:.1f}yr, "
                 f"assumed undated={assumed_undated_age:.1f}yr, data_span={data_span:.1f}yr)")
    else:
        global_waa = 3.0

    if unknown_gross > 0 and total_gross_for_age > 0 and global_waa > 0:
        implied_unknown_age = (global_waa * total_gross_for_age - known_age_weighted) / unknown_gross
        default_age_years = max(implied_unknown_age, 1.0)
        log.info(f"WAA-backsolved default age for remaining "
                 f"{sum(1 for i in active_isrcs if i not in release_dates)} ISRCs: "
                 f"{default_age_years:.1f}yr (global WAA={global_waa:.1f}yr)")
    else:
        default_age_years = global_waa

    # Extract waterfall ratios from analytics (used as ultimate fallback)
    wf = analytics.get('waterfall', {}).get('overall', {})
    if wf and wf.get('gross', 0) > 0:
        wf_gross = wf['gross']
        wf_fee_ratio = abs(wf.get('fees', 0)) / wf_gross
        wf_net_receipts = wf.get('net_receipts', wf_gross - abs(wf.get('fees', 0)))
        wf_tp = wf.get('third_party', 0)
        wf_tp_ratio = abs(wf_tp) / wf_net_receipts if wf_net_receipts > 0 else 0
    else:
        wf_fee_ratio = 0.15
        wf_tp_ratio = 0.0

    # --- Build per-payor fee rates and artist shares ---
    # Priority: explicit payor_configs > analytics payor_summaries > waterfall fallback
    payor_fee_rates = {}     # {code: float 0-1}
    payor_artist_shares = {} # {code: float 0-1}
    payor_tp_ratios = {}     # {code: float 0-1}
    analytics_payors_raw = analytics.get('payor_summaries', {})
    # Convert list to dict keyed by 'code' if needed
    if isinstance(analytics_payors_raw, list):
        analytics_payors = {p.get('code', p.get('name', '')): p for p in analytics_payors_raw}
    else:
        analytics_payors = analytics_payors_raw

    for code in payor_results.keys():
        pc = config.payor_configs.get(code, {})

        # Fee rate
        if 'fee_rate' in pc:
            payor_fee_rates[code] = float(pc['fee_rate'])
        else:
            # Auto-detect from analytics payor_summaries
            ps = analytics_payors.get(code, {})
            fee_str = ps.get('fee', ps.get('fee_pct', ''))
            if isinstance(fee_str, str) and '%' in fee_str:
                try:
                    payor_fee_rates[code] = float(fee_str.replace('%', '').strip()) / 100.0
                except ValueError:
                    payor_fee_rates[code] = wf_fee_ratio
            elif isinstance(fee_str, (int, float)):
                val = float(fee_str)
                payor_fee_rates[code] = val / 100.0 if val > 1 else val
            else:
                payor_fee_rates[code] = wf_fee_ratio

        # Artist share (from explicit config only — not in analytics)
        payor_artist_shares[code] = float(pc.get('artist_share', 0.0))

        # Third-party ratio (from explicit config, else waterfall fallback)
        if 'tp_ratio' in pc:
            payor_tp_ratios[code] = float(pc['tp_ratio'])
        else:
            payor_tp_ratios[code] = wf_tp_ratio

    log.info(f"Per-payor fee rates: {payor_fee_rates}")
    log.info(f"Per-payor artist shares: {payor_artist_shares}")

    # Map ISRCs to payors (for per-payor model)
    isrc_to_payors = {}
    for (pcode, isrc_id), g in ltm_by_payor_isrc.items():
        if g > 0:
            isrc_to_payors.setdefault(isrc_id, []).append(pcode)

    # Run per-ISRC projections with per-ISRC overrides
    isrc_projections = {}
    for isrc in active_isrcs:
        ltm_gross_val = round(float(ltm_by_isrc.get(isrc, 0)), 2)

        # Apply FX conversion if payor has fx config
        fx_multiplier = 1.0
        fx_currency = None
        isrc_payors = isrc_to_payors.get(isrc, [])
        for pcode in isrc_payors:
            pc = config.payor_configs.get(pcode, {})
            fx_ccy = pc.get('fx_currency')
            if fx_ccy and fx_ccy != config.base_currency:
                fx_rate = pc.get('fx_rate') or config.fx_rates.get(fx_ccy, 1.0)
                fx_multiplier = fx_rate
                fx_currency = fx_ccy
                break

        baseline = {
            'isrc': isrc,
            'gross': round(ltm_gross_val * fx_multiplier, 2),
            'gross_original': ltm_gross_val,
            'fx_currency': fx_currency,
            'fx_multiplier': fx_multiplier,
            'fees': 0,
            'net_receipts': 0,
            'net_earnings': 0,
            'release_date': release_dates.get(isrc),
        }

        # Per-ISRC fee override > per-payor fee > waterfall fallback
        isrc_fee = config.isrc_fee_overrides.get(isrc)
        if isrc_fee is None:
            for pcode in isrc_payors:
                if pcode in payor_fee_rates:
                    isrc_fee = payor_fee_rates[pcode]
                    break
            if isrc_fee is None:
                isrc_fee = wf_fee_ratio

        isrc_label = config.isrc_label_shares.get(isrc)
        isrc_tp = config.isrc_3p_shares.get(isrc)

        # Artist share: per-ISRC (not yet supported) > per-payor > 0
        isrc_artist_share = 0.0
        for pcode in isrc_payors:
            if pcode in payor_artist_shares and payor_artist_shares[pcode] > 0:
                isrc_artist_share = payor_artist_shares[pcode]
                break

        # Third-party ratio: per-ISRC override > per-payor > waterfall fallback
        if isrc_tp is not None:
            effective_tp = isrc_tp
        else:
            effective_tp = wf_tp_ratio
            for pcode in isrc_payors:
                if pcode in payor_tp_ratios:
                    effective_tp = payor_tp_ratios[pcode]
                    break

        effective_fee = isrc_fee
        if isrc_label is not None:
            effective_fee = 1.0 - isrc_label

        baseline['fees'] = round(baseline['gross'] * effective_fee, 2)
        baseline['net_receipts'] = round(baseline['gross'] * (1 - effective_fee), 2)
        baseline['artist_share'] = round(baseline['net_receipts'] * isrc_artist_share, 2)
        after_artist = baseline['net_receipts'] - baseline['artist_share']
        baseline['net_earnings'] = round(after_artist * (1 - effective_tp), 2)

        # Decay curve: per-ISRC override > genre_overrides > genre_default
        curve_name = config.isrc_decay_curves.get(isrc) or \
                     config.genre_overrides.get(isrc) or \
                     config.genre_default

        rd_str = release_dates.get(isrc)
        # Terminal blend: 1.0 for known/inferred dates, 0.5 for WAA-default
        # (uncertain age → blend terminal growth with flat rate).
        # Trajectory-estimated ISRCs keep 1.0 — their age match is reasonable.
        t_blend = 1.0
        if not rd_str:
            pseudo_date = date(
                forecast_start.year - int(default_age_years),
                7, 1
            )
            rd_str = pseudo_date.isoformat()
            t_blend = 0.5  # WAA-backsolved default age — uncertain

        projections = project_isrc(
            baseline, rd_str, config, genre=curve_name,
            forecast_start_date=forecast_start,
            fee_ratio=effective_fee, tp_ratio=effective_tp,
            label_share=isrc_label, tp_share=isrc_tp,
            curve_name=curve_name,
            artist_share=isrc_artist_share,
            terminal_blend=t_blend,
            payor_code=isrc_primary_payor.get(isrc),
        )

        isrc_projections[isrc] = {
            'baseline': baseline,
            'projections': projections,
            'genre': curve_name,
            'release_date': release_dates.get(isrc),
            'title': meta_lookup.get(isrc, {}).get('title', isrc),
            'artist': meta_lookup.get(isrc, {}).get('artist', ''),
            'payors': isrc_payors,
            'fee_ratio': effective_fee,
            'tp_ratio': effective_tp,
            'artist_share': isrc_artist_share,
            'label_share': isrc_label,
        }

    # Aggregate projections (both tracks)
    year_totals = _aggregate_projections(isrc_projections, config.horizon_years,
                                         forecast_start.year)

    total_gross = sum(yt['gross'] for yt in year_totals)
    total_net_excl = sum(yt['net_earnings_excl'] for yt in year_totals)
    total_net_incl = sum(yt['net_earnings_incl'] for yt in year_totals)

    # NTM = Year 1 forecast
    ntm = {
        'gross': year_totals[0]['gross'] if year_totals else 0,
        'net_excl': year_totals[0].get('net_earnings_excl', 0) if year_totals else 0,
        'net_incl': year_totals[0].get('net_earnings_incl', 0) if year_totals else 0,
    }

    # --- Per-payor aggregation ---
    per_payor = {}
    for code in payor_results.keys():
        pc = config.payor_configs.get(code, {})
        income_rights = pc.get('income_rights', config.rights_included)

        # Find ISRCs belonging to this payor
        payor_isrcs = [isrc for isrc, payors in isrc_to_payors.items()
                       if code in payors and isrc in isrc_projections]

        pp_agg_fields = ['gross', 'fees_original', 'net_receipts_excl', 'artist_share_excl',
                         'third_party_excl', 'net_earnings_excl', 'fee_savings', 'tp_savings',
                         'net_earnings_incl']

        payor_year_totals = []
        for y in range(1, config.horizon_years + 1):
            yt = {'year': y, 'calendar_year': forecast_start.year + y}
            for fld in pp_agg_fields:
                yt[fld] = 0.0
            for isrc in payor_isrcs:
                projs = isrc_projections[isrc]['projections']
                if y - 1 < len(projs):
                    p = projs[y - 1]
                    for fld in pp_agg_fields:
                        yt[fld] += p.get(fld, 0)
            for fld in pp_agg_fields:
                yt[fld] = round(yt[fld], 2)
            payor_year_totals.append(yt)

        # Payor LTM gross
        payor_ltm_gross = sum(
            isrc_projections[isrc]['baseline']['gross'] for isrc in payor_isrcs
        )

        per_payor[code] = {
            'name': payor_code_map.get(code, code),
            'income_rights': income_rights,
            'year_totals': payor_year_totals,
            'isrc_count': len(payor_isrcs),
            'ltm_gross': round(payor_ltm_gross, 2),
            'isrcs': payor_isrcs,
            'ltm_start': payor_ltm_starts.get(code),
            'ltm_end': payor_max_periods.get(code),
        }

    # --- By income type aggregation ---
    income_types = {}
    for code, pp_data in per_payor.items():
        it = pp_data['income_rights']
        if it not in income_types:
            income_types[it] = []
        income_types[it].append(code)

    by_income_type = {}
    for it, codes in income_types.items():
        it_year_totals = []
        for y in range(1, config.horizon_years + 1):
            yt = {'year': y, 'calendar_year': forecast_start.year + y}
            for fld in ['gross', 'fees_original', 'net_receipts_excl', 'artist_share_excl',
                        'third_party_excl', 'net_earnings_excl', 'fee_savings', 'tp_savings',
                        'net_earnings_incl']:
                yt[fld] = sum(
                    per_payor[c]['year_totals'][y - 1].get(fld, 0)
                    for c in codes if y - 1 < len(per_payor[c]['year_totals'])
                )
                yt[fld] = round(yt[fld], 2)
            it_year_totals.append(yt)
        by_income_type[it] = {'year_totals': it_year_totals}

    # LTM reference values
    ltm_net = analytics.get('ltm_net_total', 0)
    ltm_gross = analytics.get('ltm_gross_total', 0)

    # DCF Analysis
    dcf = compute_dcf_analysis(year_totals, config, ltm_net=ltm_net)

    # Sensitivity Tables (3x3)
    sensitivity = compute_sensitivity_tables(year_totals, config, ltm_net=ltm_net)

    # Returns (only when purchase_price > 0)
    unlevered_returns = {}
    levered_returns = {}
    virtu_levered_returns = None
    irr_sensitivity = {}
    levered_sensitivity = {}
    if config.purchase_price > 0:
        unlevered_returns = _compute_unlevered_returns_dated(year_totals, config, forecast_start)
        levered_returns = _compute_levered_returns_dated(year_totals, config, forecast_start)
        virtu_levered_returns = compute_virtu_levered_returns(year_totals, config, forecast_start)
        irr_sensitivity = compute_irr_moic_sensitivity(year_totals, config, forecast_start)
        levered_sensitivity = compute_levered_irr_moic_sensitivity(year_totals, config, forecast_start)

    # Cohort analysis
    cohorts = compute_forecast_cohorts(isrc_projections, meta_lookup, {}, config, forecast_start)

    # WA release date per payor
    wa_release_dates = compute_wa_release_date_per_payor(payor_results, ltm_start_period)

    # Backward-compatible summary
    tgr = config.terminal_growth if config.terminal_growth is not None else 0.01
    primary_npv = dcf.get('terminal_multiple', {}).get('incl', {}).get('implied_valuation', 0)
    perp_npv = dcf.get('perpetuity_growth', {}).get('incl', {}).get('implied_valuation', 0)

    summary = {
        'terminal_value': dcf.get('terminal_multiple', {}).get('incl', {}).get('terminal_value', 0),
        'terminal_growth': tgr,
        'discount_rate': config.discount_rate,
        'npv': primary_npv,
        'npv_perpetuity': perp_npv,
        'implied_multiple_net': round(primary_npv / ltm_net, 2) if ltm_net > 0 else 0,
        'implied_multiple_gross': round(primary_npv / ltm_gross, 2) if ltm_gross > 0 else 0,
        'ltm_gross': round(ltm_gross, 2),
        'ltm_net': round(ltm_net, 2),
        'ntm_gross': ntm['gross'],
        'ntm_net_excl': ntm['net_excl'],
        'ntm_net_incl': ntm['net_incl'],
        'closing_amount': config.closing_amount,
    }

    # Top ISRCs by projected total
    top_isrcs = []
    for isrc, ip in isrc_projections.items():
        proj_total = sum(p['gross'] for p in ip['projections'])
        ltm_g = ip['baseline']['gross']
        top_isrcs.append({
            'isrc': isrc,
            'title': ip.get('title', isrc),
            'artist': ip.get('artist', ''),
            'ltm_gross': ltm_g,
            'projected_total': round(proj_total, 2),
            'genre': ip.get('genre', config.genre_default),
            'release_date': ip.get('release_date', ''),
            'pct_ltm': round(ltm_g / isrc_gross_total * 100, 2) if isrc_gross_total > 0 else 0,
            'fee_ratio': ip.get('fee_ratio', wf_fee_ratio),
            'tp_ratio': ip.get('tp_ratio', wf_tp_ratio),
        })
    top_isrcs.sort(key=lambda x: x['projected_total'], reverse=True)

    # LTM waterfall row — computed from per-ISRC baselines (which use per-payor rates)
    ltm_isrc_gross = round(isrc_gross_total, 2)
    ltm_fees = round(sum(ip['baseline']['fees'] for ip in isrc_projections.values()), 2)
    ltm_nr = round(sum(ip['baseline']['net_receipts'] for ip in isrc_projections.values()), 2)
    ltm_artist = round(sum(ip['baseline'].get('artist_share', 0) for ip in isrc_projections.values()), 2)
    ltm_tp = round(sum(
        (ip['baseline']['net_receipts'] - ip['baseline'].get('artist_share', 0)) * ip.get('tp_ratio', 0)
        for ip in isrc_projections.values()
    ), 2)
    ltm_ne = round(sum(ip['baseline']['net_earnings'] for ip in isrc_projections.values()), 2)

    # Serialize isrc_projections for export (include per-year gross for song listing)
    isrc_proj_export = {}
    for isrc, ip in isrc_projections.items():
        isrc_proj_export[isrc] = {
            'baseline': ip['baseline'],
            'projections': ip['projections'],
            'genre': ip.get('genre', ''),
            'release_date': ip.get('release_date', ''),
            'title': ip.get('title', isrc),
            'artist': ip.get('artist', ''),
            'payors': ip.get('payors', []),
            'fee_ratio': ip.get('fee_ratio', wf_fee_ratio),
            'tp_ratio': ip.get('tp_ratio', wf_tp_ratio),
            'artist_share': ip.get('artist_share', 0.0),
            'label_share': ip.get('label_share'),
        }

    result = {
        'config': config.to_dict(),
        'forecast_start': forecast_start.isoformat(),
        'isrc_count': len(isrc_projections),
        'isrc_coverage': round(coverage_ratio, 4),
        'payor_ltm_warnings': payor_ltm_warnings,
        'aggregate': {
            'year_totals': year_totals,
            'total_gross': round(total_gross, 2),
            'total_net_excl': round(total_net_excl, 2),
            'total_net_incl': round(total_net_incl, 2),
            'total_net': round(total_net_incl, 2),
        },
        'ntm': ntm,
        'ltm_waterfall': {
            'gross': ltm_isrc_gross,
            'gross_all': round(ltm_gross, 2),
            'fees': ltm_fees,
            'net_receipts': ltm_nr,
            'artist_share': ltm_artist,
            'third_party': ltm_tp,
            'net_earnings': ltm_ne,
            'coverage_pct': round(coverage_ratio * 100, 1),
        },
        'per_payor': per_payor,
        'by_income_type': by_income_type,
        'dcf': dcf,
        'sensitivity': sensitivity,
        'irr_sensitivity': irr_sensitivity,
        'levered_sensitivity': levered_sensitivity,
        'unlevered_returns': unlevered_returns,
        'levered_returns': levered_returns,
        'virtu_levered_returns': virtu_levered_returns,
        'cohorts': cohorts,
        'wa_release_dates': wa_release_dates,
        'summary': summary,
        'top_isrcs': top_isrcs[:50],
        'isrc_projections': isrc_proj_export,
        'isrc_metadata': meta_lookup,
    }

    return result


def _compute_unlevered_returns_dated(year_totals: List[dict], config: ForecastConfig,
                                     forecast_start: date) -> dict:
    """Compute unlevered returns with proper dates from forecast_start."""
    if config.purchase_price <= 0:
        return {}

    ne_key = 'net_earnings_incl'
    purchase = config.purchase_price
    exit_mult = config.exit_multiple

    # Use close_date as Y0 for XIRR (standard LBO convention: annual periods
    # from close date). Falls back to forecast_start if close_date not set.
    y0_date = forecast_start
    if config.close_date:
        try:
            y0_date = date.fromisoformat(config.close_date[:10])
        except (ValueError, TypeError):
            pass

    schedule = []
    total_cf = 0.0

    for yt in year_totals:
        ufcf = yt.get(ne_key, 0)
        is_exit = (yt['year'] == config.horizon_years)
        exit_proceeds = yt.get(ne_key, 0) * exit_mult if is_exit else 0
        total_flow = ufcf + exit_proceeds

        schedule.append({
            'year': yt['year'],
            'calendar_year': yt.get('calendar_year', 0),
            'ufcf': round(ufcf, 2),
            'exit_proceeds': round(exit_proceeds, 2),
            'total': round(total_flow, 2),
        })
        total_cf += total_flow

    moic = total_cf / purchase if purchase > 0 else 0

    # XIRR with annual periods from close date (standard LBO convention).
    # Cash flows arrive at anniversary of close date each year.
    xirr_flows = [(y0_date, -purchase)]
    for s in schedule:
        yr_offset = s['year']
        cf_date = date(y0_date.year + yr_offset, y0_date.month,
                       min(y0_date.day, 28))  # avoid Feb 29 issues
        xirr_flows.append((cf_date, s['total']))

    irr = _xirr(xirr_flows)
    exit_ev = year_totals[-1].get(ne_key, 0) * exit_mult if year_totals else 0

    return {
        'irr': round(irr, 4) if irr is not None else None,
        'moic': round(moic, 2),
        'purchase_price': round(purchase, 2),
        'exit_ev': round(exit_ev, 2),
        'total_cash_flows': round(total_cf, 2),
        'schedule': schedule,
    }


def _compute_levered_returns_dated(year_totals: List[dict], config: ForecastConfig,
                                   forecast_start: date,
                                   debt_override: Optional[float] = None,
                                   equity_override: Optional[float] = None) -> dict:
    """Compute levered returns with proper dates from forecast_start.

    If debt_override/equity_override are set, uses those instead of
    purchase_price * ltv (used for Virtu variant where debt is sized
    against Virtu DCF valuation, not purchase price).
    Interest rates use SOFR forward curve when available.
    """
    if config.purchase_price <= 0:
        return {}

    ne_key = 'net_earnings_incl'
    purchase = config.purchase_price
    debt_initial = debt_override if debt_override is not None else purchase * config.ltv
    equity = equity_override if equity_override is not None else purchase - debt_initial
    sweep = config.cash_flow_sweep

    if equity <= 0:
        return {}

    balance = debt_initial
    debt_schedule = []
    total_lfcf = 0.0
    year_interest_rates = []

    for yt in year_totals:
        ufcf = yt.get(ne_key, 0)
        opening = balance

        # Per-year interest rate from SOFR curve or flat
        interest_rate = get_interest_rate_for_year(yt['year'], config, forecast_start)
        year_interest_rates.append(interest_rate)

        est_interest = opening * interest_rate
        for _ in range(10):
            available = max(0, ufcf - est_interest) * sweep
            principal = min(opening, available)
            closing = opening - principal
            new_interest = ((opening + closing) / 2) * interest_rate
            if abs(new_interest - est_interest) < 0.01:
                est_interest = new_interest
                break
            est_interest = new_interest

        interest = est_interest
        available = max(0, ufcf - interest) * sweep
        principal = min(opening, available)
        closing = opening - principal
        lfcf = ufcf - interest - principal

        debt_schedule.append({
            'year': yt['year'],
            'calendar_year': yt.get('calendar_year', 0),
            'ufcf': round(ufcf, 2),
            'opening_balance': round(opening, 2),
            'interest_rate': round(interest_rate, 4),
            'interest': round(interest, 2),
            'principal': round(principal, 2),
            'closing_balance': round(closing, 2),
            'lfcf': round(lfcf, 2),
        })

        total_lfcf += lfcf
        balance = closing

    last_ne = year_totals[-1].get(ne_key, 0) if year_totals else 0

    # Exit enterprise value — always uses config.exit_multiple
    # (Virtu variant differs only in debt/equity split, not exit multiple)
    exit_ev = last_ne * config.exit_multiple

    remaining_debt = balance
    exit_equity = exit_ev - remaining_debt

    total_distributions = total_lfcf + exit_equity
    moic = total_distributions / equity if equity > 0 else 0

    # XIRR with annual periods from close date (same convention as unlevered)
    y0_date = forecast_start
    if config.close_date:
        try:
            y0_date = date.fromisoformat(config.close_date[:10])
        except (ValueError, TypeError):
            pass

    xirr_flows = [(y0_date, -equity)]
    for i, ds in enumerate(debt_schedule):
        yr_offset = ds['year']
        cf_date = date(y0_date.year + yr_offset, y0_date.month,
                       min(y0_date.day, 28))
        cf = ds['lfcf']
        if i == len(debt_schedule) - 1:
            cf += exit_equity
        xirr_flows.append((cf_date, cf))

    irr = _xirr(xirr_flows)

    # Representative interest rate (first year, for display)
    display_rate = year_interest_rates[0] if year_interest_rates else (
        max(config.sofr_rate, config.sofr_floor) + config.sofr_spread)

    return {
        'irr': round(irr, 4) if irr is not None else None,
        'moic': round(moic, 2),
        'equity': round(equity, 2),
        'debt_initial': round(debt_initial, 2),
        'interest_rate': round(display_rate, 4),
        'exit_ev': round(exit_ev, 2),
        'exit_multiple_used': round(exit_ev / last_ne, 1) if last_ne > 0 else 0,
        'remaining_debt': round(remaining_debt, 2),
        'exit_equity': round(exit_equity, 2),
        'total_lfcf': round(total_lfcf, 2),
        'debt_schedule': debt_schedule,
    }


def compute_virtu_levered_returns(year_totals: List[dict], config: ForecastConfig,
                                  forecast_start: date) -> Optional[dict]:
    """Compute levered returns using Virtu methodology.

    Key difference from CMG levered returns:
    - CMG: Debt = Purchase_Price × LTV
    - Virtu: Debt = Virtu_DCF_PG_Incl × LTV  (DCF at virtu_wacc > PP → more debt)
    - Equity = Purchase_Price - Virtu_Debt    (much less equity than CMG)
    - Same UFCF, same exit multiple, same SOFR-based interest
    - Less equity + same exit proceeds = dramatically higher IRR/MOIC

    Returns None if virtu_wacc is not configured.
    """
    if config.virtu_wacc is None:
        return None

    # Step 1: Compute Virtu DCF (perpetuity growth, incl synergies) at virtu_wacc
    virtu_wacc = config.virtu_wacc
    tgr = config.terminal_growth if config.terminal_growth is not None else 0.01
    n = config.horizon_years
    ne_key = 'net_earnings_incl'

    # PV of forecast cash flows at virtu_wacc
    pv_cfs = 0.0
    for yt in year_totals:
        y = yt['year']
        ne = yt.get(ne_key, 0)
        df = 1 / ((1 + virtu_wacc) ** y)
        pv_cfs += ne * df

    # Terminal value (perpetuity growth method)
    last_ne = year_totals[-1].get(ne_key, 0) if year_totals else 0
    if virtu_wacc > tgr:
        tv = last_ne * (1 + tgr) / (virtu_wacc - tgr)
    else:
        tv = last_ne * 20  # Cap
    exit_df = 1 / ((1 + virtu_wacc) ** n) if n > 0 else 1
    pv_tv = tv * exit_df
    virtu_dcf_pg_incl = pv_cfs + pv_tv

    # Step 2: Size debt against Virtu DCF value (not purchase price)
    virtu_debt = virtu_dcf_pg_incl * config.ltv
    virtu_equity = config.purchase_price - virtu_debt

    if virtu_equity <= 0:
        log.warning('Virtu equity <= 0 (DCF=%.0f, debt=%.0f, PP=%.0f) — skipping',
                    virtu_dcf_pg_incl, virtu_debt, config.purchase_price)
        return None

    # Step 3: Levered returns with overridden debt/equity
    # (Same UFCF, same exit multiple, same SOFR interest — just different capital structure)
    result = _compute_levered_returns_dated(
        year_totals, config, forecast_start,
        debt_override=virtu_debt,
        equity_override=virtu_equity,
    )

    if result:
        result['virtu_dcf_pg_incl'] = round(virtu_dcf_pg_incl, 2)
        result['methodology'] = 'virtu'

    return result


def compute_irr_moic_sensitivity(year_totals: List[dict],
                                  config: ForecastConfig,
                                  forecast_start: date) -> dict:
    """Compute IRR/MOIC sensitivity grid across purchase prices × exit multiples.

    Returns {purchase_prices, exit_multiples, irr_matrix, moic_matrix,
             xltm_values, xntm_values}.
    Default: 5 purchase prices (centered ±20%), 7 exit multiples (9-15x).
    """
    ne_key = 'net_earnings_incl'
    if not year_totals:
        return {}

    # Build grid axes
    if config.irr_purchase_prices:
        prices = sorted(config.irr_purchase_prices)
    elif config.purchase_price > 0:
        base = config.purchase_price
        prices = [round(base * m, 0) for m in [0.80, 0.90, 1.0, 1.10, 1.20]]
    else:
        return {}

    if config.irr_exit_multiples:
        exit_mults = sorted(config.irr_exit_multiples)
    else:
        exit_mults = [9.0, 10.0, 11.0, 12.0, 13.0, 14.0, 15.0]

    # Compute LTM/NTM for xLTM/xNTM
    ltm_ne = 0
    ntm_ne = year_totals[0].get(ne_key, 0) if year_totals else 0

    irr_matrix = []
    moic_matrix = []
    xltm_values = []
    xntm_values = []

    for pp in prices:
        if pp <= 0:
            irr_matrix.append([None] * len(exit_mults))
            moic_matrix.append([None] * len(exit_mults))
            xltm_values.append(None)
            xntm_values.append(None)
            continue

        xltm_values.append(round(pp / ltm_ne, 1) if ltm_ne > 0 else None)
        xntm_values.append(round(pp / ntm_ne, 1) if ntm_ne > 0 else None)

        irr_row = []
        moic_row = []
        for em in exit_mults:
            # Unlevered CF: year 0 = -pp, years 1-N = NE incl, exit year += NE * em
            total_cf = 0.0
            xirr_flows = [(forecast_start, -pp)]
            for yt in year_totals:
                ne = yt.get(ne_key, 0)
                is_exit = (yt['year'] == config.horizon_years)
                exit_proc = ne * em if is_exit else 0
                total_flow = ne + exit_proc
                cy = yt.get('calendar_year', forecast_start.year + yt['year'])
                xirr_flows.append((date(cy, 7, 1), total_flow))
                total_cf += total_flow

            moic = total_cf / pp if pp > 0 else 0
            irr = _xirr(xirr_flows)

            irr_row.append(round(irr, 4) if irr is not None else None)
            moic_row.append(round(moic, 2))

        irr_matrix.append(irr_row)
        moic_matrix.append(moic_row)

    return {
        'purchase_prices': prices,
        'exit_multiples': exit_mults,
        'irr_matrix': irr_matrix,
        'moic_matrix': moic_matrix,
        'xltm_values': xltm_values,
        'xntm_values': xntm_values,
    }


def compute_levered_irr_moic_sensitivity(year_totals: List[dict],
                                          config: ForecastConfig,
                                          forecast_start: date) -> dict:
    """Compute levered IRR/MOIC sensitivity grid across purchase prices × exit multiples.

    Uses the same axes as the unlevered grid. For each (price, exit_mult) combo,
    runs the full debt schedule to get levered IRR and MOIC.
    Returns {purchase_prices, exit_multiples, irr_matrix, moic_matrix}.
    """
    ne_key = 'net_earnings_incl'
    if not year_totals:
        return {}

    # Build grid axes (same logic as unlevered)
    if config.irr_purchase_prices:
        prices = sorted(config.irr_purchase_prices)
    elif config.purchase_price > 0:
        base = config.purchase_price
        prices = [round(base * m, 0) for m in [0.80, 0.90, 1.0, 1.10, 1.20]]
    else:
        return {}

    if config.irr_exit_multiples:
        exit_mults = sorted(config.irr_exit_multiples)
    else:
        exit_mults = [9.0, 10.0, 11.0, 12.0, 13.0, 14.0, 15.0]

    # Close date for XIRR
    y0_date = forecast_start
    if config.close_date:
        try:
            y0_date = date.fromisoformat(config.close_date[:10])
        except (ValueError, TypeError):
            pass

    irr_matrix = []
    moic_matrix = []

    for pp in prices:
        if pp <= 0:
            irr_matrix.append([None] * len(exit_mults))
            moic_matrix.append([None] * len(exit_mults))
            continue

        irr_row = []
        moic_row = []
        for em in exit_mults:
            # Debt schedule for this (pp, em) combo
            debt_initial = pp * config.ltv
            equity = pp - debt_initial
            if equity <= 0:
                irr_row.append(None)
                moic_row.append(None)
                continue

            balance = debt_initial
            total_lfcf = 0.0
            xirr_flows = [(y0_date, -equity)]

            for i, yt in enumerate(year_totals):
                ufcf = yt.get(ne_key, 0)
                opening = balance
                interest_rate = get_interest_rate_for_year(yt['year'], config, forecast_start)

                # Iterative interest/principal solve
                est_interest = opening * interest_rate
                for _ in range(10):
                    available = max(0, ufcf - est_interest) * config.cash_flow_sweep
                    principal = min(opening, available)
                    closing = opening - principal
                    new_interest = ((opening + closing) / 2) * interest_rate
                    if abs(new_interest - est_interest) < 0.01:
                        est_interest = new_interest
                        break
                    est_interest = new_interest

                interest = est_interest
                available = max(0, ufcf - interest) * config.cash_flow_sweep
                principal = min(opening, available)
                closing = opening - principal
                lfcf = ufcf - interest - principal

                total_lfcf += lfcf
                balance = closing

                # Exit at horizon
                is_exit = (yt['year'] == config.horizon_years)
                last_ne = ufcf
                exit_equity_cf = 0
                if is_exit:
                    exit_ev = last_ne * em
                    exit_equity_cf = exit_ev - balance

                cf = lfcf + exit_equity_cf
                cf_date = date(y0_date.year + yt['year'], y0_date.month,
                               min(y0_date.day, 28))
                xirr_flows.append((cf_date, cf))

            # Final exit equity for MOIC
            last_ne_val = year_totals[-1].get(ne_key, 0) if year_totals else 0
            exit_ev = last_ne_val * em
            exit_equity_val = exit_ev - balance
            total_distributions = total_lfcf + exit_equity_val
            moic = total_distributions / equity if equity > 0 else 0

            irr = _xirr(xirr_flows)
            irr_row.append(round(irr, 4) if irr is not None else None)
            moic_row.append(round(moic, 2))

        irr_matrix.append(irr_row)
        moic_matrix.append(moic_row)

    return {
        'purchase_prices': prices,
        'exit_multiples': exit_mults,
        'irr_matrix': irr_matrix,
        'moic_matrix': moic_matrix,
    }


def compute_forecast_cohorts(isrc_projections: dict, isrc_metadata: dict,
                              historical_by_isrc: dict,
                              config: ForecastConfig,
                              forecast_start: date) -> dict:
    """Age-bucket analysis: Pre 1, 1-2, 2-3, ... 9-10, 10+.

    Returns {buckets: [{label, historical_gross: {year: val}, forecast_gross: {year: val}}]}.
    """
    buckets_def = [
        ('Pre 1', 0, 1), ('1-2', 1, 2), ('2-3', 2, 3), ('3-4', 3, 4),
        ('4-5', 4, 5), ('5-6', 5, 6), ('6-7', 6, 7), ('7-8', 7, 8),
        ('8-9', 8, 9), ('9-10', 9, 10), ('10+', 10, 999),
    ]

    # Assign each ISRC to an age bucket based on release date
    bucket_isrcs = {b[0]: [] for b in buckets_def}

    for isrc, ip in isrc_projections.items():
        rd_str = ip.get('release_date')
        if rd_str:
            try:
                rd = date.fromisoformat(str(rd_str)[:10])
                age = (forecast_start - rd).days / 365.25
            except (ValueError, TypeError):
                age = 3.0
        else:
            age = 3.0

        for label, lo, hi in buckets_def:
            if lo <= age < hi:
                bucket_isrcs[label].append(isrc)
                break

    # Aggregate per bucket
    buckets = []
    for label, lo, hi in buckets_def:
        isrcs = bucket_isrcs[label]
        forecast_by_year = {}
        ltm_gross = 0

        for isrc in isrcs:
            ip = isrc_projections.get(isrc, {})
            ltm_gross += ip.get('baseline', {}).get('gross', 0)
            for proj in ip.get('projections', []):
                y = proj['year']
                forecast_by_year[y] = forecast_by_year.get(y, 0) + proj.get('gross', 0)

        buckets.append({
            'label': label,
            'isrc_count': len(isrcs),
            'ltm_gross': round(ltm_gross, 2),
            'forecast_by_year': {k: round(v, 2) for k, v in forecast_by_year.items()},
        })

    return {'buckets': buckets}


def compute_wa_release_date_per_payor(payor_results: dict,
                                       ltm_start_period: int = 0) -> dict:
    """Compute weighted-average release date per payor + overall.

    Returns {payor_code: {wa_date, days_old, ltm_gross, pct_of_ltm}, overall: {...}}.
    """
    import pandas as pd
    from datetime import date as _date

    today = _date.today()
    per_payor = {}
    total_gross = 0

    for code, pr in payor_results.items():
        detail = pr.detail
        if detail is None or detail.empty:
            continue

        if ltm_start_period:
            ltm_d = detail[detail['period'] >= ltm_start_period]
        else:
            ltm_d = detail

        if 'identifier' not in ltm_d.columns or 'gross' not in ltm_d.columns:
            continue

        payor_gross = float(ltm_d['gross'].sum())
        total_gross += payor_gross

        weighted_days = 0.0
        gross_with_dates = 0.0

        if 'release_date' in ltm_d.columns:
            isrc_agg = ltm_d.groupby('identifier').agg(
                {'gross': 'sum', 'release_date': 'first'}).reset_index()
            for _, row in isrc_agg.iterrows():
                g = float(row['gross']) if pd.notna(row['gross']) else 0
                if g <= 0:
                    continue
                rd = str(row.get('release_date', '')).strip()
                if not rd or rd in ('', 'nan', 'None', 'NaT') or len(rd) < 4:
                    continue
                try:
                    rd_clean = rd[:10]
                    release = _date.fromisoformat(rd_clean)
                    days_old = (today - release).days
                    weighted_days += days_old * g
                    gross_with_dates += g
                except (ValueError, TypeError):
                    continue

        if gross_with_dates > 0:
            wa_days = weighted_days / gross_with_dates
            wa_date = today - __import__('datetime').timedelta(days=int(wa_days))
            per_payor[code] = {
                'wa_date': wa_date.isoformat(),
                'days_old': round(wa_days, 0),
                'ltm_gross': round(payor_gross, 2),
            }
        else:
            per_payor[code] = {
                'wa_date': None,
                'days_old': None,
                'ltm_gross': round(payor_gross, 2),
            }

    # Compute pct_of_ltm
    for code, info in per_payor.items():
        info['pct_of_ltm'] = round(info['ltm_gross'] / total_gross * 100, 1) if total_gross > 0 else 0

    # Overall
    overall_weighted = sum(
        (info['days_old'] or 0) * info['ltm_gross']
        for info in per_payor.values() if info['days_old'] is not None
    )
    overall_gross_with_dates = sum(
        info['ltm_gross']
        for info in per_payor.values() if info['days_old'] is not None
    )
    if overall_gross_with_dates > 0:
        overall_days = overall_weighted / overall_gross_with_dates
        overall_date = today - __import__('datetime').timedelta(days=int(overall_days))
    else:
        overall_days = 0
        overall_date = today

    return {
        'per_payor': per_payor,
        'overall': {
            'wa_date': overall_date.isoformat(),
            'days_old': round(overall_days, 0),
            'total_gross': round(total_gross, 2),
        },
    }


def _empty_forecast(config: ForecastConfig) -> dict:
    return {
        'config': config.to_dict(),
        'forecast_start': date.today().isoformat(),
        'isrc_count': 0,
        'aggregate': {'year_totals': [], 'total_gross': 0, 'total_net_excl': 0,
                       'total_net_incl': 0, 'total_net': 0},
        'ltm_waterfall': {'gross': 0, 'fees': 0, 'net_receipts': 0,
                          'third_party': 0, 'net_earnings': 0},
        'ntm': {'gross': 0, 'net_excl': 0, 'net_incl': 0},
        'dcf': {'terminal_multiple': {'excl': {}, 'incl': {}},
                'perpetuity_growth': {'excl': {}, 'incl': {}}},
        'sensitivity': {'terminal_multiple': {}, 'perpetuity_growth': {}},
        'irr_sensitivity': {},
        'levered_sensitivity': {},
        'unlevered_returns': {},
        'levered_returns': {},
        'virtu_levered_returns': None,
        'summary': {'terminal_value': 0, 'npv': 0, 'npv_perpetuity': 0,
                     'implied_multiple_net': 0, 'implied_multiple_gross': 0,
                     'ltm_gross': 0, 'ltm_net': 0,
                     'terminal_growth': 0, 'discount_rate': config.discount_rate},
        'top_isrcs': [],
        'per_payor': {},
        'by_income_type': {},
        'cohorts': {'buckets': []},
        'wa_release_dates': {},
        'isrc_projections': {},
        'isrc_metadata': {},
    }


def _gather_release_dates(payor_results: dict) -> Dict[str, str]:
    """Collect release dates from all payors."""
    dates = {}
    for code, pr in payor_results.items():
        if hasattr(pr, 'detail') and pr.detail is not None and 'Release Date' in pr.detail.columns:
            id_col = 'ISRC' if 'ISRC' in pr.detail.columns else 'identifier'
            if id_col in pr.detail.columns:
                for _, row in pr.detail[[id_col, 'Release Date']].drop_duplicates(id_col).iterrows():
                    isrc = str(row[id_col]).strip()
                    rd = str(row['Release Date']).strip()
                    if isrc and rd and len(rd) >= 4 and rd != 'nan' and isrc not in dates:
                        dates[isrc] = rd
        if hasattr(pr, 'isrc_meta') and 'release_date' in pr.isrc_meta.columns:
            for _, row in pr.isrc_meta.iterrows():
                isrc = str(row['identifier']).strip()
                rd = str(row.get('release_date', '')).strip()
                if isrc and rd and len(rd) >= 4 and rd != 'nan' and isrc not in dates:
                    dates[isrc] = rd
    return dates


def _gather_metadata(payor_results: dict) -> Dict[str, dict]:
    """Collect title/artist metadata from all payors."""
    meta = {}
    for code, pr in payor_results.items():
        if hasattr(pr, 'isrc_meta'):
            for _, row in pr.isrc_meta.iterrows():
                isrc = str(row['identifier']).strip()
                if isrc and isrc not in meta:
                    meta[isrc] = {
                        'title': str(row.get('title', isrc))[:60],
                        'artist': str(row.get('artist', ''))[:40],
                    }
    return meta


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def export_forecast_excel(result: dict, output_path: str, deal_name: str = ''):
    """Export forecast to Excel with full model parity — 13 sheet types."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl required for Excel export")
        return None

    wb = openpyxl.Workbook()

    # Shared styles
    hdr = Font(bold=True, size=12)
    hdr2 = Font(bold=True, size=11)
    bf = Font(bold=True)
    mf = '#,##0'
    mf2 = '#,##0.00'
    pf = '0.0%'
    pf2 = '0.00%'
    mxf = '0.0x'
    blue = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    green = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    gray = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    yellow = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    dk_blue = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    dk_blue_font = Font(bold=True, color='FFFFFF', size=11)

    config = result.get('config', {})
    summary = result.get('summary', {})
    year_totals = result.get('aggregate', {}).get('year_totals', [])
    ltm_wf = result.get('ltm_waterfall', {})
    dcf = result.get('dcf', {})
    sens = result.get('sensitivity', {})
    irr_sens = result.get('irr_sensitivity', {})
    unlev = result.get('unlevered_returns', {})
    lev = result.get('levered_returns', {})
    virtu_lev = result.get('virtu_levered_returns') or {}
    per_payor = result.get('per_payor', {})
    by_income = result.get('by_income_type', {})
    cohorts = result.get('cohorts', {})
    wa_dates = result.get('wa_release_dates', {})
    isrc_proj = result.get('isrc_projections', {})
    isrc_meta = result.get('isrc_metadata', {})
    top_isrcs = result.get('top_isrcs', [])
    n_years = len(year_totals)
    horizon = config.get('horizon_years', 5)
    dr = config.get('discount_rate', 0.09375)
    tgr = summary.get('terminal_growth', 0.01)

    def _fill_row(ws, row, col_start, values, font=None, fill=None, fmt=None):
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=col_start + i, value=v)
            if font:
                c.font = font
            if fill:
                c.fill = fill
            if fmt:
                c.number_format = fmt

    def _section_hdr(ws, row, text, cols=4, fill_color=None):
        fill_c = fill_color or blue
        for c in range(1, cols + 1):
            ws.cell(row=row, column=c).fill = fill_c
        ws.cell(row=row, column=1, value=text).font = hdr
        if fill_c == dk_blue:
            ws.cell(row=row, column=1).font = dk_blue_font

    # ====================================================================
    # Sheet 1: Summary - Overview
    # ====================================================================
    ws1 = wb.active
    ws1.title = 'Summary - Overview'

    ws1['A1'] = deal_name or 'Forecast Model'
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws1['A2'].font = Font(size=10, color='666666')

    r = 4
    _section_hdr(ws1, r, 'Deal Summary', 3, dk_blue)
    r += 1
    deal_items = [
        ('Opportunity Name', config.get('opportunity_name', deal_name)),
        ('Opportunity Details', config.get('opportunity_details', 'Catalog Acquisition')),
        ('Rights Included', config.get('rights_included', 'Masters')),
        ('Deal Type', config.get('deal_type', 'Catalog')),
        ('Cash Date', config.get('cash_date', '')),
        ('Close Date', config.get('close_date', '')),
        ('WA Catalog Age', f"{summary.get('ltm_gross', 0):.0f}"),
        ('Period End Date', result.get('forecast_start', '')),
        ('ISRCs Projected', result.get('isrc_count', 0)),
    ]
    for lbl, val in deal_items:
        ws1.cell(row=r, column=1, value=lbl).font = bf
        ws1.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    _section_hdr(ws1, r, 'Underwriting Assumptions', 3, dk_blue)
    r += 1
    for lbl, val, fmt in [
        ('CMG WACC', dr, pf2), ('Virtu WACC', config.get('virtu_wacc') or 0, pf2),
        ('Terminal Growth Rate', tgr, pf), ('Exit Multiple', config.get('exit_multiple', 15), mxf),
        ('Default Decay Curve', config.get('genre_default', 'default'), None),
        ('Horizon Years', config.get('horizon_years', 5), '0'),
    ]:
        ws1.cell(row=r, column=1, value=lbl).font = bf
        c = ws1.cell(row=r, column=2, value=val)
        if fmt:
            c.number_format = fmt
        r += 1

    r += 1
    _section_hdr(ws1, r, 'Purchase Structure', 3, dk_blue)
    r += 1
    pp = config.get('purchase_price', 0)
    hb = config.get('holdback', 0)
    pcdpcdr = config.get('pcdpcdr', 0)
    closing = pp - hb - pcdpcdr
    for lbl, val in [
        ('Total Purchase Price', pp), ('PCDPCDR', pcdpcdr),
        ('Holdback', hb), ('Closing Amount', closing),
    ]:
        ws1.cell(row=r, column=1, value=lbl).font = bf
        ws1.cell(row=r, column=2, value=val).number_format = mf
        r += 1

    r += 1
    _section_hdr(ws1, r, 'Catalog Earnings Summary', 4, dk_blue)
    r += 1
    ltm_g = summary.get('ltm_gross', 0)
    ltm_n = summary.get('ltm_net', 0)
    ntm_g = summary.get('ntm_gross', 0)
    ntm_n = summary.get('ntm_net_incl', 0)
    for lbl, ltm_v, ntm_v in [
        ('Gross', ltm_g, ntm_g), ('Net Earnings (Incl)', ltm_n, ntm_n),
    ]:
        ws1.cell(row=r, column=1, value=lbl).font = bf
        ws1.cell(row=r, column=2, value='LTM')
        ws1.cell(row=r, column=3, value=ltm_v).number_format = mf
        r += 1
        ws1.cell(row=r, column=2, value='NTM')
        ws1.cell(row=r, column=3, value=ntm_v).number_format = mf
        if pp > 0:
            ws1.cell(row=r, column=4, value=f'{pp / ntm_v:.1f}x' if ntm_v > 0 else 'N/A')
        r += 1

    r += 1
    _section_hdr(ws1, r, 'Returns', 3, dk_blue)
    r += 1
    returns_data = []
    if unlev:
        returns_data.extend([
            ('Unlevered IRR', unlev.get('irr') or 0, pf),
            ('Unlevered MOIC', unlev.get('moic', 0), '0.00x'),
        ])
    if lev:
        returns_data.extend([
            ('Levered IRR (CMG)', lev.get('irr') or 0, pf),
            ('Levered MOIC (CMG)', lev.get('moic', 0), '0.00x'),
        ])
    if virtu_lev:
        returns_data.extend([
            ('Virtu Levered IRR', virtu_lev.get('irr') or 0, pf),
            ('Virtu Levered MOIC', virtu_lev.get('moic', 0), '0.00x'),
        ])
    for lbl, val, fmt in returns_data:
        ws1.cell(row=r, column=1, value=lbl).font = bf
        ws1.cell(row=r, column=2, value=val).number_format = fmt
        r += 1

    r += 1
    _section_hdr(ws1, r, 'Valuation Details', 3, dk_blue)
    r += 1
    for lbl, val in [
        ('DCF (TM, Excl)', dcf.get('terminal_multiple', {}).get('excl', {}).get('implied_valuation', 0)),
        ('DCF (TM, Incl)', dcf.get('terminal_multiple', {}).get('incl', {}).get('implied_valuation', 0)),
        ('DCF (PG, Excl)', dcf.get('perpetuity_growth', {}).get('excl', {}).get('implied_valuation', 0)),
        ('DCF (PG, Incl)', dcf.get('perpetuity_growth', {}).get('incl', {}).get('implied_valuation', 0)),
        ('Purchase Price', pp),
    ]:
        ws1.cell(row=r, column=1, value=lbl).font = bf
        ws1.cell(row=r, column=2, value=val).number_format = mf
        r += 1

    # Unlevered + Levered CF schedules
    r += 1
    if unlev and unlev.get('schedule'):
        _section_hdr(ws1, r, 'Unlevered Cash Flow Schedule', n_years + 2)
        r += 1
        cols = [''] + [f"Yr {yt['year']}" for yt in year_totals]
        _fill_row(ws1, r, 1, cols, font=bf, fill=gray)
        r += 1
        ws1.cell(row=r, column=1, value='UFCF').font = bf
        for i, s in enumerate(unlev.get('schedule', []), 2):
            ws1.cell(row=r, column=i, value=s['ufcf']).number_format = mf
        r += 1
        ws1.cell(row=r, column=1, value='Exit Proceeds').font = bf
        for i, s in enumerate(unlev.get('schedule', []), 2):
            ws1.cell(row=r, column=i, value=s['exit_proceeds']).number_format = mf
        r += 1
        ws1.cell(row=r, column=1, value='Total').font = bf
        for i, s in enumerate(unlev.get('schedule', []), 2):
            ws1.cell(row=r, column=i, value=s['total']).number_format = mf
        r += 2

    if lev and lev.get('debt_schedule'):
        _section_hdr(ws1, r, 'Levered Cash Flow Schedule', n_years + 2)
        r += 1
        cols = [''] + [f"Yr {yt['year']}" for yt in year_totals]
        _fill_row(ws1, r, 1, cols, font=bf, fill=gray)
        r += 1
        for field_lbl, field_key in [('UFCF', 'ufcf'), ('Interest', 'interest'),
                                      ('Principal', 'principal'), ('LFCF', 'lfcf')]:
            ws1.cell(row=r, column=1, value=field_lbl).font = bf
            for i, ds in enumerate(lev['debt_schedule'], 2):
                ws1.cell(row=r, column=i, value=ds[field_key]).number_format = mf
            r += 1

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 14

    # ====================================================================
    # Sheet 2: Valuation & Returns
    # ====================================================================
    ws2 = wb.create_sheet('Valuation & Returns')

    r = 1
    _section_hdr(ws2, r, 'Assumptions / Inputs', 4, dk_blue)
    r += 1
    for lbl, val, fmt in [
        ('WACC (CMG)', dr, pf2), ('Virtu WACC', config.get('virtu_wacc') or 0, pf2),
        ('Terminal Growth Rate', tgr, pf), ('Exit Multiple', config.get('exit_multiple', 15), mxf),
        ('Purchase Price', pp, mf), ('LTV', config.get('ltv', 0.55), pf),
    ]:
        ws2.cell(row=r, column=1, value=lbl).font = bf
        ws2.cell(row=r, column=2, value=val).number_format = fmt
        r += 1

    r += 1
    _section_hdr(ws2, r, 'Financing Assumptions', 4, dk_blue)
    r += 1
    for lbl, val, fmt in [
        ('SOFR Rate', config.get('sofr_rate', 0.045), pf),
        ('SOFR Floor', config.get('sofr_floor', 0.02), pf),
        ('SOFR Spread', config.get('sofr_spread', 0.0275), pf2),
        ('Cash Flow Sweep', config.get('cash_flow_sweep', 1.0), pf),
        ('Debt Amount', pp * config.get('ltv', 0.55), mf),
        ('Equity', pp * (1 - config.get('ltv', 0.55)), mf),
    ]:
        ws2.cell(row=r, column=1, value=lbl).font = bf
        ws2.cell(row=r, column=2, value=val).number_format = fmt
        r += 1

    # DCF sections
    for method_lbl, method_key in [('DCF Analysis — Terminal Multiple', 'terminal_multiple'),
                                     ('DCF Analysis — Perpetuity Growth', 'perpetuity_growth')]:
        r += 1
        _section_hdr(ws2, r, method_lbl, 4, blue)
        r += 1
        method_data = dcf.get(method_key, {})
        for track_lbl, track_key in [('Excl Synergies', 'excl'), ('Incl Synergies', 'incl')]:
            ws2.cell(row=r, column=1, value=track_lbl).font = bf
            ws2.cell(row=r, column=1).fill = gray
            r += 1
            t = method_data.get(track_key, {})
            for lbl, val, fmt in [
                ('PV of Cash Flows', t.get('pv_cash_flows', 0), mf),
                ('Terminal Value', t.get('terminal_value', 0), mf),
                ('PV of Terminal Value', t.get('pv_terminal_value', 0), mf),
                ('Implied Valuation', t.get('implied_valuation', 0), mf),
                ('Implied LTM Multiple', t.get('implied_ltm_multiple', 0), mxf),
            ]:
                ws2.cell(row=r, column=1, value=lbl)
                c = ws2.cell(row=r, column=2, value=val)
                c.number_format = fmt
                if lbl == 'Implied Valuation':
                    c.font = Font(bold=True, color='1D4ED8')
                r += 1
            r += 1

    # 3x3 sensitivity tables
    for method_lbl, method_key in [('Sensitivity — Terminal Multiple', 'terminal_multiple'),
                                     ('Sensitivity — Perpetuity Growth', 'perpetuity_growth')]:
        ms = sens.get(method_key, {})
        if not ms:
            continue
        r += 1
        _section_hdr(ws2, r, method_lbl, 6, yellow)
        r += 1
        wacc_vals = ms.get('wacc_values', [])
        if method_key == 'terminal_multiple':
            col_vals = ms.get('exit_mult_values', [])
        else:
            col_vals = ms.get('tgr_values', [])
        for track_lbl, matrix_key in [('Excl', 'matrix_excl'), ('Incl', 'matrix_incl')]:
            ws2.cell(row=r, column=1, value=f'{track_lbl} Synergies').font = bf
            r += 1
            # Header
            ws2.cell(row=r, column=1, value='WACC').font = bf
            for ci, cv in enumerate(col_vals, 2):
                ws2.cell(row=r, column=ci, value=cv).font = bf
                ws2.cell(row=r, column=ci).fill = gray
            r += 1
            matrix = ms.get(matrix_key, [])
            for ri, wp in enumerate(wacc_vals):
                ws2.cell(row=r, column=1, value=f'{wp:.1f}%').font = bf
                ws2.cell(row=r, column=1).fill = gray
                if ri < len(matrix):
                    for ci, v in enumerate(matrix[ri], 2):
                        ws2.cell(row=r, column=ci, value=v).number_format = mf
                r += 1
            r += 1

    # Unlevered returns schedule
    if unlev:
        r += 1
        _section_hdr(ws2, r, 'Unlevered Returns', n_years + 3, blue)
        r += 1
        _fill_row(ws2, r, 1, ['', 'Yr 0'] + [f"Yr {y}" for y in range(1, horizon + 1)], font=bf, fill=gray)
        r += 1
        ws2.cell(row=r, column=1, value='UFCF').font = bf
        ws2.cell(row=r, column=2, value=-pp).number_format = mf
        for i, s in enumerate(unlev.get('schedule', []), 3):
            ws2.cell(row=r, column=i, value=s['total']).number_format = mf
        r += 1
        ws2.cell(row=r, column=1, value='IRR').font = bf
        ws2.cell(row=r, column=2, value=unlev.get('irr') or 0).number_format = pf
        ws2.cell(row=r, column=1 + horizon + 1, value='MOIC').font = bf
        ws2.cell(row=r, column=1 + horizon + 2, value=unlev.get('moic', 0)).number_format = '0.00x'
        r += 1

    # IRR/MOIC sensitivity grids
    if irr_sens and irr_sens.get('purchase_prices'):
        r += 1
        _section_hdr(ws2, r, 'IRR Sensitivity', len(irr_sens['exit_multiples']) + 3, yellow)
        r += 1
        prices = irr_sens['purchase_prices']
        exit_ms = irr_sens['exit_multiples']
        # Header
        ws2.cell(row=r, column=1, value='Purchase Price').font = bf
        ws2.cell(row=r, column=2, value='xLTM').font = bf
        for ci, em in enumerate(exit_ms, 3):
            ws2.cell(row=r, column=ci, value=f'{em:.0f}x').font = bf
            ws2.cell(row=r, column=ci).fill = gray
        r += 1
        irr_mat = irr_sens.get('irr_matrix', [])
        xltm = irr_sens.get('xltm_values', [])
        for ri, price in enumerate(prices):
            ws2.cell(row=r, column=1, value=price).number_format = mf
            ws2.cell(row=r, column=2, value=f'{xltm[ri]:.1f}x' if ri < len(xltm) and xltm[ri] else '')
            if ri < len(irr_mat):
                for ci, v in enumerate(irr_mat[ri], 3):
                    c = ws2.cell(row=r, column=ci, value=v if v is not None else '')
                    if v is not None:
                        c.number_format = pf
            r += 1
        r += 1

        _section_hdr(ws2, r, 'MOIC Sensitivity', len(exit_ms) + 3, yellow)
        r += 1
        ws2.cell(row=r, column=1, value='Purchase Price').font = bf
        ws2.cell(row=r, column=2, value='xLTM').font = bf
        for ci, em in enumerate(exit_ms, 3):
            ws2.cell(row=r, column=ci, value=f'{em:.0f}x').font = bf
            ws2.cell(row=r, column=ci).fill = gray
        r += 1
        moic_mat = irr_sens.get('moic_matrix', [])
        for ri, price in enumerate(prices):
            ws2.cell(row=r, column=1, value=price).number_format = mf
            ws2.cell(row=r, column=2, value=f'{xltm[ri]:.1f}x' if ri < len(xltm) and xltm[ri] else '')
            if ri < len(moic_mat):
                for ci, v in enumerate(moic_mat[ri], 3):
                    c = ws2.cell(row=r, column=ci, value=v if v is not None else '')
                    if v is not None:
                        c.number_format = '0.00x'
            r += 1

    # Levered returns
    if lev and lev.get('debt_schedule'):
        r += 1
        _section_hdr(ws2, r, 'Levered Returns (CMG)', n_years + 3, green)
        r += 1
        _fill_row(ws2, r, 1, ['', 'Yr 0'] + [f"Yr {y}" for y in range(1, horizon + 1)], font=bf, fill=gray)
        r += 1
        for field_lbl, field_key in [('UFCF', 'ufcf'), ('Interest', 'interest'),
                                      ('Principal', 'principal'), ('LFCF', 'lfcf')]:
            ws2.cell(row=r, column=1, value=field_lbl).font = bf
            for i, ds in enumerate(lev['debt_schedule'], 3):
                ws2.cell(row=r, column=i, value=ds[field_key]).number_format = mf
            r += 1
        ws2.cell(row=r, column=1, value='IRR').font = bf
        ws2.cell(row=r, column=2, value=lev.get('irr') or 0).number_format = pf
        ws2.cell(row=r, column=1 + horizon + 1, value='MOIC').font = bf
        ws2.cell(row=r, column=1 + horizon + 2, value=lev.get('moic', 0)).number_format = '0.00x'
        r += 1

    # Virtu levered returns
    if virtu_lev and virtu_lev.get('debt_schedule'):
        r += 1
        _section_hdr(ws2, r, 'Virtu Levered Returns', n_years + 3, green)
        r += 1
        _fill_row(ws2, r, 1, ['', 'Yr 0'] + [f"Yr {y}" for y in range(1, horizon + 1)], font=bf, fill=gray)
        r += 1
        for field_lbl, field_key in [('UFCF', 'ufcf'), ('Interest', 'interest'),
                                      ('Principal', 'principal'), ('LFCF', 'lfcf')]:
            ws2.cell(row=r, column=1, value=field_lbl).font = bf
            for i, ds in enumerate(virtu_lev['debt_schedule'], 3):
                ws2.cell(row=r, column=i, value=ds[field_key]).number_format = mf
            r += 1
        ws2.cell(row=r, column=1, value='IRR').font = bf
        ws2.cell(row=r, column=2, value=virtu_lev.get('irr') or 0).number_format = pf
        ws2.cell(row=r, column=1 + horizon + 1, value='MOIC').font = bf
        ws2.cell(row=r, column=1 + horizon + 2, value=virtu_lev.get('moic', 0)).number_format = '0.00x'

    ws2.column_dimensions['A'].width = 26
    for c in range(2, 12):
        ws2.column_dimensions[get_column_letter(c)].width = 16

    # ====================================================================
    # Sheet 3: Summary - Earnings
    # ====================================================================
    ws3 = wb.create_sheet('Summary - Earnings')

    wf_fields = [
        ('Gross Revenue', 'gross'), ('Less: Distribution Fees', 'fees_original'),
        ('Net Receipts', 'net_receipts_excl'), ('Less: Third Party', 'third_party_excl'),
        ('NE (Excl Synergies)', 'net_earnings_excl'),
        ('Fee Savings', 'fee_savings'), ('3P Savings', 'tp_savings'),
        ('NE (Incl Synergies)', 'net_earnings_incl'),
    ]
    yr_cols = ['LTM'] + [f"Yr {yt['year']} ({yt['calendar_year']})" for yt in year_totals] + ['Total']

    r = 1
    # --- By Payor ---
    for code, pp_data in per_payor.items():
        _section_hdr(ws3, r, f"{pp_data.get('name', code)} ({pp_data.get('income_rights', '')})",
                     len(yr_cols) + 1, blue)
        r += 1
        _fill_row(ws3, r, 1, [''] + yr_cols, font=bf, fill=gray)
        r += 1
        p_yt = pp_data.get('year_totals', [])
        p_ltm_g = pp_data.get('ltm_gross', 0)
        for lbl, fld in wf_fields:
            ws3.cell(row=r, column=1, value=lbl).font = bf
            # LTM
            if fld == 'gross':
                ws3.cell(row=r, column=2, value=p_ltm_g).number_format = mf
            else:
                ws3.cell(row=r, column=2, value=0).number_format = mf
            total_v = 0
            for ci, yt in enumerate(p_yt, 3):
                v = yt.get(fld, 0)
                ws3.cell(row=r, column=ci, value=v).number_format = mf
                total_v += v
            ws3.cell(row=r, column=len(p_yt) + 3, value=round(total_v, 2)).number_format = mf
            r += 1
        r += 1

    # --- By Income Type ---
    for it_name, it_data in by_income.items():
        _section_hdr(ws3, r, f"By Income Type: {it_name}", len(yr_cols) + 1, green)
        r += 1
        _fill_row(ws3, r, 1, [''] + yr_cols, font=bf, fill=gray)
        r += 1
        it_yt = it_data.get('year_totals', [])
        for lbl, fld in wf_fields:
            ws3.cell(row=r, column=1, value=lbl).font = bf
            ws3.cell(row=r, column=2, value=0).number_format = mf
            total_v = 0
            for ci, yt in enumerate(it_yt, 3):
                v = yt.get(fld, 0)
                ws3.cell(row=r, column=ci, value=v).number_format = mf
                total_v += v
            ws3.cell(row=r, column=len(it_yt) + 3, value=round(total_v, 2)).number_format = mf
            r += 1
        r += 1

    # --- Total ---
    _section_hdr(ws3, r, 'Total', len(yr_cols) + 1, dk_blue)
    r += 1
    _fill_row(ws3, r, 1, [''] + yr_cols, font=bf, fill=gray)
    r += 1
    for lbl, fld in wf_fields:
        ws3.cell(row=r, column=1, value=lbl).font = bf
        if fld == 'gross':
            ws3.cell(row=r, column=2, value=ltm_wf.get('gross', 0)).number_format = mf
        elif fld == 'net_earnings_excl':
            ws3.cell(row=r, column=2, value=ltm_wf.get('net_earnings', 0)).number_format = mf
        else:
            ws3.cell(row=r, column=2, value=0).number_format = mf
        total_v = 0
        for ci, yt in enumerate(year_totals, 3):
            v = yt.get(fld, 0)
            ws3.cell(row=r, column=ci, value=v).number_format = mf
            total_v += v
        ws3.cell(row=r, column=n_years + 3, value=round(total_v, 2)).number_format = mf
        r += 1

    ws3.column_dimensions['A'].width = 26
    for c in range(2, n_years + 5):
        ws3.column_dimensions[get_column_letter(c)].width = 16

    # ====================================================================
    # Sheet 4: Finance Forecast (monthly)
    # ====================================================================
    ws4 = wb.create_sheet('Finance Forecast')
    ws4.cell(row=1, column=1, value='Finance Forecast — Monthly Model Outputs').font = hdr
    ws4.cell(row=2, column=1, value='Monthly columns for manual actuals tracking').font = Font(size=10, color='666666')

    # Generate monthly column headers from forecast_start
    fs_str = result.get('forecast_start', '')
    if fs_str:
        try:
            fs = date.fromisoformat(fs_str[:10])
        except ValueError:
            fs = date.today()
    else:
        fs = date.today()

    monthly_cols = []
    for y in range(1, horizon + 1):
        for m in range(1, 13):
            md = date(fs.year + y, m, 1)
            monthly_cols.append(md.strftime('%b %Y'))

    r = 4
    ws4.cell(row=r, column=1, value='MODEL OUTPUTS').font = hdr
    ws4.cell(row=r, column=1).fill = blue
    r += 1
    _fill_row(ws4, r, 1, [''] + monthly_cols[:24], font=bf, fill=gray)  # Show max 2 years monthly
    r += 1
    for lbl in ['Gross Revenue', 'Less: Dist Fee', 'Less: Royalty COGS', 'NE (Excl)', 'Fee Savings', 'NE (Incl)']:
        ws4.cell(row=r, column=1, value=lbl).font = bf
        r += 1

    r += 2
    ws4.cell(row=r, column=1, value='ACTUAL RECEIPTS').font = hdr
    ws4.cell(row=r, column=1).fill = green
    r += 1
    _fill_row(ws4, r, 1, [''] + monthly_cols[:24], font=bf, fill=gray)
    r += 1
    for lbl in ['Actual Gross', 'Actuals vs Forecast', '% Difference']:
        ws4.cell(row=r, column=1, value=lbl).font = bf
        r += 1

    ws4.column_dimensions['A'].width = 22
    for c in range(2, 26):
        ws4.column_dimensions[get_column_letter(c)].width = 12

    # ====================================================================
    # Sheet 5: Catalog Detail (cohorts + top songs)
    # ====================================================================
    ws5 = wb.create_sheet('Catalog Detail')

    r = 1
    _section_hdr(ws5, r, 'Cohort Analysis by Age Bucket', n_years + 3, dk_blue)
    r += 1
    _fill_row(ws5, r, 1, ['Bucket', 'ISRCs', 'LTM Gross'] + [f"Yr {y}" for y in range(1, horizon + 1)],
              font=bf, fill=gray)
    r += 1

    buckets = cohorts.get('buckets', [])
    for b in buckets:
        ws5.cell(row=r, column=1, value=b['label']).font = bf
        ws5.cell(row=r, column=2, value=b.get('isrc_count', 0))
        ws5.cell(row=r, column=3, value=b.get('ltm_gross', 0)).number_format = mf
        for y in range(1, horizon + 1):
            ws5.cell(row=r, column=3 + y, value=b.get('forecast_by_year', {}).get(y, 0)).number_format = mf
        r += 1

    # Total row
    ws5.cell(row=r, column=1, value='Total').font = bf
    ws5.cell(row=r, column=1).fill = gray
    ws5.cell(row=r, column=2, value=sum(b.get('isrc_count', 0) for b in buckets))
    ws5.cell(row=r, column=3, value=sum(b.get('ltm_gross', 0) for b in buckets)).number_format = mf
    for y in range(1, horizon + 1):
        ws5.cell(row=r, column=3 + y, value=sum(
            b.get('forecast_by_year', {}).get(y, 0) for b in buckets)).number_format = mf
    r += 2

    # Top songs
    _section_hdr(ws5, r, 'Top Songs', n_years + 8, dk_blue)
    r += 1
    _fill_row(ws5, r, 1, ['#', 'ISRC', 'Title', 'Artist', 'Release Year', 'Years Old', '% LTM',
                           'LTM Gross'] + [f"Yr {y}" for y in range(1, horizon + 1)],
              font=bf, fill=gray)
    r += 1

    for rank, ti in enumerate(top_isrcs[:100], 1):
        isrc_id = ti.get('isrc', '')
        ws5.cell(row=r, column=1, value=rank)
        ws5.cell(row=r, column=2, value=isrc_id)
        ws5.cell(row=r, column=3, value=ti.get('title', ''))
        ws5.cell(row=r, column=4, value=ti.get('artist', ''))
        rd = ti.get('release_date', '')
        ws5.cell(row=r, column=5, value=str(rd)[:4] if rd else '')
        if rd:
            try:
                age = (fs - date.fromisoformat(str(rd)[:10])).days / 365.25
                ws5.cell(row=r, column=6, value=round(age, 1))
            except (ValueError, TypeError):
                pass
        ws5.cell(row=r, column=7, value=ti.get('pct_ltm', 0)).number_format = '0.00%'
        ws5.cell(row=r, column=8, value=ti.get('ltm_gross', 0)).number_format = mf

        # Per-year forecast gross
        ip = isrc_proj.get(isrc_id, {})
        for y in range(1, horizon + 1):
            projs = ip.get('projections', [])
            if y - 1 < len(projs):
                ws5.cell(row=r, column=8 + y, value=projs[y - 1].get('gross', 0)).number_format = mf
        r += 1

    ws5.column_dimensions['A'].width = 5
    ws5.column_dimensions['B'].width = 16
    ws5.column_dimensions['C'].width = 30
    ws5.column_dimensions['D'].width = 20
    for c in range(5, n_years + 9):
        ws5.column_dimensions[get_column_letter(c)].width = 14

    # ====================================================================
    # Sheet 6: Metadata
    # ====================================================================
    ws6 = wb.create_sheet('Metadata')

    meta_headers = ['ISRC', 'Track Title', 'Artists', 'Release Date', 'Decay Curve',
                    'Reversion Date', '3P Share', 'Label Share', 'Fee Rate']
    _fill_row(ws6, 1, 1, meta_headers, font=bf, fill=blue)

    r = 2
    reversions = config.get('reversions', {})
    for isrc_id, ip in isrc_proj.items():
        ws6.cell(row=r, column=1, value=isrc_id)
        ws6.cell(row=r, column=2, value=ip.get('title', ''))
        ws6.cell(row=r, column=3, value=ip.get('artist', ''))
        ws6.cell(row=r, column=4, value=ip.get('release_date', ''))
        ws6.cell(row=r, column=5, value=ip.get('genre', ''))
        ws6.cell(row=r, column=6, value=reversions.get(isrc_id, ''))
        ws6.cell(row=r, column=7, value=ip.get('tp_ratio', 0)).number_format = pf
        ls = ip.get('label_share')
        ws6.cell(row=r, column=8, value=ls if ls is not None else '').number_format = pf if ls is not None else '@'
        ws6.cell(row=r, column=9, value=ip.get('fee_ratio', 0)).number_format = pf
        r += 1

    for c, w in enumerate([16, 30, 20, 12, 16, 12, 10, 10, 10], 1):
        ws6.column_dimensions[get_column_letter(c)].width = w

    # ====================================================================
    # Sheet 7: WA Release Date
    # ====================================================================
    ws7 = wb.create_sheet('WA Release Date')

    r = 1
    _section_hdr(ws7, r, 'Weighted Average Release Date by Payor', 6, dk_blue)
    r += 1
    _fill_row(ws7, r, 1, ['Payor', 'WA Release Date', 'Days Old', 'LTM Gross', '% of LTM'], font=bf, fill=gray)
    r += 1

    wa_per_payor = wa_dates.get('per_payor', {})
    for code, info in wa_per_payor.items():
        ws7.cell(row=r, column=1, value=per_payor.get(code, {}).get('name', code))
        ws7.cell(row=r, column=2, value=info.get('wa_date', ''))
        ws7.cell(row=r, column=3, value=info.get('days_old') or '')
        ws7.cell(row=r, column=4, value=info.get('ltm_gross', 0)).number_format = mf
        ws7.cell(row=r, column=5, value=(info.get('pct_of_ltm', 0) / 100)).number_format = pf
        r += 1

    r += 1
    overall = wa_dates.get('overall', {})
    ws7.cell(row=r, column=1, value='Overall WA Catalog Release Date').font = bf
    ws7.cell(row=r, column=2, value=overall.get('wa_date', ''))
    ws7.cell(row=r, column=3, value=overall.get('days_old', ''))

    for c, w in enumerate([20, 16, 12, 14, 10], 1):
        ws7.column_dimensions[get_column_letter(c)].width = w

    # ====================================================================
    # Sheet 8: Song Listing (combined)
    # ====================================================================
    ws8 = wb.create_sheet('Song Listing')

    song_hdrs = ['ISRC', 'Title', 'Artist', 'Rank LTM', 'LTM Gross'] + \
                [f"Yr {y}" for y in range(1, horizon + 1)]
    _fill_row(ws8, 1, 1, song_hdrs, font=bf, fill=blue)

    sorted_isrcs = sorted(top_isrcs, key=lambda x: x.get('ltm_gross', 0), reverse=True)
    r = 2
    for rank, ti in enumerate(sorted_isrcs, 1):
        isrc_id = ti.get('isrc', '')
        ws8.cell(row=r, column=1, value=isrc_id)
        ws8.cell(row=r, column=2, value=ti.get('title', ''))
        ws8.cell(row=r, column=3, value=ti.get('artist', ''))
        ws8.cell(row=r, column=4, value=rank)
        ws8.cell(row=r, column=5, value=ti.get('ltm_gross', 0)).number_format = mf
        ip = isrc_proj.get(isrc_id, {})
        for y in range(1, horizon + 1):
            projs = ip.get('projections', [])
            if y - 1 < len(projs):
                ws8.cell(row=r, column=5 + y, value=projs[y - 1].get('gross', 0)).number_format = mf
        r += 1

    # Total row
    ws8.cell(row=r, column=1, value='Total').font = bf
    ws8.cell(row=r, column=1).fill = gray
    ws8.cell(row=r, column=5, value=sum(t.get('ltm_gross', 0) for t in sorted_isrcs)).number_format = mf
    for y in range(1, horizon + 1):
        ws8.cell(row=r, column=5 + y, value=sum(
            isrc_proj.get(t.get('isrc', ''), {}).get('projections', [{}] * y)[y - 1].get('gross', 0)
            for t in sorted_isrcs if y - 1 < len(isrc_proj.get(t.get('isrc', ''), {}).get('projections', []))
        )).number_format = mf

    ws8.column_dimensions['A'].width = 16
    ws8.column_dimensions['B'].width = 30
    ws8.column_dimensions['C'].width = 20
    for c in range(4, horizon + 6):
        ws8.column_dimensions[get_column_letter(c)].width = 14

    # ====================================================================
    # Sheet 9: Song Listing per Payor (one sheet per payor)
    # ====================================================================
    for code, pp_data in per_payor.items():
        name = pp_data.get('name', code)
        # Sheet name max 31 chars
        sheet_name = f"Songs_{name}"[:31]
        wsp = wb.create_sheet(sheet_name)

        p_hdrs = ['ISRC', 'Title', 'Artist', 'Payor', 'Income Type', 'LTM Gross'] + \
                 [f"Yr {y}" for y in range(1, horizon + 1)]
        _fill_row(wsp, 1, 1, p_hdrs, font=bf, fill=blue)

        payor_isrcs = pp_data.get('isrcs', [])
        payor_isrc_data = [(isrc_id, isrc_proj.get(isrc_id, {})) for isrc_id in payor_isrcs]
        payor_isrc_data.sort(key=lambda x: x[1].get('baseline', {}).get('gross', 0), reverse=True)

        r = 2
        for isrc_id, ip in payor_isrc_data:
            wsp.cell(row=r, column=1, value=isrc_id)
            wsp.cell(row=r, column=2, value=ip.get('title', ''))
            wsp.cell(row=r, column=3, value=ip.get('artist', ''))
            wsp.cell(row=r, column=4, value=name)
            wsp.cell(row=r, column=5, value=pp_data.get('income_rights', ''))
            wsp.cell(row=r, column=6, value=ip.get('baseline', {}).get('gross', 0)).number_format = mf
            for y in range(1, horizon + 1):
                projs = ip.get('projections', [])
                if y - 1 < len(projs):
                    wsp.cell(row=r, column=6 + y, value=projs[y - 1].get('gross', 0)).number_format = mf
            r += 1

        wsp.column_dimensions['A'].width = 16
        wsp.column_dimensions['B'].width = 28
        wsp.column_dimensions['C'].width = 18
        wsp.column_dimensions['D'].width = 14
        wsp.column_dimensions['E'].width = 14
        for c in range(6, horizon + 7):
            wsp.column_dimensions[get_column_letter(c)].width = 14

    # ====================================================================
    # Sheet 10: Decay Curve
    # ====================================================================
    ws10 = wb.create_sheet('Decay Curve')

    r = 1
    _section_hdr(ws10, r, 'Decay Curves (FTI Report)', 13, dk_blue)
    r += 1
    _fill_row(ws10, r, 1, ['Curve'] + [f"Yr {y}" for y in range(1, 11)] + ['Terminal'], font=bf, fill=gray)
    r += 1

    for key, curve_data in DECAY_CURVES.items():
        ws10.cell(row=r, column=1, value=curve_data.get('label', key)).font = bf
        rates = curve_data.get('rates', {})
        for y in range(1, 11):
            ws10.cell(row=r, column=1 + y, value=rates.get(y, 0)).number_format = '0.00%'
        ws10.cell(row=r, column=12, value=curve_data.get('terminal', 0.01)).number_format = '0.00%'
        r += 1

    r += 1
    ws10.cell(row=r, column=1, value='Virtu Aliases:').font = bf
    r += 1
    for alias, target in DECAY_CURVE_ALIASES.items():
        ws10.cell(row=r, column=1, value=alias)
        ws10.cell(row=r, column=2, value=f'→ {target}')
        r += 1

    ws10.column_dimensions['A'].width = 22
    for c in range(2, 14):
        ws10.column_dimensions[get_column_letter(c)].width = 10

    # ====================================================================
    # Sheet 11: SOFR
    # ====================================================================
    ws11 = wb.create_sheet('SOFR')

    r = 1
    _section_hdr(ws11, r, 'SOFR Forward Curve', 4, dk_blue)
    r += 1

    sofr_curve = config.get('sofr_curve', [])
    if sofr_curve:
        _fill_row(ws11, r, 1, ['Date', 'Rate', 'Floor', 'Effective Rate'], font=bf, fill=gray)
        r += 1
        floor = config.get('sofr_floor', 0.02)
        spread = config.get('sofr_spread', 0.0275)
        for entry in sofr_curve:
            ws11.cell(row=r, column=1, value=entry.get('date', ''))
            rate = entry.get('rate', 0)
            ws11.cell(row=r, column=2, value=rate).number_format = pf2
            ws11.cell(row=r, column=3, value=floor).number_format = pf2
            ws11.cell(row=r, column=4, value=max(rate, floor) + spread).number_format = pf2
            r += 1
    else:
        ws11.cell(row=r, column=1, value='Flat SOFR Rate').font = bf
        ws11.cell(row=r, column=2, value=config.get('sofr_rate', 0.045)).number_format = pf2
        r += 1
        ws11.cell(row=r, column=1, value='Floor').font = bf
        ws11.cell(row=r, column=2, value=config.get('sofr_floor', 0.02)).number_format = pf2
        r += 1
        ws11.cell(row=r, column=1, value='Spread').font = bf
        ws11.cell(row=r, column=2, value=config.get('sofr_spread', 0.0275)).number_format = pf2
        r += 1
        eff = max(config.get('sofr_rate', 0.045), config.get('sofr_floor', 0.02)) + config.get('sofr_spread', 0.0275)
        ws11.cell(row=r, column=1, value='Effective Interest Rate').font = bf
        ws11.cell(row=r, column=2, value=eff).number_format = pf2

    for c, w in enumerate([16, 12, 10, 16], 1):
        ws11.column_dimensions[get_column_letter(c)].width = w

    # ====================================================================
    # Sheet 12: [Payor]_Model (one per payor)
    # ====================================================================
    for code, pp_data in per_payor.items():
        name = pp_data.get('name', code)
        sheet_name = f"{name}_Model"[:31]
        wsm = wb.create_sheet(sheet_name)

        r = 1
        _section_hdr(wsm, r, f'{name} — Model', n_years + 8, dk_blue)
        r += 1

        # Summary data
        wsm.cell(row=r, column=1, value='WA Release Date').font = bf
        wa_info = wa_per_payor.get(code, {})
        wsm.cell(row=r, column=2, value=wa_info.get('wa_date', ''))
        r += 1

        # Waterfall
        p_yt = pp_data.get('year_totals', [])
        _fill_row(wsm, r, 1, [''] + [f"Yr {yt['year']}" for yt in p_yt], font=bf, fill=gray)
        r += 1
        for lbl, fld in wf_fields:
            wsm.cell(row=r, column=1, value=lbl).font = bf
            for ci, yt in enumerate(p_yt, 2):
                wsm.cell(row=r, column=ci, value=yt.get(fld, 0)).number_format = mf
            r += 1

        r += 1
        # Assumptions
        wsm.cell(row=r, column=1, value='Assumptions').font = hdr
        wsm.cell(row=r, column=1).fill = gray
        r += 1
        pc = config.get('payor_configs', {}).get(code, {})
        # Format LTM window as YYYY-MM
        ltm_s = pp_data.get('ltm_start')
        ltm_e = pp_data.get('ltm_end')
        ltm_window_str = ''
        if ltm_s and ltm_e:
            ltm_window_str = f"{ltm_s // 100}-{ltm_s % 100:02d} to {ltm_e // 100}-{ltm_e % 100:02d}"
        assumptions_rows = [
            ('Payor', name),
            ('Income Rights', pp_data.get('income_rights', '')),
            ('Deal Type', config.get('deal_type', 'Catalog')),
            ('LTM Window', ltm_window_str),
            ('FX Currency', pc.get('fx_currency', config.get('base_currency', 'USD'))),
            ('FX Rate', pc.get('fx_rate', 1.0)),
            ('Fee Rate', pc.get('fee_rate', '')),
        ]
        # Synergy assumptions
        if pc.get('synergy'):
            syn_fee = pc.get('synergy_new_fee_rate', config.get('new_fee_rate', ''))
            syn_start = pc.get('synergy_start_year', config.get('synergy_start_year', ''))
            syn_ramp = pc.get('synergy_ramp_months', config.get('synergy_ramp_months', ''))
            assumptions_rows.append(('Synergy Fee Rate', syn_fee))
            assumptions_rows.append(('Synergy Start Year', syn_start))
            assumptions_rows.append(('Synergy Ramp (months)', syn_ramp))
        for lbl, val in assumptions_rows:
            wsm.cell(row=r, column=1, value=lbl).font = bf
            wsm.cell(row=r, column=2, value=val)
            r += 1

        r += 1
        # Per-ISRC rows
        _section_hdr(wsm, r, 'Per-ISRC Detail', n_years + 8, blue)
        r += 1
        _fill_row(wsm, r, 1, ['ISRC', 'Title', 'Artist', 'Curve', '% LTM',
                               'Release Date', 'Label Share', '3P Share'] +
                  [f"Yr {y}" for y in range(1, horizon + 1)],
                  font=bf, fill=gray)
        r += 1

        payor_isrcs = pp_data.get('isrcs', [])
        for isrc_id in payor_isrcs:
            ip = isrc_proj.get(isrc_id, {})
            b = ip.get('baseline', {})
            wsm.cell(row=r, column=1, value=isrc_id)
            wsm.cell(row=r, column=2, value=ip.get('title', ''))
            wsm.cell(row=r, column=3, value=ip.get('artist', ''))
            wsm.cell(row=r, column=4, value=ip.get('genre', ''))
            ltm_g_isrc = b.get('gross', 0)
            pct = ltm_g_isrc / pp_data.get('ltm_gross', 1) * 100 if pp_data.get('ltm_gross', 0) > 0 else 0
            wsm.cell(row=r, column=5, value=round(pct, 2)).number_format = '0.00%'
            wsm.cell(row=r, column=6, value=ip.get('release_date', ''))
            ls = ip.get('label_share')
            wsm.cell(row=r, column=7, value=ls if ls is not None else '').number_format = pf if ls is not None else '@'
            wsm.cell(row=r, column=8, value=ip.get('tp_ratio', 0)).number_format = pf
            for y in range(1, horizon + 1):
                projs = ip.get('projections', [])
                if y - 1 < len(projs):
                    wsm.cell(row=r, column=8 + y, value=projs[y - 1].get('gross', 0)).number_format = mf
            r += 1

        wsm.column_dimensions['A'].width = 16
        wsm.column_dimensions['B'].width = 28
        wsm.column_dimensions['C'].width = 18
        for c in range(4, horizon + 9):
            wsm.column_dimensions[get_column_letter(c)].width = 14

    # ====================================================================
    # Sheet 13: M&A Summary Grid
    # ====================================================================
    ws13 = wb.create_sheet('M&A Summary Grid')

    _section_hdr(ws13, 1, 'M&A Summary Grid', 14, dk_blue)
    ma_headers = ['Close Date', 'Financials Date', 'Name', 'Deal Type', 'Transaction Structure',
                  'Upfront', 'Deferred', 'Total/TEV', 'LTM Gross', 'LTM Net', 'NTM Gross', 'NTM Net',
                  'xLTM Net', 'xNTM Net']
    _fill_row(ws13, 2, 1, ma_headers, font=bf, fill=gray)

    # Single row for this deal
    r = 3
    ntm_data = result.get('ntm', {})
    ws13.cell(row=r, column=1, value=config.get('close_date', ''))
    ws13.cell(row=r, column=2, value=result.get('forecast_start', ''))
    ws13.cell(row=r, column=3, value=config.get('opportunity_name', deal_name))
    ws13.cell(row=r, column=4, value=config.get('deal_type', 'Catalog'))
    ws13.cell(row=r, column=5, value='Asset Purchase')
    ws13.cell(row=r, column=6, value=closing).number_format = mf
    ws13.cell(row=r, column=7, value=hb).number_format = mf
    ws13.cell(row=r, column=8, value=pp).number_format = mf
    ws13.cell(row=r, column=9, value=ltm_g).number_format = mf
    ws13.cell(row=r, column=10, value=ltm_n).number_format = mf
    ws13.cell(row=r, column=11, value=ntm_data.get('gross', 0)).number_format = mf
    ws13.cell(row=r, column=12, value=ntm_data.get('net_incl', 0)).number_format = mf
    ws13.cell(row=r, column=13, value=round(pp / ltm_n, 1) if ltm_n > 0 else '').number_format = mxf if ltm_n > 0 else '@'
    ntm_n_val = ntm_data.get('net_incl', 0)
    ws13.cell(row=r, column=14, value=round(pp / ntm_n_val, 1) if ntm_n_val > 0 else '').number_format = mxf if ntm_n_val > 0 else '@'

    for c in range(1, 15):
        ws13.column_dimensions[get_column_letter(c)].width = 14

    # Save
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    log.info("Forecast Excel saved: %s (%d sheets)", output_path, len(wb.sheetnames))
    return output_path
