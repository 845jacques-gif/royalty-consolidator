"""
Royalty Statement Consolidator - Multi-Payor Edition
Ingests royalty statements from any payor (PDF, CSV, Excel),
auto-detects columns, consolidates, and populates a multi-tab financial model.
"""

import argparse
import calendar
import json
import os
import re
import shutil
import sys
import time
from dataclasses import dataclass, field
from datetime import date
from typing import Dict, List, Optional
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

import logging
import pandas as pd
import openpyxl

log = logging.getLogger('royalty')

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False


# ---------------------------------------------------------------------------
# MusicBrainz ISRC Lookup (free, no credentials required)
# ---------------------------------------------------------------------------

_isrc_cache = {}  # ISRC -> {'release_date': ..., 'track_name': ..., 'artist_name': ...}
_isrc_cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'isrc_cache.json')

# Lazy DB module reference
_db_mod = None

def _db():
    """Lazy-load db module and check availability."""
    global _db_mod
    if _db_mod is None:
        try:
            import db as _d
            _db_mod = _d
        except ImportError:
            return None
    return _db_mod if _db_mod.is_available() else None


def _load_isrc_cache():
    """Load ISRC lookup cache from DB first, then disk."""
    global _isrc_cache
    dbm = _db()
    if dbm:
        try:
            _isrc_cache = dbm.load_full_isrc_cache_db()
            if _isrc_cache:
                return
        except Exception as e:
            log.debug("DB isrc_cache load failed: %s", e)
    if os.path.exists(_isrc_cache_path):
        try:
            with open(_isrc_cache_path, 'r') as f:
                _isrc_cache = json.load(f)
        except (json.JSONDecodeError, IOError):
            _isrc_cache = {}


def _save_isrc_cache():
    """Save ISRC lookup cache to DB + disk."""
    dbm = _db()
    if dbm:
        try:
            dbm.save_isrc_cache_db(_isrc_cache)
        except Exception as e:
            log.debug("DB isrc_cache save failed: %s", e)
    try:
        with open(_isrc_cache_path, 'w') as f:
            json.dump(_isrc_cache, f, indent=2)
    except IOError as e:
        log.warning("Failed to save isrc_cache.json: %s", e)


def lookup_isrc_musicbrainz(isrc: str) -> dict:
    """Look up a track by ISRC on MusicBrainz. Returns dict with release_date, track_name, artist_name."""
    global _isrc_cache

    isrc = isrc.strip().upper()

    # Check cache first
    if isrc in _isrc_cache:
        return _isrc_cache[isrc]

    url = f"https://musicbrainz.org/ws/2/recording?query=isrc:{isrc}&fmt=json"
    req = Request(url, headers={"User-Agent": "RoyaltyConsolidator/1.0 (contact@example.com)"})

    result = {'release_date': '', 'track_name': '', 'artist_name': ''}
    try:
        with urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())

        recordings = data.get("recordings", [])
        if recordings:
            r = recordings[0]
            artist_credit = r.get("artist-credit", [{}])
            artist_name = artist_credit[0].get("name", "") if artist_credit else ""
            result = {
                'release_date': r.get("first-release-date", ""),
                'track_name': r.get("title", ""),
                'artist_name': artist_name,
            }
    except (HTTPError, URLError) as e:
        if hasattr(e, 'code') and e.code == 503:
            time.sleep(2)  # Back off on rate limit
        else:
            log.debug("MusicBrainz HTTP error for %s: %s", isrc, e)
    except Exception as e:
        log.debug("MusicBrainz lookup error for %s: %s", isrc, e)

    # Cache result (even empty ones to avoid re-lookups)
    _isrc_cache[isrc] = result
    return result


def lookup_isrcs_batch(isrcs: List[str], progress_callback=None) -> Dict[str, dict]:
    """Look up multiple ISRCs via MusicBrainz with rate limiting (1 req/sec).
    Returns {isrc: {release_date, track_name, artist_name}}."""
    _load_isrc_cache()

    results = {}
    new_lookups = 0
    not_found = []

    for i, isrc in enumerate(isrcs):
        if not isrc or not isinstance(isrc, str):
            continue

        isrc = isrc.strip().upper()

        # Check if already cached
        if isrc in _isrc_cache:
            results[isrc] = _isrc_cache[isrc]
        else:
            result = lookup_isrc_musicbrainz(isrc)
            results[isrc] = result
            new_lookups += 1

            if not result.get('release_date'):
                not_found.append(isrc)

            # MusicBrainz rate limit: 1 request per second
            time.sleep(1.1)

        if progress_callback and (i + 1) % 20 == 0:
            progress_callback(i + 1, len(isrcs))

    if new_lookups > 0:
        _save_isrc_cache()

    # Report ISRCs without release dates
    if not_found:
        log.warning("[MusicBrainz] %d ISRCs missing release dates: %s", len(not_found),
                    ', '.join(not_found[:10]) + ('...' if len(not_found) > 10 else ''))

    return results


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

STATEMENT_TYPES = {
    'masters': 'Masters',
    'publishing': 'Publishing',
    'neighboring': 'Neighboring Rights',
    'pro': 'PRO (Performance)',
    'sync': 'Sync',
    'other': 'Other',
}


@dataclass
class PayorConfig:
    """Configuration for a single payor."""
    code: str              # Short code: B1, B2, RJ — maps to {code}_Model tab
    name: str              # Display name: "Believe 15%", "RecordJet"
    statements_dir: str    # Directory containing statement files
    fmt: str               # 'auto'
    fee: float             # Distribution fee as decimal (0.15 = 15%)
    source_currency: str = 'USD'
    fx_rate: float = 1.0   # Kept for backward compat; always 1.0 (conversion at dashboard time)
    statement_type: str = 'masters'  # masters, publishing, neighboring, pro, sync, other
    deal_type: str = 'artist'                  # 'artist' or 'label' — whose perspective the earnings are from
    artist_split: Optional[float] = None       # Split % — your share after distro fees (e.g. 50 = you keep 50%)
    calc_payable: bool = False                 # Toggle: calculate payable amount from %
    payable_pct: float = 0.0                   # Payable share percentage
    calc_third_party: bool = False             # Toggle: calculate third party amount from %
    third_party_pct: float = 0.0               # Third party share percentage
    matching_right: Optional[bool] = None      # Whether payor has matching right
    contract_term: Optional[str] = None        # e.g. "3 years", "Life of copyright"
    territory: Optional[str] = None            # e.g. "Worldwide", "North America"
    advance: Optional[float] = None            # Advance amount in deal currency
    contract_pdf_path: Optional[str] = None    # Path to uploaded contract PDF
    contract_summary: Optional[Dict] = None    # Gemini-extracted contract summary
    expected_start: Optional[int] = None       # Expected first period YYYYMM (for missing month detection)
    expected_end: Optional[int] = None         # Expected last period YYYYMM (for missing month detection)
    gcs_files: Optional[List[dict]] = None     # [{name, gcs_path}] — stream from GCS instead of local dir


@dataclass
class PayorResult:
    """Aggregated data for one payor."""
    config: PayorConfig
    isrc_meta: pd.DataFrame    # Unique ISRCs with title, artist, total_gross
    monthly: pd.DataFrame      # ISRC × period with gross, net, sales
    detail: pd.DataFrame       # ISRC × period × store × media_type
    pivot_gross: pd.DataFrame  # Pivot: rows=ISRC, cols=period, vals=gross
    by_store: pd.DataFrame
    file_count: int
    detected_currencies: List[str] = field(default_factory=list)  # Currencies found in files
    file_inventory: List[dict] = field(default_factory=list)  # Per-file metadata
    quality_warnings: List[str] = field(default_factory=list)  # Data quality warnings


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def period_to_end_of_month(period):
    """Convert a YYYYMM int to the last day of that month as mm/dd/yyyy string."""
    year = int(str(period)[:4])
    month = int(str(period)[4:6])
    last_day = calendar.monthrange(year, month)[1]
    return f"{month:02d}/{last_day:02d}/{year}"


def month_to_column(year, month):
    """Convert calendar year+month to the model tab column index.

    Monthly columns start at column 29 (AC) = Jan 2020.
    """
    offset = (year - 2020) * 12 + (month - 1)
    return 29 + offset


_MONTH_NAMES = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}


def parse_period_from_filename(filename):
    """Extract YYYYMM period from filename."""
    # Match patterns like: "PLYGRND 202201", "2022-01", "2022_01", "202201"
    m = re.search(r'(\d{4})[\s._-]?(\d{2})', filename)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        if 2000 <= year <= 2099 and 1 <= month <= 12:
            return year * 100 + month

    # Match M_DD_YYYY / MM_DD_YYYY / MM/DD/YYYY full date patterns:
    # "1_31_2022", "10_31_2022", "4_30-2024", "12/31/2023"
    m_full = re.search(r'(\d{1,2})[\s._/-](\d{1,2})[\s._/-](\d{4})', filename)
    if m_full:
        month, day, year = int(m_full.group(1)), int(m_full.group(2)), int(m_full.group(3))
        if 2000 <= year <= 2099 and 1 <= month <= 12 and 1 <= day <= 31:
            return year * 100 + month

    # Match MM-YY / MM.YY patterns: "12-23 ADA", "03-23", "06_24"
    # Negative lookahead avoids matching "1_31" in "1_31_2022" (day followed by year)
    m_mmyy = re.search(r'(?:^|[^\d])(\d{1,2})[\s._-](\d{2})(?![_.\-/]\d)', filename)
    if m_mmyy:
        month, year_short = int(m_mmyy.group(1)), int(m_mmyy.group(2))
        if 1 <= month <= 12 and 20 <= year_short <= 99:
            return (2000 + year_short) * 100 + month

    # Match text month + year: "January 2022", "Nov 2025"
    m2 = re.search(r'([A-Za-z]{3,})\s*(\d{4})', filename)
    if m2:
        month_str = m2.group(1)[:3].lower()
        year = int(m2.group(2))
        if month_str in _MONTH_NAMES and 2000 <= year <= 2099:
            return year * 100 + _MONTH_NAMES[month_str]

    # Match text month + 2-digit year: "November 24", "Jan 25"
    m3 = re.search(r'([A-Za-z]{3,})\s*(\d{2})(?:\D|$)', filename)
    if m3:
        month_str = m3.group(1)[:3].lower()
        year_short = int(m3.group(2))
        if month_str in _MONTH_NAMES and 20 <= year_short <= 99:
            return (2000 + year_short) * 100 + _MONTH_NAMES[month_str]

    return None


def parse_period_from_path(filepath, statements_dir):
    """Extract YYYYMM period from filepath, checking filename then each parent folder name."""
    filename = os.path.basename(filepath)
    period = parse_period_from_filename(filename)
    if period:
        return period, 'filename'

    # Walk parent folders between the file and the statements root
    rel = os.path.relpath(filepath, statements_dir)
    parts = os.path.normpath(rel).split(os.sep)
    # Check folder parts from innermost to outermost (skip the filename itself)
    for part in reversed(parts[:-1]):
        period = parse_period_from_filename(part)
        if period:
            return period, f'folder:{part}'

    return None, None


# ---------------------------------------------------------------------------
# Universal column auto-detection
# ---------------------------------------------------------------------------

# Each key is the standard schema field; the list has lowercase substrings to
# match against the source column name (checked in order — first match wins).
COLUMN_PATTERNS = {
    'identifier': ['isrc', 'identifier', 'track id', 'asset id', 'recording id', 'track code'],
    'iswc': ['iswc', 'international standard musical work code'],
    'upc': ['upc', 'ean', 'product code', 'barcode', 'catalog'],
    'other_identifier': ['custom id', 'internal id', 'client id', 'deal id', 'release id'],
    'title': ['track title', 'song name', 'track name', 'title', 'track', 'song'],
    'artist': ['track artist', 'artist', 'performer', 'act', 'band', 'creator'],
    'product_title': ['product title', 'album', 'release', 'product', 'bundle',
                      'album name', 'release title'],
    'store': ['distributor', 'store', 'platform', 'dsp', 'service',
              'retailer', 'partner', 'channel', 'source'],
    'sales': ['quantity', 'units', 'streams', 'plays', 'count',
              'qty', 'volume', 'number of', 'total units', 'sales'],
    'media_type': ['download type', 'usage type', 'transaction type',
                   'content type', 'sale type', 'revenue type', 'type',
                   'media type', 'delivery type'],
    'gross': ['gross revenue account', 'gross_revenue_account', 'gross account',
              'gross', 'ppu total', 'revenue', 'earning gross', 'total revenue',
              'gross revenue', 'amount', 'total amount',
              'earning_gross', 'price', 'retail'],
    'net': ['net share account', 'net_share_account', 'net account',
            'net', 'royalty', 'earning net', 'earnings', 'payout', 'net revenue',
            'your share', 'payable', 'earning_net', 'net amount', 'royalties'],
    'fees': ['fee', 'commission', 'distribution fee', 'service fee', 'deduction'],
    'period': ['period', 'reporting period', 'statement period', 'statement date',
               'sale period', 'accounting period', 'royalty period', 'month'],
    'country': ['country', 'territory', 'region', 'market'],
    'release_date': ['release date', 'release_date', 'date of release', 'first release'],
}


CURRENCY_HINTS = {
    'EUR': ['eur', 'euro', '€'],
    'USD': ['usd', 'dollar', '$', 'us '],
    'GBP': ['gbp', 'pound', '£', 'sterling'],
    'CAD': ['cad', 'canadian'],
    'AUD': ['aud', 'australian'],
    'JPY': ['jpy', 'yen', '¥'],
    'SEK': ['sek', 'krona'],
    'NOK': ['nok'],
    'DKK': ['dkk'],
    'CHF': ['chf', 'franc'],
    'BRL': ['brl', 'real'],
}


def _detect_currency(df_columns):
    """Detect currency from column names. Returns currency code or 'Unknown'."""
    all_cols = ' '.join(str(c).lower() for c in df_columns)
    for code, hints in CURRENCY_HINTS.items():
        for hint in hints:
            if hint in all_cols:
                return code
    return 'Unknown'


def _fuzzy_match_columns(df_columns):
    """Match source column names to our standard schema.

    Returns a dict mapping standard field -> source column name.
    """
    mapped = {}
    used = set()
    lower_cols = {c: c.strip().lower() for c in df_columns}

    # Columns that look like percentages — never match as revenue fields
    pct_cols = {orig for orig, low in lower_cols.items()
                if low.endswith('%') or low.endswith('pct') or low.endswith('percent')
                or 'royalty%' in low or 'royalty %' in low}

    # Disqualifiers: prevent cross-matching ambiguous terms like "earnings"
    # e.g. "Net Earnings" should not match gross, "Gross Revenue" should not match net
    # Also: prefer "account currency" columns over "sale currency" columns (FUGA/Harbour)
    has_account_gross = any('account' in low and 'gross' in low for low in lower_cols.values())
    has_account_net = any('account' in low and 'net' in low for low in lower_cols.values())
    field_disqualifiers = {
        'gross': ('net',) + (('sale',) if has_account_gross else ()),
        'net': ('gross',) + (('sale',) if has_account_net else ()),
    }

    for field, patterns in COLUMN_PATTERNS.items():
        disqualifiers = field_disqualifiers.get(field, ())
        for pattern in patterns:
            for orig, low in lower_cols.items():
                if orig in used:
                    continue
                # Skip percentage columns for revenue fields (gross/net)
                if field in ('gross', 'net') and orig in pct_cols:
                    continue
                if pattern == low or pattern in low:
                    # Skip if column contains a disqualifier word for this field
                    if disqualifiers and any(dq in low for dq in disqualifiers):
                        continue
                    mapped[field] = orig
                    used.add(orig)
                    break
            if field in mapped:
                break

    return mapped


# ---------------------------------------------------------------------------
# Sample-based column validation
# ---------------------------------------------------------------------------

# Patterns for validating detected column content
_ISRC_RE = re.compile(r'^[A-Z]{2}[A-Z0-9]{3}\d{7}$')
_UPC_RE = re.compile(r'^\d{12,13}$')
_ISWC_RE = re.compile(r'^T\d{9,10}$')
_COUNTRY_RE = re.compile(r'^[A-Z]{2}$')


def _validate_column_detection(df: pd.DataFrame, mapping: dict, n_sample: int = 10) -> List[str]:
    """Validate auto-detected column mappings by sampling data rows.

    Checks that values in mapped columns match expected formats.
    Returns a list of warning strings for suspicious mappings.
    """
    warnings = []
    if len(df) == 0:
        return warnings

    sample = df.head(n_sample)

    validators = {
        'identifier': (_ISRC_RE, 'ISRC format (e.g. USRC12345678)'),
        'upc': (_UPC_RE, 'UPC format (12-13 digits)'),
        'iswc': (_ISWC_RE, 'ISWC format (e.g. T0123456789)'),
        'country': (_COUNTRY_RE, '2-letter country code'),
    }

    numeric_fields = {'gross', 'net', 'fees', 'sales'}

    for field_name, src_col in mapping.items():
        if src_col not in sample.columns:
            continue

        col_data = sample[src_col].dropna().astype(str).str.strip()
        if len(col_data) == 0:
            continue

        # Check format validators
        if field_name in validators:
            pattern, expected = validators[field_name]
            match_count = col_data.apply(lambda v: bool(pattern.match(v))).sum()
            match_pct = match_count / len(col_data) * 100
            if match_pct < 30 and len(col_data) >= 3:
                warnings.append(
                    f'Column "{src_col}" mapped to {field_name} — only {match_pct:.0f}% of sampled '
                    f'values match expected {expected}'
                )

        # Check numeric fields actually contain numbers
        if field_name in numeric_fields:
            try:
                numeric_vals = pd.to_numeric(col_data.str.replace(r'[$€£¥,]', '', regex=True), errors='coerce')
                non_null = numeric_vals.notna().sum()
                if non_null / len(col_data) < 0.5 and len(col_data) >= 3:
                    warnings.append(
                        f'Column "{src_col}" mapped to {field_name} — '
                        f'only {non_null}/{len(col_data)} sampled values are numeric'
                    )
            except Exception:
                pass

    return warnings


# ---------------------------------------------------------------------------
# Duplicate row detection (across files)
# ---------------------------------------------------------------------------

def detect_duplicate_rows(detail_df: pd.DataFrame) -> List[str]:
    """Flag rows that appear identical across different files.

    Checks for rows with the same ISRC + period + gross amount,
    which may indicate duplicate data imported from overlapping statements.
    Returns a list of warning strings.
    """
    warnings = []
    if detail_df is None or len(detail_df) == 0:
        return warnings

    key_cols = ['identifier', 'period']
    if not all(c in detail_df.columns for c in key_cols + ['gross']):
        return warnings

    # Group by ISRC + period and look for exact gross duplicates
    grouped = detail_df.groupby(key_cols)['gross'].agg(['sum', 'count']).reset_index()
    dups = grouped[grouped['count'] > 1]

    if len(dups) > 0:
        # Check if any have exact same gross (likely true duplicates vs legitimate splits)
        dup_detail = detail_df.merge(dups[key_cols], on=key_cols, how='inner')
        exact_dups = (
            dup_detail.groupby(key_cols + ['gross'])
            .size()
            .reset_index(name='n')
        )
        exact_dups = exact_dups[exact_dups['n'] > 1]

        if len(exact_dups) > 0:
            n_rows = exact_dups['n'].sum() - len(exact_dups)  # extra duplicate rows
            sample_isrcs = exact_dups['identifier'].head(3).tolist()
            warnings.append(
                f'{n_rows} potential duplicate rows detected (same ISRC + period + gross amount). '
                f'Sample ISRCs: {", ".join(str(i) for i in sample_isrcs)}'
            )

    return warnings


# ---------------------------------------------------------------------------
# Cross-file ISRC validation
# ---------------------------------------------------------------------------

def validate_cross_payor_isrcs(payor_results: Dict[str, 'PayorResult']) -> List[str]:
    """Warn when an ISRC appears across payors with conflicting title/artist metadata.

    Returns a list of warning strings.
    """
    warnings = []
    if len(payor_results) < 2:
        return warnings

    # Collect title/artist per ISRC across all payors
    isrc_info = {}  # {isrc: [(payor_code, title, artist), ...]}
    for code, pr in payor_results.items():
        if pr.isrc_meta is None or len(pr.isrc_meta) == 0:
            continue
        for _, row in pr.isrc_meta.iterrows():
            isrc = str(row.get('identifier', '')).strip()
            if not isrc:
                continue
            title = str(row.get('title', '')).strip().lower()
            artist = str(row.get('artist', '')).strip().lower()
            if isrc not in isrc_info:
                isrc_info[isrc] = []
            isrc_info[isrc].append((code, title, artist))

    conflicts = []
    for isrc, entries in isrc_info.items():
        if len(entries) < 2:
            continue
        # Check for conflicting titles or artists
        titles = set(t for _, t, _ in entries if t)
        artists = set(a for _, _, a in entries if a)
        if len(titles) > 1 or len(artists) > 1:
            payor_codes = [e[0] for e in entries]
            conflicts.append((isrc, payor_codes, titles, artists))

    if conflicts:
        n = len(conflicts)
        samples = conflicts[:3]
        details = []
        for isrc, payors, titles, artists in samples:
            details.append(f'{isrc} ({", ".join(payors)})')
        warnings.append(
            f'{n} ISRCs have conflicting title/artist metadata across payors. '
            f'Examples: {"; ".join(details)}'
        )

    return warnings


def _find_header_row(filepath, ext, max_scan=50):
    """Scan the first rows of a CSV/Excel file for the real header row.

    Returns the 0-based row index of the best header candidate, or 0 if
    row 0 already looks like a header.
    """
    keywords = set()
    for patterns in COLUMN_PATTERNS.values():
        for p in patterns:
            keywords.add(p)

    if ext == '.csv':
        # Read raw lines to avoid pandas column-count inference issues
        import csv, io
        rows = []
        try:
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                for i, line in enumerate(f):
                    if i >= max_scan:
                        break
                    try:
                        parsed = next(csv.reader(io.StringIO(line)))
                        rows.append(parsed)
                    except Exception as e:
                        log.debug("CSV line parse fallback in %s row %d: %s", filepath, i, e)
                        rows.append([line.strip()])
        except Exception as e:
            log.debug("_find_header_row CSV open failed for %s: %s", filepath, e)
            return 0
    else:
        try:
            preview = pd.read_excel(filepath, sheet_name=0, header=None,
                                    nrows=max_scan, dtype=str)
            rows = []
            for idx in range(len(preview)):
                row_vals = [str(v).strip() for v in preview.iloc[idx] if pd.notna(v)]
                rows.append(row_vals)
        except Exception as e:
            log.debug("_find_header_row Excel read failed for %s: %s", filepath, e)
            return 0

    best_row = 0
    best_hits = 0
    for idx, row_vals in enumerate(rows):
        hits = 0
        for val in row_vals:
            val_low = val.strip().lower()
            for kw in keywords:
                if kw == val_low or kw in val_low:
                    hits += 1
                    break
        if hits > best_hits:
            best_hits = hits
            best_row = idx

    return best_row if best_hits >= 3 else 0


CSV_CHUNKSIZE = 100_000  # rows per chunk for large CSV reads


def _read_raw_dataframe(filepath, filename):
    """Read any file (PDF, CSV, Excel) into a raw DataFrame.
    Large CSVs are read in chunks to limit peak memory usage.
    """
    ext = os.path.splitext(filename)[1].lower()

    if ext == '.csv':
        df = None
        # Try common encodings
        for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
            try:
                chunks = []
                for chunk in pd.read_csv(filepath, encoding=enc, chunksize=CSV_CHUNKSIZE,
                                         low_memory=False):
                    chunks.append(chunk)
                if chunks:
                    df = pd.concat(chunks, ignore_index=True)
                    del chunks
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
            except Exception as e:
                log.debug("CSV read failed for %s (enc=%s): %s", filename, enc, e)
                break
        # Check if we got useful columns; if not, scan for the real header row
        if df is None or len(_fuzzy_match_columns(df.columns)) < 2:
            header_row = _find_header_row(filepath, ext)
            if header_row > 0:
                for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
                    try:
                        chunks = []
                        for chunk in pd.read_csv(filepath, skiprows=header_row, encoding=enc,
                                                  chunksize=CSV_CHUNKSIZE, low_memory=False):
                            chunks.append(chunk)
                        if chunks:
                            df = pd.concat(chunks, ignore_index=True)
                            del chunks
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
                    except Exception as e:
                        log.debug("CSV read with skiprows failed for %s (enc=%s): %s", filename, enc, e)
                        break
        return df

    elif ext in ('.xlsx', '.xls', '.xlsb'):
        # Auto-pick the best sheet (e.g. "Digital Sales" over "Summary")
        engine = 'pyxlsb' if ext == '.xlsb' else None
        try:
            xls = pd.ExcelFile(filepath, engine=engine)
            from mapper import _pick_best_sheet
            best_sheet = _pick_best_sheet(xls.sheet_names)
        except Exception as e:
            log.debug("Excel sheet pick failed for %s: %s", filename, e)
            best_sheet = 0
        try:
            df = pd.read_excel(filepath, sheet_name=best_sheet, engine=engine)
        except Exception as e:
            log.debug("Excel read failed for %s (sheet=%s): %s", filename, best_sheet, e)
            df = None
        if df is None or len(_fuzzy_match_columns(df.columns)) < 2:
            header_row = _find_header_row(filepath, ext)
            if header_row > 0:
                try:
                    df = pd.read_excel(filepath, sheet_name=best_sheet, skiprows=header_row, engine=engine)
                except Exception as e:
                    log.debug("Excel read with skiprows failed for %s: %s", filename, e)
        return df

    elif ext == '.pdf':
        if not HAS_PDF:
            log.warning("pdfplumber not installed, skipping %s", filename)
            return None
        frames = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(table[0])]
                    rows = table[1:]
                    frames.append(pd.DataFrame(rows, columns=header))
        if not frames:
            return None
        return pd.concat(frames, ignore_index=True)

    return None


def parse_file_universal(filepath, filename, fmt='auto', fallback_period=None):
    """Parse any statement file into the standard schema.

    Returns (DataFrame, detected_currency) or (None, None).
    fmt='auto' uses column auto-detection.
    fallback_period: optional YYYYMM int to use when auto-detection fails.
    """
    df = _read_raw_dataframe(filepath, filename)
    if df is None or df.empty:
        return None, None

    df.columns = [str(c).strip() for c in df.columns]
    detected_currency = _detect_currency(df.columns)

    # --- Auto-detect columns ---
    col_map = _fuzzy_match_columns(df.columns)

    if 'identifier' not in col_map and 'gross' not in col_map:
        cols_found = list(df.columns)
        log.warning("Could not detect columns in %s. Found: %s", filename, cols_found[:15])
        return None, None

    # Sample-based validation of detected columns
    col_warnings = _validate_column_detection(df, col_map)
    for w in col_warnings:
        log.warning("[%s] %s", filename, w)

    # Derive period
    if 'period' in col_map:
        raw_period = df[col_map['period']]

        def _parse_period_value(val):
            s = str(val).strip()
            # Try YYYYMM from digits only (e.g. "202201", "2022-01")
            digits = re.sub(r'[^0-9]', '', s)
            if len(digits) >= 6 and digits[:6].isdigit():
                p = int(digits[:6])
                if 1 <= p % 100 <= 12:
                    return p
            # Try text month + year (e.g. "Nov 2025", "January 2022")
            _MONTH_MAP = {
                'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
            }
            m = re.search(r'([A-Za-z]{3,})\s*(\d{4})', s)
            if m:
                month_str = m.group(1)[:3].lower()
                year = int(m.group(2))
                if month_str in _MONTH_MAP and 2000 <= year <= 2099:
                    return year * 100 + _MONTH_MAP[month_str]
            # Try year + text month (e.g. "2025 November")
            m2 = re.search(r'(\d{4})\s*([A-Za-z]{3,})', s)
            if m2:
                year = int(m2.group(1))
                month_str = m2.group(2)[:3].lower()
                if month_str in _MONTH_MAP and 2000 <= year <= 2099:
                    return year * 100 + _MONTH_MAP[month_str]
            return 0

        df['_period'] = raw_period.apply(_parse_period_value)
    else:
        period = parse_period_from_filename(filename)
        if period is None:
            # Try parent folder names for period hints
            parent = os.path.dirname(filepath)
            while parent and len(parent) > 3:
                folder_name = os.path.basename(parent)
                period = parse_period_from_filename(folder_name)
                if period:
                    break
                parent = os.path.dirname(parent)
        if period is None:
            # Try peeking inside file for date columns
            try:
                period = peek_statement_date(filepath, filename)
            except Exception as e:
                log.debug("peek_statement_date failed for %s: %s", filename, e)
        if period is None:
            # Try to find a date-like value in any column
            for c in df.columns:
                sample = str(df[c].dropna().iloc[0]) if len(df[c].dropna()) > 0 else ''
                m = re.search(r'(\d{4})[\s._/-]?(\d{2})', sample)
                if m:
                    period = int(f"{m.group(1)}{m.group(2)}")
                    break
        if period is None and fallback_period:
            period = fallback_period
        if period is None:
            log.warning("No period found for %s, skipping", filename)
            return None
        df['_period'] = period

    n = len(df)

    def _get(field, default_val=''):
        if field in col_map:
            return df[col_map[field]]
        return pd.Series([default_val] * n)

    def _get_numeric(field):
        if field in col_map:
            col = df[col_map[field]]
            # Strip currency symbols and thousands separators before converting
            if not pd.api.types.is_numeric_dtype(col):
                col = col.astype(str).str.replace(r'[$€£¥,]', '', regex=True).str.strip()
            return pd.to_numeric(col, errors='coerce').fillna(0)
        return pd.Series([0.0] * n)

    gross_vals = _get_numeric('gross')
    net_vals = _get_numeric('net')
    fees_vals = _get_numeric('fees')

    # If fees not directly available, compute as gross - net (when both are nonzero)
    if 'fees' not in col_map and gross_vals.abs().sum() > 0 and net_vals.abs().sum() > 0:
        fees_vals = gross_vals - net_vals

    result = pd.DataFrame({
        'identifier': _get('identifier'),
        'iswc': _get('iswc'),
        'upc': _get('upc'),
        'other_identifier': _get('other_identifier'),
        'title': _get('title'),
        'artist': _get('artist'),
        'product_title': _get('product_title'),
        'store': _get('store'),
        'media_type': _get('media_type'),
        'period': df['_period'],
        'gross': gross_vals,
        'net': net_vals,
        'fees': fees_vals,
        'sales': _get_numeric('sales'),
        'country': _get('country'),
        'release_date': _get('release_date'),
    })

    # Capture KEEP columns (unmapped source columns, prefixed with KEEP_)
    mapped_src_cols = set(col_map.values())
    for src_col in df.columns:
        if src_col.startswith('_'):
            continue
        if src_col not in mapped_src_cols:
            keep_name = f'KEEP_{src_col}'
            result[keep_name] = df[src_col]

    return result, detected_currency


# ---------------------------------------------------------------------------
# Phase 2: Parse with explicit user mapping
# ---------------------------------------------------------------------------

def parse_file_with_mapping(filepath, filename, column_mapping, remove_top=0,
                            remove_bottom=0, header_row=None, sheet=None,
                            keep_columns=None, fallback_period=None):
    """Parse a file using an explicit user-provided column mapping (Phase 2).

    column_mapping: {source_col_name: canonical_field_name} from the mapping UI
    keep_columns: list of source column names the user toggled as KEEP
    Returns (DataFrame, detected_currency) or (None, None).
    """
    df = _read_raw_dataframe(filepath, filename)
    if df is None or df.empty:
        return None, None

    df.columns = [str(c).strip() for c in df.columns]
    detected_currency = _detect_currency(df.columns)

    # Apply header row offset if specified — use skiprows to avoid C parser
    # errors from inconsistent field counts in rows before the header
    if header_row is not None and header_row > 0:
        skip = list(range(header_row))
        ext = os.path.splitext(filepath)[1].lower()
        if ext in ('.xlsx', '.xls', '.xlsb'):
            engine = 'pyxlsb' if ext == '.xlsb' else None
            df = pd.read_excel(filepath, sheet_name=sheet or 0, header=header_row, dtype=str, engine=engine)
        elif ext == '.csv':
            for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
                try:
                    df = pd.read_csv(filepath, header=0, skiprows=skip,
                                     dtype=str, encoding=enc)
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
                except pd.errors.ParserError:
                    try:
                        df = pd.read_csv(filepath, header=0, skiprows=skip,
                                         dtype=str, encoding=enc, on_bad_lines='skip')
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
        df.columns = [str(c).strip() for c in df.columns]

    # Remove top/bottom rows
    if remove_top > 0 and remove_top < len(df):
        df = df.iloc[remove_top:]
    if remove_bottom > 0 and remove_bottom < len(df):
        df = df.iloc[:-remove_bottom]
    df = df.reset_index(drop=True)

    if df.empty:
        return None, None

    # Build reverse mapping: canonical -> source_col
    canonical_to_src = {}
    for src_col, canonical in column_mapping.items():
        if canonical and src_col in df.columns:
            canonical_to_src[canonical] = src_col

    n = len(df)

    def _get(field, default_val=''):
        if field in canonical_to_src:
            return df[canonical_to_src[field]]
        return pd.Series([default_val] * n)

    def _get_numeric(field):
        if field in canonical_to_src:
            col = df[canonical_to_src[field]]
            if not pd.api.types.is_numeric_dtype(col):
                col = col.astype(str).str.replace(r'[$€£¥,]', '', regex=True).str.strip()
            return pd.to_numeric(col, errors='coerce').fillna(0)
        return pd.Series([0.0] * n)

    # Handle period
    if 'period' in canonical_to_src:
        raw_period = df[canonical_to_src['period']]

        def _parse_period_value(val):
            s = str(val).strip()
            digits = re.sub(r'[^0-9]', '', s)
            if len(digits) >= 6 and digits[:6].isdigit():
                p = int(digits[:6])
                if 1 <= p % 100 <= 12:
                    return p
            _MONTH_MAP = {
                'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
            }
            m = re.search(r'([A-Za-z]{3,})\s*(\d{4})', s)
            if m:
                month_str = m.group(1)[:3].lower()
                year = int(m.group(2))
                if month_str in _MONTH_MAP and 2000 <= year <= 2099:
                    return year * 100 + _MONTH_MAP[month_str]
            m2 = re.search(r'(\d{4})\s*([A-Za-z]{3,})', s)
            if m2:
                year = int(m2.group(1))
                month_str = m2.group(2)[:3].lower()
                if month_str in _MONTH_MAP and 2000 <= year <= 2099:
                    return year * 100 + _MONTH_MAP[month_str]
            return 0

        period_col = raw_period.apply(_parse_period_value)
    else:
        period = parse_period_from_filename(filename)
        # Try parent folder names
        if period is None:
            parent = os.path.dirname(filepath)
            while parent and len(parent) > 3:
                folder_name = os.path.basename(parent)
                period = parse_period_from_filename(folder_name)
                if period:
                    break
                parent = os.path.dirname(parent)
        # Try peeking inside file content
        if period is None:
            try:
                period = peek_statement_date(filepath, filename)
            except Exception as e:
                log.debug("peek_statement_date failed for %s: %s", filename, e)
                pass
        if period is None and fallback_period:
            period = fallback_period
        if period is None:
            log.warning("No period found for %s, skipping", filename)
            return None, None
        period_col = pd.Series([period] * n)

    gross_vals = _get_numeric('gross')
    net_vals = _get_numeric('net')
    fees_vals = _get_numeric('fees')

    if 'fees' not in canonical_to_src and gross_vals.abs().sum() > 0 and net_vals.abs().sum() > 0:
        fees_vals = gross_vals - net_vals

    result = pd.DataFrame({
        'identifier': _get('identifier'),
        'iswc': _get('iswc'),
        'upc': _get('upc'),
        'other_identifier': _get('other_identifier'),
        'title': _get('title'),
        'artist': _get('artist'),
        'product_title': _get('product_title'),
        'store': _get('store'),
        'media_type': _get('media_type'),
        'period': period_col,
        'gross': gross_vals,
        'net': net_vals,
        'fees': fees_vals,
        'sales': _get_numeric('sales'),
        'country': _get('country'),
        'release_date': _get('release_date'),
    })

    # Fee percent field for waterfall
    if 'fee_pct' in canonical_to_src:
        result['fee_pct'] = _get_numeric('fee_pct')

    # KEEP columns: only ones the user explicitly toggled
    if keep_columns:
        mapped_src_cols = set(canonical_to_src.values())
        for src_col in keep_columns:
            if src_col in df.columns and src_col not in mapped_src_cols:
                keep_name = f'KEEP_{src_col}'
                result[keep_name] = df[src_col]
    else:
        # Default: keep all unmapped columns
        mapped_src_cols = set(canonical_to_src.values())
        for src_col in df.columns:
            if src_col.startswith('_'):
                continue
            if src_col not in mapped_src_cols:
                keep_name = f'KEEP_{src_col}'
                result[keep_name] = df[src_col]

    return result, detected_currency


# ---------------------------------------------------------------------------
# Per-payor loading and aggregation
# ---------------------------------------------------------------------------

def peek_statement_date(filepath: str, filename: str) -> Optional[int]:
    """Peek inside a file to find a statement date/period without full parsing.
    Returns YYYYMM int or None.
    """
    ext = os.path.splitext(filename)[1].lower()
    try:
        if ext == '.csv':
            df = pd.read_csv(filepath, nrows=20, encoding='utf-8', on_bad_lines='skip')
        elif ext in ('.xlsx', '.xls', '.xlsb'):
            engine = 'pyxlsb' if ext == '.xlsb' else None
            df = pd.read_excel(filepath, nrows=20, engine=engine)
        else:
            return None
    except Exception as e:
        log.debug("peek_statement_date read failed for %s: %s", filename, e)
        return None

    if df is None or df.empty:
        return None

    # Normalise column headers
    col_lower = {c: str(c).strip().lower() for c in df.columns}

    # Check for known period/date column names
    period_names = ['period', 'reporting period', 'statement period', 'statement date',
                    'sale period', 'accounting period', 'royalty period', 'month',
                    'reporting_period', 'statement_date', 'sale_period', 'pay period']

    for orig_col, low in col_lower.items():
        if low in period_names:
            # Try to parse first non-null value
            vals = df[orig_col].dropna()
            if vals.empty:
                continue
            sample = str(vals.iloc[0]).strip()
            # Try YYYYMM, YYYY-MM
            m = re.search(r'(\d{4})[\s._/-]?(\d{2})', sample)
            if m:
                yr, mo = int(m.group(1)), int(m.group(2))
                if 2000 <= yr <= 2099 and 1 <= mo <= 12:
                    return yr * 100 + mo
            # Try month name + year
            m2 = re.search(r'([A-Za-z]{3,})\s*(\d{4})', sample)
            if m2:
                ms = m2.group(1)[:3].lower()
                yr = int(m2.group(2))
                if ms in _MONTH_NAMES and 2000 <= yr <= 2099:
                    return yr * 100 + _MONTH_NAMES[ms]
            # Try MM/DD/YYYY or DD/MM/YYYY
            m3 = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', sample)
            if m3:
                a, b, yr = int(m3.group(1)), int(m3.group(2)), int(m3.group(3))
                if 2000 <= yr <= 2099:
                    mo = a if 1 <= a <= 12 else b
                    return yr * 100 + mo

    # Scan all columns for date-like values as last resort
    for orig_col in df.columns:
        vals = df[orig_col].dropna()
        if vals.empty:
            continue
        sample = str(vals.iloc[0]).strip()
        m = re.search(r'(\d{4})[\s._/-](\d{2})', sample)
        if m:
            yr, mo = int(m.group(1)), int(m.group(2))
            if 2000 <= yr <= 2099 and 1 <= mo <= 12:
                return yr * 100 + mo

    return None


def _file_hash(filepath: str) -> str:
    """Compute a fast content hash (SHA-256) for duplicate detection."""
    import hashlib
    h = hashlib.sha256()
    with open(filepath, 'rb') as fh:
        for chunk in iter(lambda: fh.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def _parse_single_file(filepath, filename, config_fmt, column_mappings, fallback_period):
    """Parse a single file — designed to run in a thread pool."""
    if column_mappings and filename in column_mappings:
        cm = column_mappings[filename]
        df, file_currency = parse_file_with_mapping(
            filepath, filename,
            column_mapping=cm.get('mapping', {}),
            remove_top=cm.get('remove_top', 0),
            remove_bottom=cm.get('remove_bottom', 0),
            header_row=cm.get('header_row'),
            sheet=cm.get('sheet'),
            keep_columns=cm.get('keep_columns'),
            fallback_period=fallback_period,
        )
    else:
        df, file_currency = parse_file_universal(filepath, filename, fmt=config_fmt,
                                                  fallback_period=fallback_period)
    return df, file_currency


def load_payor_statements(config: PayorConfig, file_dates: Optional[Dict[str, str]] = None,
                          column_mappings: Optional[Dict[str, dict]] = None,
                          progress_cb=None) -> Optional[PayorResult]:
    """Load and aggregate all statement files for one payor.

    file_dates: optional {filename: "MM/DD/YY"} from the date extraction modal,
                used as fallback when auto-detection fails to find a period.
    column_mappings: optional {filename: {mapping_dict, remove_top, remove_bottom, header_row, sheet, keep_columns}}
                     When provided for a file, uses parse_file_with_mapping() instead of parse_file_universal().
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import time as _time
    _t_start = _time.time()
    gcs_mode = bool(config.gcs_files)
    log.info("[%s] Loading %s (%s, %d files)", config.code, config.name,
             'GCS' if gcs_mode else config.statements_dir,
             len(config.gcs_files) if gcs_mode else 0)

    detail_chunks = []  # pre-aggregated detail per file (much smaller than raw)
    meta_chunks = []
    total_raw_rows = 0
    file_count = 0
    currencies_seen = set()
    file_inventory = []
    fee_detected = config.fee > 0  # Skip auto-detect if user provided a fee
    seen_hashes = {}  # hash -> filename (for duplicate detection)

    SUPPORTED_EXT = ('.csv', '.xlsx', '.xls', '.xlsb', '.pdf')

    # ---- Phase 1: Discover files and deduplicate ----
    # file_tasks: (filepath_or_none, filename, rel_folder, path_period, path_source, fallback_period, gcs_path_or_none)
    file_tasks = []

    if gcs_mode:
        # GCS mode: files are in cloud storage, will be downloaded one-at-a-time during processing.
        # Zip files are downloaded, extracted to temp, and inner files added as tasks.
        import tempfile as _tmpmod
        try:
            import storage as _st_discover
        except ImportError:
            _st_discover = None

        for entry in config.gcs_files:
            f = entry.get('name', '')
            gcs_path = entry.get('gcs_path', '')
            if not f or not gcs_path:
                continue
            if f.startswith('~$'):
                continue
            ext = os.path.splitext(f)[1].lower()

            # Handle zip files: download, extract, add inner files as individual tasks
            if ext == '.zip' and _st_discover:
                fd, tmp_zip = _tmpmod.mkstemp(suffix='.zip')
                os.close(fd)
                try:
                    _st_discover.download_to_file(gcs_path, tmp_zip)
                    import zipfile
                    with zipfile.ZipFile(tmp_zip, 'r') as zf:
                        tmp_dir = _tmpmod.mkdtemp()
                        zf.extractall(tmp_dir)
                        for inner_name in sorted(zf.namelist()):
                            if inner_name.endswith('/') or inner_name.startswith('__MACOSX'):
                                continue
                            inner_ext = os.path.splitext(inner_name)[1].lower()
                            if inner_ext not in SUPPORTED_EXT:
                                continue
                            inner_path = os.path.join(tmp_dir, inner_name)
                            base_name = os.path.basename(inner_name)
                            pp = parse_period_from_filename(base_name)
                            ps = 'filename' if pp else None
                            fbp = None
                            if file_dates and base_name in file_dates:
                                try:
                                    ds = file_dates[base_name].split('/')
                                    if len(ds) == 3:
                                        yyyy = 2000 + int(ds[2]) if int(ds[2]) < 100 else int(ds[2])
                                        fbp = yyyy * 100 + int(ds[0])
                                except (ValueError, IndexError):
                                    pass
                            # Local path (already extracted), no GCS path
                            file_tasks.append((inner_path, base_name, '', pp, ps, fbp, None))
                    # Keep GCS blob for potential reprocessing (lifecycle rule handles cleanup)
                except Exception as e:
                    log.error("GCS zip extraction failed for %s: %s", f, e)
                finally:
                    try:
                        os.remove(tmp_zip)
                    except OSError:
                        pass
                continue

            if ext not in SUPPORTED_EXT:
                continue

            # Period detection from filename only (no full path in GCS)
            path_period = parse_period_from_filename(f)
            path_source = 'filename' if path_period else None

            fallback_period = None
            if file_dates and f in file_dates:
                try:
                    date_str = file_dates[f]
                    parts = date_str.split('/')
                    if len(parts) == 3:
                        mm, dd, yy = int(parts[0]), int(parts[1]), int(parts[2])
                        yyyy = 2000 + yy if yy < 100 else yy
                        fallback_period = yyyy * 100 + mm
                except (ValueError, IndexError):
                    pass

            file_tasks.append((None, f, '', path_period, path_source, fallback_period, gcs_path))
    else:
        # Local mode: walk filesystem
        for root, dirs, files in os.walk(config.statements_dir):
            for f in sorted(files):
                if f.startswith('~$'):
                    continue
                ext = os.path.splitext(f)[1].lower()
                if ext not in SUPPORTED_EXT:
                    continue
                filepath = os.path.join(root, f)

                # ---- Duplicate file detection ----
                try:
                    fhash = _file_hash(filepath)
                except OSError:
                    fhash = None
                if fhash and fhash in seen_hashes:
                    rel_folder = os.path.relpath(root, config.statements_dir)
                    if rel_folder == '.':
                        rel_folder = ''
                    file_inventory.append({
                        'filename': f,
                        'folder': rel_folder,
                        'period': None,
                        'period_source': 'none',
                        'rows': 0,
                        'status': 'duplicate',
                        'gross': 0,
                        'duplicate_of': seen_hashes[fhash],
                    })
                    log.warning("  Duplicate file skipped: %s (same as %s)", f, seen_hashes[fhash])
                    continue
                if fhash:
                    seen_hashes[fhash] = f

                # Detect period from filepath (filename + folder names)
                path_period, path_source = parse_period_from_path(filepath, config.statements_dir)
                rel_folder = os.path.relpath(root, config.statements_dir)
                if rel_folder == '.':
                    rel_folder = ''

                # Check file_dates for a fallback period
                fallback_period = None
                if file_dates and f in file_dates:
                    try:
                        date_str = file_dates[f]
                        parts = date_str.split('/')
                        if len(parts) == 3:
                            mm, dd, yy = int(parts[0]), int(parts[1]), int(parts[2])
                            yyyy = 2000 + yy if yy < 100 else yy
                            fallback_period = yyyy * 100 + mm
                    except (ValueError, IndexError):
                        pass

                file_tasks.append((filepath, f, rel_folder, path_period, path_source, fallback_period, None))

    # ---- Phase 2+3: Parse and process files with bounded memory ----
    # Parse files in small batches (BATCH_SIZE at a time in parallel),
    # then immediately process + pre-aggregate each batch before starting
    # the next. This caps peak memory at ~BATCH_SIZE raw DataFrames.
    import gc
    _t_parse = _time.time()
    log.info("[%s] Phase 1 (discover): %d files in %.1fs", config.code, len(file_tasks), _t_parse - _t_start)

    PARSE_BATCH = 8  # files per batch — balances memory vs throughput

    # Lazy-import storage for GCS download-on-demand
    _storage = None
    if gcs_mode:
        try:
            import storage as _storage
            import tempfile as _tmpmod
        except ImportError:
            log.error("storage module not available for GCS mode")
            return None

    for batch_start in range(0, len(file_tasks), PARSE_BATCH):
        batch_end = min(batch_start + PARSE_BATCH, len(file_tasks))
        batch = file_tasks[batch_start:batch_end]

        # For GCS files, download this batch in parallel threads
        temp_paths = {}  # j -> temp_path (to clean up after processing)
        if gcs_mode:
            def _download_one(j, fn, gcs_path):
                ext = os.path.splitext(fn)[1]
                fd, tmp = _tmpmod.mkstemp(suffix=ext)
                os.close(fd)
                try:
                    _storage.download_to_file(gcs_path, tmp)
                    return j, tmp
                except Exception as e:
                    log.error("  GCS download failed for %s: %s", fn, e)
                    try:
                        os.remove(tmp)
                    except OSError:
                        pass
                    return j, None

            with ThreadPoolExecutor(max_workers=PARSE_BATCH) as dl_pool:
                dl_futures = {}
                for j, (_, fn, _, _, _, _, gcs_path) in enumerate(batch):
                    if gcs_path:
                        dl_futures[dl_pool.submit(_download_one, j, fn, gcs_path)] = j
                for fut in as_completed(dl_futures):
                    j, tmp = fut.result()
                    temp_paths[j] = tmp

        # Parse this batch in parallel (download + parse overlap for local files)
        n_workers = min(len(batch), PARSE_BATCH)
        parsed_batch = [None] * len(batch)

        def _do_parse(idx, fp, fn, fbp):
            df, fc = _parse_single_file(fp, fn, config.fmt, column_mappings, fbp)
            return idx, df, fc

        with ThreadPoolExecutor(max_workers=n_workers) as pool:
            futures = {}
            for j, (fp, fn, _rf, _pp, _ps, fbp, _gp) in enumerate(batch):
                # Use temp path for GCS files, original path for local
                actual_path = temp_paths.get(j, fp) if gcs_mode else fp
                if actual_path is None:
                    parsed_batch[j] = (None, None)
                    continue
                futures[pool.submit(_do_parse, j, actual_path, fn, fbp)] = j
            for future in as_completed(futures):
                try:
                    idx, df, fc = future.result()
                    parsed_batch[idx] = (df, fc)
                except Exception as e:
                    idx = futures[future]
                    log.error("  Failed to parse %s: %s", batch[idx][1], e)
                    parsed_batch[idx] = (None, None)

        # Process each file in this batch sequentially, then free it
        for j, (filepath, f, rel_folder, path_period, path_source, fallback_period, gcs_path) in enumerate(batch):
            df, file_currency = parsed_batch[j]
            parsed_batch[j] = None  # free immediately

            # Clean up temp file immediately after parsing (before pre-agg)
            if j in temp_paths and temp_paths[j]:
                try:
                    os.remove(temp_paths[j])
                except OSError:
                    pass

            if df is None:
                file_inventory.append({
                    'filename': f,
                    'folder': rel_folder,
                    'period': path_period,
                    'period_source': path_source or 'none',
                    'rows': 0,
                    'status': 'skipped',
                    'gross': 0,
                })
                log.warning("Could not parse %s, skipping", f)
                continue

            if file_currency and file_currency != 'Unknown':
                currencies_seen.add(file_currency)

            # Store original gross for FX Original column
            df['gross_original'] = df['gross'].copy()

            # Auto-detect source currency if user chose "auto"
            if config.source_currency == 'auto' and file_currency and file_currency != 'Unknown':
                config.source_currency = file_currency

            # Auto-detect fee from data when fee is 0 and both gross & net exist
            if not fee_detected and config.fee == 0 and df['gross'].abs().sum() > 0 and df['net'].abs().sum() > 0:
                total_g = df['gross'].abs().sum()
                total_n = df['net'].abs().sum()
                config.fee = round((total_g - total_n) / total_g, 4)
                fee_detected = True

            # Fallback: if net is all zeros but gross has data, derive net from fee
            if df['net'].abs().sum() == 0 and df['gross'].abs().sum() > 0 and config.fee > 0:
                df['net'] = df['gross'] * (1 - config.fee)
                df['fees'] = df['gross'] * config.fee

            # Reverse: if gross is all zeros but net has data, derive gross from fee
            if df['gross'].abs().sum() == 0 and df['net'].abs().sum() > 0:
                if config.fee > 0:
                    df['gross'] = df['net'] / (1 - config.fee)
                    df['fees'] = df['gross'] - df['net']
                else:
                    df['gross'] = df['net']

            # Drop rows with missing ISRC
            df['identifier'] = df['identifier'].astype(str).str.strip()
            df = df[df['identifier'].ne('') & df['identifier'].ne('nan') & df['identifier'].notna()]

            # Determine actual periods found in the parsed data
            data_periods = sorted(df['period'].unique().tolist()) if len(df) > 0 else []
            file_gross = float(df['gross'].sum()) if len(df) > 0 else 0

            file_inventory.append({
                'filename': f,
                'folder': rel_folder,
                'period': data_periods[0] if len(data_periods) == 1 else path_period,
                'periods': [int(p) for p in data_periods],
                'period_source': 'data' if data_periods else (path_source or 'none'),
                'rows': len(df),
                'status': 'ok',
                'gross': round(file_gross, 2),
            })

            log.info("  [%d/%d] %s (%s rows)", batch_start + j + 1, len(file_tasks), f, f"{len(df):,}")

            # Ensure new columns exist with defaults
            for col in ['iswc', 'upc', 'other_identifier', 'country', 'release_date']:
                if col not in df.columns:
                    df[col] = ''
            for col in ['fees', 'gross_original']:
                if col not in df.columns:
                    df[col] = 0.0

            # Metadata per ISRC (first occurrence)
            meta = (
                df.groupby('identifier')
                .agg({'title': 'first', 'artist': 'first', 'product_title': 'first',
                      'iswc': 'first', 'upc': 'first', 'release_date': 'first'})
                .reset_index()
            )
            meta_chunks.append(meta)

            # ---- Pre-aggregate to detail level immediately ----
            # 200K raw rows → ~10-50K detail rows per file
            keep_cols = [c for c in df.columns if c.startswith('KEEP_')]
            file_agg = {
                'title': 'first', 'artist': 'first', 'product_title': 'first',
                'iswc': 'first', 'upc': 'first', 'other_identifier': 'first',
                'release_date': 'first',
                'gross': 'sum', 'net': 'sum', 'fees': 'sum', 'sales': 'sum',
                'gross_original': 'sum',
            }
            for kc in keep_cols:
                file_agg[kc] = 'first'
            detail_chunk = (
                df.groupby(['identifier', 'period', 'store', 'media_type', 'country'])
                .agg(file_agg)
                .reset_index()
            )
            total_raw_rows += len(df)
            del df  # free raw rows immediately
            detail_chunks.append(detail_chunk)
            file_count += 1

            if progress_cb:
                try:
                    progress_cb(file_count, f)
                except Exception:
                    pass

        # Keep GCS blobs for potential reprocessing (lifecycle rule handles cleanup)

        # Force gc after each batch to reclaim freed raw DataFrames + temp files
        del parsed_batch
        temp_paths.clear()
        gc.collect()

    _t_agg = _time.time()
    log.info("[%s] Phase 2+3 (parse + process): %d files in %.1fs", config.code, file_count, _t_agg - _t_parse)

    if not detail_chunks:
        log.warning("No files found for %s", config.name)
        return None

    log.info("Aggregating %d files (%s raw rows → %s pre-aggregated rows)...",
             file_count, f"{total_raw_rows:,}",
             f"{sum(len(d) for d in detail_chunks):,}")

    # Merge pre-aggregated chunks — re-aggregate to handle overlapping keys
    # (same ISRC+period+store+media_type+country across files)
    combined_detail = pd.concat(detail_chunks, ignore_index=True)
    del detail_chunks

    # Detect all KEEP_ columns across files (different files may have different ones)
    keep_cols = [c for c in combined_detail.columns if c.startswith('KEEP_')]
    detail_agg_dict = {
        'title': 'first', 'artist': 'first', 'product_title': 'first',
        'iswc': 'first', 'upc': 'first', 'other_identifier': 'first',
        'release_date': 'first',
        'gross': 'sum', 'net': 'sum', 'fees': 'sum', 'sales': 'sum',
        'gross_original': 'sum',
    }
    for kc in keep_cols:
        detail_agg_dict[kc] = 'first'

    detail = (
        combined_detail.groupby(['identifier', 'period', 'store', 'media_type', 'country'])
        .agg(detail_agg_dict)
        .reset_index()
    )
    detail['statement_date'] = detail['period'].apply(period_to_end_of_month)

    # Monthly: derive from detail (already aggregated, so just roll up)
    monthly = (
        detail.groupby(['identifier', 'period'])
        .agg({'gross': 'sum', 'net': 'sum', 'fees': 'sum', 'sales': 'sum',
              'gross_original': 'sum'})
        .reset_index()
    )
    monthly['statement_date'] = monthly['period'].apply(period_to_end_of_month)

    # Store aggregate: derive from detail
    by_store = (
        detail.groupby('store')
        .agg({'gross': 'sum', 'net': 'sum', 'sales': 'sum'})
        .reset_index()
        .sort_values('gross', ascending=False)
    )
    by_store.columns = ['Store', 'Total Gross', 'Total Net', 'Total Sales']
    del combined_detail

    all_meta = pd.concat(meta_chunks, ignore_index=True)
    del meta_chunks
    isrc_meta = all_meta.drop_duplicates('identifier', keep='first').reset_index(drop=True)
    del all_meta

    # Pivot: rows=ISRC, cols=period, vals=gross
    pivot_gross = monthly.pivot_table(
        index='identifier', columns='period', values='gross', aggfunc='sum', fill_value=0
    )

    # Total gross per ISRC
    isrc_meta['total_gross'] = isrc_meta['identifier'].map(pivot_gross.sum(axis=1))
    isrc_meta = isrc_meta.sort_values('total_gross', ascending=False).reset_index(drop=True)

    _t_done = _time.time()
    log.info("[%s] Phase 4 (final agg): %.1fs", config.code, _t_done - _t_agg)

    currency_str = ', '.join(sorted(currencies_seen)) if currencies_seen else config.source_currency
    log.info("%s: %d files, %s ISRCs, $%s total gross, currency: %s — total %.1fs",
             config.name, file_count, f"{len(isrc_meta):,}",
             f"{isrc_meta['total_gross'].sum():,.2f}", currency_str,
             _t_done - _t_start)

    # Duplicate row detection across files
    quality_warnings = detect_duplicate_rows(detail)
    for w in quality_warnings:
        log.warning("[%s] %s", config.code, w)

    return PayorResult(
        config=config,
        isrc_meta=isrc_meta,
        monthly=monthly,
        detail=detail,
        pivot_gross=pivot_gross,
        by_store=by_store,
        file_count=file_count,
        detected_currencies=sorted(currencies_seen) if currencies_seen else [config.source_currency],
        file_inventory=file_inventory,
        quality_warnings=quality_warnings,
    )


def load_all_payors(configs: List[PayorConfig], file_dates: Optional[Dict[str, str]] = None,
                    column_mappings_by_payor: Optional[Dict[str, Dict[str, dict]]] = None,
                    progress_cb=None) -> Dict[str, PayorResult]:
    """Load statements for all payors (parallel when multiple payors).

    file_dates: optional {filename: "MM/DD/YY"} from the date extraction modal,
                used as fallback when auto-detection fails.
    column_mappings_by_payor: optional {payor_code: {filename: mapping_info}}
                              Phase 2 explicit column mappings per payor.
    progress_cb: optional callback(files_done, current_filename) for UI progress.
    """
    if len(configs) <= 1:
        # Single payor — no need for threading overhead
        results = {}
        for cfg in configs:
            cm = column_mappings_by_payor.get(cfg.code) if column_mappings_by_payor else None
            result = load_payor_statements(cfg, file_dates=file_dates, column_mappings=cm,
                                           progress_cb=progress_cb)
            if result is not None:
                results[cfg.code] = result
    else:
        # Multiple payors — parse in parallel
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def _load_one(cfg):
            cm = column_mappings_by_payor.get(cfg.code) if column_mappings_by_payor else None
            return cfg.code, load_payor_statements(cfg, file_dates=file_dates, column_mappings=cm,
                                                   progress_cb=progress_cb)

        results = {}
        with ThreadPoolExecutor(max_workers=min(len(configs), 4)) as pool:
            futures = {pool.submit(_load_one, cfg): cfg.code for cfg in configs}
            for future in as_completed(futures):
                try:
                    code, result = future.result()
                    if result is not None:
                        results[code] = result
                except Exception as e:
                    log.error("Payor %s failed: %s", futures[future], e)

    # Cross-payor ISRC validation
    cross_warnings = validate_cross_payor_isrcs(results)
    if cross_warnings:
        for w in cross_warnings:
            log.warning("[cross-payor] %s", w)
        # Attach cross-payor warnings to all payors
        for pr in results.values():
            pr.quality_warnings.extend(cross_warnings)

    return results


# ---------------------------------------------------------------------------
# Consolidated Excel output (all payors combined)
# ---------------------------------------------------------------------------

def _build_detail_23col(pr: 'PayorResult', deal_name: str = '',
                        formulas: Optional[Dict[str, str]] = None) -> pd.DataFrame:
    """Build the 23+ column consolidated detail DataFrame for one payor.

    Output columns in order:
        Statement Date, Royalty Type, Payor, ISRC, ISWC, UPC, Other Identifier,
        Title, Artist, Release Date, Source, Deal, Media Type, Territory,
        FX Original, Units, Gross Earnings, Fees, Net Receipts, Payable Share,
        Third Party Share, Net Earnings, + any KEEP_ columns

    formulas: optional {field: "=expression"} from Phase 2 waterfall calc step.
    """
    d = pr.detail.copy()
    cfg = pr.config
    n = len(d)

    # Helper to safely get column or default
    def _col(name, default=''):
        return d[name] if name in d.columns else pd.Series([default] * n, dtype=object)

    net_vals = d['net'] if 'net' in d.columns else pd.Series([0.0] * n)

    # Share calculation: driven by PayorConfig toggles (set on upload page)
    if cfg.calc_payable and cfg.payable_pct > 0:
        payable_share = net_vals * (cfg.payable_pct / 100.0)
    else:
        # Use the configured artist split %
        split_pct = cfg.artist_split if cfg.artist_split is not None else 100.0
        payable_share = net_vals * (split_pct / 100.0)

    if cfg.calc_third_party and cfg.third_party_pct > 0:
        third_party_share = net_vals * (cfg.third_party_pct / 100.0)
    else:
        third_party_share = net_vals - payable_share

    net_earnings = payable_share - third_party_share

    out = pd.DataFrame({
        'Statement Date': d['statement_date'] if 'statement_date' in d.columns else '',
        'Royalty Type': STATEMENT_TYPES.get(cfg.statement_type, cfg.statement_type),
        'Payor': cfg.name,
        'ISRC': _col('identifier'),
        'ISWC': _col('iswc'),
        'UPC': _col('upc'),
        'Other Identifier': _col('other_identifier'),
        'Title': _col('title'),
        'Artist': _col('artist'),
        'Release Date': _col('release_date'),
        'Release Date Source': _col('release_date_source') if 'release_date_source' in d.columns else '',
        'Source': _col('store'),
        'Deal': deal_name or '',
        'Media Type': _col('media_type'),
        'Territory': d['country'] if 'country' in d.columns and d['country'].astype(str).str.strip().ne('').any()
                     else (cfg.territory or ''),
        'FX Original': d['gross_original'] if 'gross_original' in d.columns else d.get('gross', 0),
        'Units': d['sales'] if 'sales' in d.columns else 0,
        'Gross Earnings': d['gross'] if 'gross' in d.columns else 0,
        'Fees': d['fees'] if 'fees' in d.columns else 0,
        'Net Receipts': net_vals,
        'Payable Share': payable_share,
        'Third Party Share': third_party_share,
        'Net Earnings': net_earnings,
    })

    # Apply user formulas from Phase 2 waterfall calc step
    if formulas:
        try:
            from formula_engine import apply_formulas
            out, _errors = apply_formulas(out, formulas)
        except ImportError:
            pass

    # Append KEEP columns
    for col in d.columns:
        if col.startswith('KEEP_'):
            out[col] = d[col]

    return out


def aggregate_detail(df: pd.DataFrame, group_by_cols: List[str]) -> pd.DataFrame:
    """Aggregate consolidated detail by groupby cols, summing numeric fields."""
    numeric_cols = ['Units', 'Gross Earnings', 'Fees', 'Net Receipts',
                    'Payable Share', 'Third Party Share', 'Net Earnings', 'FX Original']
    numeric_cols = [c for c in numeric_cols if c in df.columns]
    non_group = [c for c in df.columns if c not in group_by_cols and c not in numeric_cols]

    agg_dict = {}
    for c in numeric_cols:
        agg_dict[c] = 'sum'
    for c in non_group:
        agg_dict[c] = 'first'

    grouped = df.groupby(group_by_cols, dropna=False).agg(agg_dict).reset_index()
    # Restore original column order
    grouped = grouped[[c for c in df.columns if c in grouped.columns]]
    return grouped


EXCEL_MAX_ROWS = 1_048_575  # 1,048,576 minus 1 for header row


def _write_df_to_excel(writer, df, sheet_name):
    """Write a DataFrame to Excel, splitting across numbered sheets if it exceeds the row limit."""
    if len(df) <= EXCEL_MAX_ROWS:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        part = 1
        for start in range(0, len(df), EXCEL_MAX_ROWS):
            chunk = df.iloc[start:start + EXCEL_MAX_ROWS]
            sname = f'{sheet_name}_{part}' if part > 1 else sheet_name
            chunk.to_excel(writer, sheet_name=sname[:31], index=False)
            part += 1


def write_consolidated_excel(payor_results: Dict[str, PayorResult], output_path,
                             deal_name: str = '', formulas: Optional[Dict[str, str]] = None,
                             aggregate_by: Optional[List[str]] = None):
    """Write a single consolidated Excel with data from all payors (23+ column schema)."""
    log.info("Writing consolidated data to: %s", output_path)

    all_clean = []
    all_summary = []
    all_monthly = []
    all_isrc_meta = []

    for code, pr in payor_results.items():
        payor_name = pr.config.name

        # Build 23-column detail
        clean = _build_detail_23col(pr, deal_name=deal_name, formulas=formulas)
        all_clean.append(clean)

        # ISRC + month summary
        summary = pr.monthly.merge(
            pr.isrc_meta[['identifier', 'title', 'artist']], on='identifier', how='left'
        )
        summary_cols = ['identifier', 'title', 'artist', 'period', 'statement_date',
                        'gross', 'net', 'fees', 'sales']
        summary_cols = [c for c in summary_cols if c in summary.columns]
        summary = summary[summary_cols]
        col_rename = {
            'identifier': 'ISRC', 'title': 'Title', 'artist': 'Artist',
            'period': 'Period', 'statement_date': 'Statement Date',
            'gross': 'Gross', 'net': 'Net', 'fees': 'Fees', 'sales': 'Units',
        }
        summary = summary.rename(columns=col_rename)
        summary.insert(0, 'Payor', payor_name)
        all_summary.append(summary)

        # Monthly totals
        mt_agg = {'gross': 'sum', 'net': 'sum', 'sales': 'sum'}
        if 'fees' in pr.monthly.columns:
            mt_agg['fees'] = 'sum'
        mt = (
            pr.monthly.groupby(['period', 'statement_date'])
            .agg(mt_agg)
            .reset_index()
        )
        mt = mt.rename(columns={'gross': 'Gross', 'net': 'Net', 'fees': 'Fees',
                                'sales': 'Units', 'period': 'Period',
                                'statement_date': 'Statement Date'})
        mt.insert(0, 'Payor', payor_name)
        all_monthly.append(mt)

        # ISRC metadata
        meta_cols = ['identifier', 'title', 'artist', 'product_title', 'total_gross']
        for extra in ['iswc', 'upc', 'release_date']:
            if extra in pr.isrc_meta.columns:
                meta_cols.append(extra)
        meta = pr.isrc_meta[[c for c in meta_cols if c in pr.isrc_meta.columns]].copy()
        meta.insert(0, 'Payor', payor_name)
        all_isrc_meta.append(meta)

    combined_clean = pd.concat(all_clean, ignore_index=True).sort_values(
        ['Statement Date', 'Payor', 'ISRC'])
    if aggregate_by:
        combined_clean = aggregate_detail(combined_clean, aggregate_by)
    combined_summary = pd.concat(all_summary, ignore_index=True).sort_values(
        ['Payor', 'ISRC', 'Period'])
    combined_monthly = pd.concat(all_monthly, ignore_index=True).sort_values(
        ['Period', 'Payor'])
    combined_meta = pd.concat(all_isrc_meta, ignore_index=True)

    # Free intermediate lists to reduce memory pressure
    del all_clean, all_summary, all_monthly, all_isrc_meta

    # Cross-payor top songs (deduped by ISRC, summed across payors)
    meta_agg = {'title': 'first', 'artist': 'first', 'total_gross': 'sum'}
    if 'release_date' in combined_meta.columns:
        meta_agg['release_date'] = 'first'
    cross_payor = (
        combined_meta.groupby('identifier')
        .agg(meta_agg)
        .reset_index()
        .sort_values('total_gross', ascending=False)
    )
    del combined_meta

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        _write_df_to_excel(writer, combined_clean, 'Consolidated')
        del combined_clean
        _write_df_to_excel(writer, combined_summary, 'By ISRC-Month')
        del combined_summary
        combined_monthly.to_excel(writer, sheet_name='Monthly Totals', index=False)
        del combined_monthly

        # Per-payor store breakdown
        for code, pr in payor_results.items():
            sheet_name = f'Stores_{code}'
            pr.by_store.to_excel(writer, sheet_name=sheet_name, index=False)

        # Top songs across all payors — use release_date from enrichment/isrc_meta (no MusicBrainz call)
        top = cross_payor.head(50).copy()
        top.columns = ['ISRC', 'Title', 'Artist', 'Total Gross'] + (['Release Date'] if 'release_date' in cross_payor.columns else [])
        top.insert(0, 'Rank', range(1, len(top) + 1))
        top.to_excel(writer, sheet_name='Top 50 Songs', index=False)

        # Full ISRC catalog
        catalog = cross_payor.copy()
        catalog.columns = ['ISRC', 'Title', 'Artist', 'Total Gross'] + (['Release Date'] if 'release_date' in cross_payor.columns else [])
        catalog.to_excel(writer, sheet_name='ISRC Catalog', index=False)

    log.info("Done writing consolidated Excel: %s", output_path)


def write_consolidated_csv(payor_results: Dict[str, PayorResult], output_path,
                           deal_name: str = '', formulas: Optional[Dict[str, str]] = None,
                           aggregate_by: Optional[List[str]] = None):
    """Write consolidated detail as a single CSV file (23+ column schema)."""
    log.info("Writing consolidated CSV to: %s", output_path)

    all_clean = []
    for code, pr in payor_results.items():
        clean = _build_detail_23col(pr, deal_name=deal_name, formulas=formulas)
        all_clean.append(clean)

    combined = pd.concat(all_clean, ignore_index=True).sort_values(
        ['Statement Date', 'Payor', 'ISRC'])
    if aggregate_by:
        combined = aggregate_detail(combined, aggregate_by)
    combined.to_csv(output_path, index=False)
    log.info("Done. %s rows", f"{len(combined):,}")
    return combined


def write_per_payor_exports(payor_results: Dict[str, PayorResult], output_dir: str,
                            deal_name: str = '', formulas: Optional[Dict[str, str]] = None,
                            aggregate_by: Optional[List[str]] = None):
    """Write one consolidated Excel file per payor. Returns dict of code -> path."""
    os.makedirs(output_dir, exist_ok=True)
    paths = {}
    for code, pr in payor_results.items():
        safe_name = pr.config.name.replace(' ', '_').replace('/', '_')
        path = os.path.join(output_dir, f'{safe_name}_consolidated.xlsx')
        log.info("Writing %s -> %s", pr.config.name, path)

        clean = _build_detail_23col(pr, deal_name=deal_name, formulas=formulas)
        if aggregate_by:
            clean = aggregate_detail(clean, aggregate_by)

        summary = pr.monthly.merge(
            pr.isrc_meta[['identifier', 'title', 'artist']], on='identifier', how='left'
        )
        summary_cols = ['identifier', 'title', 'artist', 'period', 'statement_date',
                        'gross', 'net', 'fees', 'sales']
        summary_cols = [c for c in summary_cols if c in summary.columns]
        summary = summary[summary_cols]
        summary = summary.rename(columns={
            'identifier': 'ISRC', 'title': 'Title', 'artist': 'Artist',
            'period': 'Period', 'statement_date': 'Statement Date',
            'gross': 'Gross', 'net': 'Net', 'fees': 'Fees', 'sales': 'Units',
        })

        mt_agg = {'gross': 'sum', 'net': 'sum', 'sales': 'sum'}
        if 'fees' in pr.monthly.columns:
            mt_agg['fees'] = 'sum'
        monthly_totals = (
            pr.monthly.groupby(['period', 'statement_date'])
            .agg(mt_agg)
            .reset_index()
        )
        monthly_totals = monthly_totals.rename(columns={
            'gross': 'Gross', 'net': 'Net', 'fees': 'Fees',
            'sales': 'Units', 'period': 'Period',
            'statement_date': 'Statement Date',
        })

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            _write_df_to_excel(writer, clean, 'Detail')
            _write_df_to_excel(writer, summary, 'By ISRC-Month')
            monthly_totals.to_excel(writer, sheet_name='Monthly Totals', index=False)
            pr.by_store.to_excel(writer, sheet_name='Stores', index=False)

        paths[code] = path
        log.info("%s: %s rows", pr.config.name, f"{len(clean):,}")

    return paths


def write_per_payor_csv_exports(payor_results: Dict[str, PayorResult], output_dir: str,
                                 deal_name: str = '', formulas: Optional[Dict[str, str]] = None,
                                 aggregate_by: Optional[List[str]] = None):
    """Write one consolidated CSV file per payor. Returns dict of code -> path."""
    os.makedirs(output_dir, exist_ok=True)
    paths = {}
    for code, pr in payor_results.items():
        safe_name = pr.config.name.replace(' ', '_').replace('/', '_')
        path = os.path.join(output_dir, f'{safe_name}_consolidated.csv')
        log.info("Writing CSV %s -> %s", pr.config.name, path)

        clean = _build_detail_23col(pr, deal_name=deal_name, formulas=formulas)
        if aggregate_by:
            clean = aggregate_detail(clean, aggregate_by)
        clean.to_csv(path, index=False)
        paths[code] = path
        log.info("%s: %s rows", pr.config.name, f"{len(clean):,}")

    return paths


def apply_enrichment_to_raw_detail(detail_df: pd.DataFrame, lookups: dict) -> pd.DataFrame:
    """Inject enriched release dates into a PayorResult.detail DataFrame.

    Match by ISRC (identifier column) first, then TITLE::ARTIST fallback.
    Sets both 'release_date' and 'release_date_source' columns on the raw detail.
    Vectorized implementation — avoids row-by-row iteration.
    """
    import re as _re

    if not lookups:
        return detail_df

    df = detail_df.copy()

    if 'release_date_source' not in df.columns:
        df['release_date_source'] = ''

    # Build lookup Series keyed by ISRC
    isrc_dates = {}
    isrc_sources = {}
    ta_dates = {}
    ta_sources = {}
    for key, entry in lookups.items():
        rd = entry.get('release_date')
        if not rd:
            continue
        src = entry.get('source', '')
        if '::' in key:
            ta_dates[key] = rd
            ta_sources[key] = src
        else:
            isrc_dates[key] = rd
            isrc_sources[key] = src

    # Vectorized ISRC key column
    has_id = 'identifier' in df.columns
    has_title = 'title' in df.columns
    has_artist = 'artist' in df.columns

    if has_id and isrc_dates:
        isrc_col = df['identifier'].fillna('').astype(str).str.strip().str.upper()
        matched_rd = isrc_col.map(isrc_dates)
        matched_src = isrc_col.map(isrc_sources)
    else:
        matched_rd = pd.Series([None] * len(df), index=df.index)
        matched_src = pd.Series([None] * len(df), index=df.index)

    # Fallback: TITLE::ARTIST for rows without ISRC match
    if has_title and has_artist and ta_dates:
        needs_fallback = matched_rd.isna()
        if needs_fallback.any():
            title_col = df.loc[needs_fallback, 'title'].fillna('').astype(str).str.strip().str.upper().str.replace(r'\s+', ' ', regex=True)
            artist_col = df.loc[needs_fallback, 'artist'].fillna('').astype(str).str.strip().str.upper().str.replace(r'\s+', ' ', regex=True)
            ta_key = title_col + '::' + artist_col
            fb_rd = ta_key.map(ta_dates)
            fb_src = ta_key.map(ta_sources)
            matched_rd.loc[needs_fallback] = matched_rd.loc[needs_fallback].fillna(fb_rd)
            matched_src.loc[needs_fallback] = matched_src.loc[needs_fallback].fillna(fb_src)

    # Apply: only fill release_date where currently empty
    has_match = matched_rd.notna()
    if has_match.any():
        if 'release_date' in df.columns:
            current_rd = df['release_date'].fillna('').astype(str).str.strip()
            empty_rd = current_rd.isin(['', 'nan', 'None', 'NaT'])
            fill_mask = has_match & empty_rd
        else:
            fill_mask = has_match
            df['release_date'] = ''

        df.loc[fill_mask, 'release_date'] = matched_rd.loc[fill_mask]
        df.loc[has_match, 'release_date_source'] = matched_src.loc[has_match]

    return df


# ---------------------------------------------------------------------------
# Populate the financial model template
# ---------------------------------------------------------------------------

def clear_data_rows(ws, start_row, max_col=120):
    """Clear data from start_row to the last used row."""
    for row in range(start_row, ws.max_row + 1):
        for col in range(2, max_col + 1):
            ws.cell(row=row, column=col).value = None


def populate_model_tab(ws, pr: PayorResult, supplemental_meta=None):
    """Populate a single payor model tab with data.

    ws: the worksheet for this payor's model tab
    pr: PayorResult with aggregated data
    supplemental_meta: optional DataFrame with ISRC, release_date, artist_share,
                       third_party_share, reversion_date, license_term
    """
    config = pr.config
    isrc_meta = pr.isrc_meta
    pivot_gross = pr.pivot_gross

    # Set config cells
    ws['D25'] = config.name
    ws['D32'] = config.fee

    # Clear existing data rows (52+)
    clear_data_rows(ws, start_row=52, max_col=120)

    isrcs = isrc_meta['identifier'].tolist()
    periods = sorted(pivot_gross.columns.tolist())

    # Pre-compute column mapping
    col_map = {}
    for period in periods:
        year = int(str(period)[:4])
        month = int(str(period)[4:6])
        col_map[period] = month_to_column(year, month)

    # Build supplemental lookup
    supp = {}
    if supplemental_meta is not None and not supplemental_meta.empty:
        for _, row in supplemental_meta.iterrows():
            isrc = str(row.get('identifier', row.get('isrc', row.get('ISRC', '')))).strip()
            if isrc:
                supp[isrc] = row

    meta_dict = isrc_meta.set_index('identifier').to_dict('index')
    start_row = 52

    for i, isrc in enumerate(isrcs):
        row = start_row + i
        meta = meta_dict[isrc]

        # B: ISRC, C: Title, D: Artist
        ws.cell(row=row, column=2, value=isrc)
        ws.cell(row=row, column=3, value=meta['title'])
        ws.cell(row=row, column=4, value=meta['artist'])

        # Supplemental: G: Release Date, I: Reversion, J: Artist Share, K: 3P Share
        if isrc in supp:
            s = supp[isrc]
            release = s.get('release_date', s.get('Release Date'))
            if pd.notna(release):
                ws.cell(row=row, column=7, value=release)
            reversion = s.get('reversion_date', s.get('Reversion Date', s.get('reversion')))
            if pd.notna(reversion):
                ws.cell(row=row, column=9, value=reversion)
            artist_share = s.get('artist_share', s.get('Artist Share', s.get('Share PLYGRND')))
            if pd.notna(artist_share):
                ws.cell(row=row, column=10, value=float(artist_share))
            tp_share = s.get('third_party_share', s.get('Third Party Splits', s.get('3P Share')))
            if pd.notna(tp_share):
                ws.cell(row=row, column=11, value=float(tp_share))

        # Monthly gross earnings
        if isrc in pivot_gross.index:
            isrc_data = pivot_gross.loc[isrc]
            for period in periods:
                val = isrc_data.get(period, 0)
                if val and val != 0:
                    ws.cell(row=row, column=col_map[period], value=round(float(val), 6))

        if (i + 1) % 200 == 0:
            print(f"      {i + 1}/{len(isrcs)} ISRCs...", flush=True)

    print(f"    {config.code}_Model: {len(isrcs)} ISRCs, "
          f"{len(periods)} months ({min(periods)}-{max(periods)})", flush=True)


def populate_metadata_sheet(ws, payor_results: Dict[str, PayorResult], supplemental_meta=None):
    """Populate the combined Metadata sheet with all unique ISRCs across payors."""
    # Gather all ISRC metadata, deduped
    all_meta = []
    for code, pr in payor_results.items():
        meta = pr.isrc_meta[['identifier', 'title', 'artist', 'product_title']].copy()
        meta['total_gross'] = pr.isrc_meta['total_gross']
        all_meta.append(meta)

    combined = pd.concat(all_meta, ignore_index=True)
    # Keep the version with the highest total_gross for each ISRC
    combined = combined.sort_values('total_gross', ascending=False)
    combined = combined.drop_duplicates('identifier', keep='first').reset_index(drop=True)

    # Build supplemental lookup
    supp = {}
    if supplemental_meta is not None and not supplemental_meta.empty:
        for _, row in supplemental_meta.iterrows():
            isrc = str(row.get('identifier', row.get('isrc', row.get('ISRC', '')))).strip()
            if isrc:
                supp[isrc] = row

    # Clear existing data (row 5+)
    clear_data_rows(ws, start_row=5, max_col=12)

    for i, (_, row) in enumerate(combined.iterrows()):
        r = 5 + i
        isrc = row['identifier']
        ws.cell(row=r, column=2, value=isrc)                   # B: ISRC
        ws.cell(row=r, column=3, value=row['title'])            # C: Track Title
        ws.cell(row=r, column=4, value=row['artist'])           # D: Artists
        ws.cell(row=r, column=6, value=row['product_title'])    # F: Album Name

        if isrc in supp:
            s = supp[isrc]
            release = s.get('release_date', s.get('Release Date'))
            if pd.notna(release):
                ws.cell(row=r, column=5, value=release)         # E: Release Date
            license_term = s.get('license_term', s.get('License Term', s.get('Term')))
            if pd.notna(license_term):
                ws.cell(row=r, column=7, value=license_term)    # G: License Term
            reversion = s.get('reversion_date', s.get('Reversion Date'))
            if pd.notna(reversion):
                ws.cell(row=r, column=8, value=reversion)       # H: Reversion Date
            tp_share = s.get('third_party_share', s.get('Third Party Splits'))
            if pd.notna(tp_share):
                ws.cell(row=r, column=9, value=tp_share)        # I: Third Party Splits
            artist_share = s.get('artist_share', s.get('Artist/Label Share', s.get('Share PLYGRND')))
            if pd.notna(artist_share):
                ws.cell(row=r, column=10, value=artist_share)   # J: Artist/Label Share

    print(f"    Metadata: {len(combined)} unique ISRCs", flush=True)


def populate_template(template_path, output_path, payor_results: Dict[str, PayorResult],
                      supplemental_meta=None):
    """Copy template and populate all payor model tabs + metadata."""
    print(f"\n  Copying template to: {output_path}", flush=True)
    shutil.copy2(template_path, output_path)

    print("  Opening workbook (this may be slow for large templates)...", flush=True)
    wb = openpyxl.load_workbook(output_path)
    sheet_names = wb.sheetnames

    # Populate each payor's model tab
    for code, pr in payor_results.items():
        tab_name = f"{code}_Model"
        if tab_name in sheet_names:
            print(f"  Populating {tab_name}...", flush=True)
            populate_model_tab(wb[tab_name], pr, supplemental_meta)
        else:
            print(f"  WARNING: Tab '{tab_name}' not found in template. "
                  f"Available: {sheet_names}", flush=True)

    # Populate combined Metadata
    if 'Metadata' in sheet_names:
        print("  Populating Metadata...", flush=True)
        populate_metadata_sheet(wb['Metadata'], payor_results, supplemental_meta)

    print("  Saving workbook...", flush=True)
    wb.save(output_path)
    wb.close()
    print(f"  Saved: {output_path}", flush=True)


# ---------------------------------------------------------------------------
# Supplemental metadata loader
# ---------------------------------------------------------------------------

def load_supplemental_metadata(paths: List[str]) -> Optional[pd.DataFrame]:
    """Load supplemental metadata files (CSV or Excel) and combine them.

    Expected columns (flexible naming):
      ISRC, Release Date, Artist Share, Third Party Splits, Reversion Date, License Term
    """
    if not paths:
        return None

    frames = []
    for p in paths:
        if not os.path.exists(p):
            print(f"  WARNING: Metadata file not found: {p}", flush=True)
            continue
        if p.endswith('.csv'):
            df = pd.read_csv(p)
        elif p.endswith('.xlsx') or p.endswith('.xls') or p.endswith('.xlsb'):
            engine = 'pyxlsb' if p.endswith('.xlsb') else None
            df = pd.read_excel(p, sheet_name=0, engine=engine)
        else:
            continue

        # Normalize column names
        col_remap = {}
        for c in df.columns:
            cl = c.strip().lower()
            if cl in ('isrc', 'identifier'):
                col_remap[c] = 'identifier'
            elif 'release' in cl and 'date' in cl:
                col_remap[c] = 'release_date'
            elif 'artist' in cl and 'share' in cl:
                col_remap[c] = 'artist_share'
            elif 'share' in cl and 'plygrnd' in cl:
                col_remap[c] = 'artist_share'
            elif '3p' in cl or 'third' in cl:
                col_remap[c] = 'third_party_share'
            elif 'reversion' in cl:
                col_remap[c] = 'reversion_date'
            elif 'license' in cl or 'term' in cl:
                col_remap[c] = 'license_term'

        df = df.rename(columns=col_remap)
        frames.append(df)

    if not frames:
        return None

    combined = pd.concat(frames, ignore_index=True)
    if 'identifier' in combined.columns:
        combined['identifier'] = combined['identifier'].astype(str).str.strip()
        combined = combined.drop_duplicates('identifier', keep='first')
    return combined


# ---------------------------------------------------------------------------
# Analytics helpers (used by both CLI and web app)
# ---------------------------------------------------------------------------

def compute_analytics(payor_results: Dict[str, PayorResult],
                      formulas: Optional[Dict[str, str]] = None,
                      enrichment_stats: Optional[dict] = None) -> dict:
    """Compute cross-payor analytics for the web app summary."""
    # Combine monthly data across all payors
    all_monthly = []
    all_meta = []
    total_files = 0

    for code, pr in payor_results.items():
        m = pr.monthly.copy()
        m['payor'] = pr.config.name
        all_monthly.append(m)

        meta = pr.isrc_meta.copy()
        meta['payor'] = pr.config.name
        all_meta.append(meta)

        total_files += pr.file_count

    monthly = pd.concat(all_monthly, ignore_index=True)
    meta = pd.concat(all_meta, ignore_index=True)

    # --- Basic stats ---
    total_gross = monthly['gross'].sum()
    total_net = monthly['net'].sum()
    periods = sorted(monthly['period'].unique().tolist())
    period_range = f"{min(periods)} - {max(periods)}" if periods else "N/A"

    # Unique ISRCs across all payors
    all_isrcs = set()
    for code, pr in payor_results.items():
        all_isrcs.update(pr.isrc_meta['identifier'].tolist())

    # --- Top songs (cross-payor) ---
    cross = (
        meta.groupby('identifier')
        .agg({'title': 'first', 'artist': 'first', 'total_gross': 'sum'})
        .reset_index()
        .sort_values('total_gross', ascending=False)
    )
    # Pre-compute per-song yearly data for top 20 YoY
    monthly['year'] = monthly['period'].astype(str).str[:4].astype(int)
    top_20_isrcs = cross.head(20)['identifier'].tolist()
    song_yearly = (
        monthly[monthly['identifier'].isin(top_20_isrcs)]
        .groupby(['identifier', 'year'])
        .agg({'gross': 'sum', 'net': 'sum'})
        .reset_index()
        .sort_values(['identifier', 'year'])
    )

    top_songs = []
    for _, row in cross.head(20).iterrows():
        isrc = str(row['identifier'])
        sy = song_yearly[song_yearly['identifier'] == isrc].sort_values('year')
        yearly = []
        for _, yr_row in sy.iterrows():
            yearly.append({'year': int(yr_row['year']),
                           'gross': round(float(yr_row['gross']), 2),
                           'net': round(float(yr_row['net']), 2)})
        # Compute latest YoY change for this song
        song_yoy = []
        yr_list = sy['year'].tolist()
        for i in range(1, len(yr_list)):
            prev_g = float(sy[sy['year'] == yr_list[i - 1]]['gross'].iloc[0])
            curr_g = float(sy[sy['year'] == yr_list[i]]['gross'].iloc[0])
            pct = ((curr_g - prev_g) / prev_g * 100) if prev_g > 0 else 0.0
            song_yoy.append({
                'period': f"{yr_list[i-1]}\u2192{yr_list[i]}",
                'pct': round(pct, 1),
                'direction': 'up' if pct >= 0 else 'down',
            })
        pct = (float(row['total_gross']) / total_gross * 100) if total_gross > 0 else 0.0
        top_songs.append({
            'isrc': isrc,
            'artist': str(row['artist'])[:30],
            'title': str(row['title'])[:40],
            'gross': f"{row['total_gross']:,.2f}",
            'gross_raw': round(float(row['total_gross']), 2),
            'net_raw': round(float(row.get('total_net', 0)), 2) if 'total_net' in row.index else 0.0,
            'pct_of_total': round(pct, 1),
            'yearly': yearly,
            'yoy': song_yoy,
        })

    # --- Annual gross earnings ---
    annual = (
        monthly.groupby('year')
        .agg({'gross': 'sum', 'net': 'sum'})
        .reset_index()
        .sort_values('year')
    )
    annual_earnings = []
    for _, row in annual.iterrows():
        annual_earnings.append({
            'year': int(row['year']),
            'gross': f"{row['gross']:,.2f}",
            'gross_raw': round(float(row['gross']), 2),
            'net': f"{row['net']:,.2f}",
            'net_raw': round(float(row['net']), 2),
        })

    # --- LTM (Last Twelve Months) earnings by song ---
    if not periods:
        # No data at all – return minimal analytics
        csym = _currency_symbol(payor_results)
        return {
            'total_files': sum(len(pr.file_inventory) for pr in payor_results.values()),
            'isrc_count': '0', 'isrc_count_raw': 0,
            'total_gross': '0', 'total_gross_raw': 0,
            'total_net': '0', 'total_net_raw': 0,
            'period_range': 'N/A',
            'top_songs': [], 'annual_earnings': [],
            'ltm_songs': [], 'yoy_decay': [],
            'payor_summaries': [],
            'monthly_trend': [], 'monthly_by_payor': {},
            'ltm_by_payor': {}, 'annual_by_payor': {},
            'top_stores': [], 'ltm_stores': [],
            'ltm_media_types': [],
            'ltm_gross_total': 0, 'ltm_gross_total_fmt': csym + '0',
            'ltm_net_total': 0, 'ltm_net_total_fmt': csym + '0',
            'ltm_yoy_pct': 0, 'ltm_yoy_direction': 'flat',
            'earnings_matrix': {}, 'earnings_years': [],
            'earnings_year_totals': {}, 'earnings_grand_total': 0,
            'earnings_grand_total_fmt': csym + '0',
            'coverage_months': [], 'coverage_rows': [],
            'currency_symbol': csym,
            'currency_code': _currency_code(payor_results),
            'payor_currencies': {code: pr.config.source_currency for code, pr in payor_results.items()},
            'waterfall': _compute_waterfall(payor_results, formulas),
            'weighted_avg_age': _compute_weighted_average_age(payor_results),
            'source_breakdown': _compute_source_breakdown(enrichment_stats),
            'cohort_analysis': {'cohorts': [], 'years': []},
            'revenue_concentration': {'top_1_pct': 0, 'top_5_pct': 0, 'top_10_pct': 0,
                                      'top_20_pct': 0, 'top_50_pct': 0, 'herfindahl_index': 0},
            'catalog_age_distribution': {'buckets': []},
            'ltm_comparison': None,
        }
    max_period = max(periods)
    max_year = int(str(max_period)[:4])
    max_month = int(str(max_period)[4:6])
    # Go back 12 months
    ltm_start_year = max_year - 1 if max_month < 12 else max_year
    ltm_start_month = max_month + 1 if max_month < 12 else 1
    if max_month == 12:
        ltm_start_year = max_year
        ltm_start_month = 1
    else:
        ltm_start_year = max_year - 1
        ltm_start_month = max_month + 1
    ltm_start_period = ltm_start_year * 100 + ltm_start_month

    ltm_monthly = monthly[monthly['period'] >= ltm_start_period]

    # --- LTM gross & net totals + YoY% ---
    ltm_gross_total = float(ltm_monthly['gross'].sum())
    ltm_net_total = float(ltm_monthly['net'].sum())

    # Prior-LTM window (12 months before LTM start) for YoY comparison
    prior_ltm_end_period = ltm_start_period - 1  # month just before LTM start
    prior_end_year = int(str(prior_ltm_end_period)[:4])
    prior_end_month = int(str(prior_ltm_end_period)[4:6])
    if prior_end_month == 0:
        prior_end_year -= 1
        prior_end_month = 12
        prior_ltm_end_period = prior_end_year * 100 + prior_end_month
    prior_start_year = prior_end_year - 1 if prior_end_month < 12 else prior_end_year
    prior_start_month = prior_end_month + 1 if prior_end_month < 12 else 1
    if prior_end_month == 12:
        prior_start_year = prior_end_year
        prior_start_month = 1
    else:
        prior_start_year = prior_end_year - 1
        prior_start_month = prior_end_month + 1
    prior_ltm_start_period = prior_start_year * 100 + prior_start_month

    prior_ltm_monthly = monthly[
        (monthly['period'] >= prior_ltm_start_period) &
        (monthly['period'] <= prior_ltm_end_period)
    ]
    prior_ltm_gross = float(prior_ltm_monthly['gross'].sum())
    prior_ltm_net = float(prior_ltm_monthly['net'].sum())
    prior_ltm_isrcs = int(prior_ltm_monthly['identifier'].nunique()) if len(prior_ltm_monthly) > 0 else 0

    if prior_ltm_gross > 0:
        ltm_yoy_pct = (ltm_gross_total - prior_ltm_gross) / prior_ltm_gross * 100
    else:
        ltm_yoy_pct = 0.0
    ltm_yoy_direction = 'up' if ltm_yoy_pct >= 0 else 'down'

    # Structured prior-LTM comparison (A4)
    ltm_isrc_count = int(ltm_monthly['identifier'].nunique()) if len(ltm_monthly) > 0 else 0
    prior_ltm = {
        'gross': round(prior_ltm_gross, 2),
        'net': round(prior_ltm_net, 2),
        'isrc_count': prior_ltm_isrcs,
        'avg_per_isrc': round(prior_ltm_gross / prior_ltm_isrcs, 2) if prior_ltm_isrcs > 0 else 0,
    }
    ltm_comparison = {
        'ltm': {
            'gross': round(ltm_gross_total, 2),
            'net': round(ltm_net_total, 2),
            'isrc_count': ltm_isrc_count,
            'avg_per_isrc': round(ltm_gross_total / ltm_isrc_count, 2) if ltm_isrc_count > 0 else 0,
        },
        'prior_ltm': prior_ltm,
        'changes': {
            'gross_pct': round(ltm_yoy_pct, 1),
            'net_pct': round(((ltm_net_total - prior_ltm_net) / prior_ltm_net * 100) if prior_ltm_net > 0 else 0, 1),
            'isrc_pct': round(((ltm_isrc_count - prior_ltm_isrcs) / prior_ltm_isrcs * 100) if prior_ltm_isrcs > 0 else 0, 1),
            'avg_per_isrc_pct': round((
                ((ltm_gross_total / ltm_isrc_count if ltm_isrc_count else 0) -
                 (prior_ltm_gross / prior_ltm_isrcs if prior_ltm_isrcs else 0)) /
                (prior_ltm_gross / prior_ltm_isrcs if prior_ltm_isrcs else 1) * 100
            ) if prior_ltm_isrcs > 0 else 0, 1),
        },
    }

    ltm_by_song = (
        ltm_monthly.groupby('identifier')
        .agg({'gross': 'sum'})
        .reset_index()
    )
    # Merge with metadata for titles
    ltm_by_song = ltm_by_song.merge(
        cross[['identifier', 'title', 'artist']], on='identifier', how='left'
    )
    ltm_by_song = ltm_by_song.sort_values('gross', ascending=False)

    ltm_songs = []
    for _, row in ltm_by_song.head(20).iterrows():
        ltm_pct = (float(row['gross']) / ltm_gross_total * 100) if ltm_gross_total > 0 else 0.0
        ltm_songs.append({
            'isrc': str(row['identifier']),
            'artist': str(row['artist'])[:30],
            'title': str(row['title'])[:40],
            'gross': f"{row['gross']:,.2f}",
            'gross_raw': round(float(row['gross']), 2),
            'pct_of_total': round(ltm_pct, 1),
        })

    # Add LTM gross to top_songs (hero card) and re-sort by LTM gross descending
    ltm_gross_lookup = dict(zip(ltm_by_song['identifier'].astype(str), ltm_by_song['gross']))
    for song in top_songs:
        ltm_val = ltm_gross_lookup.get(song['isrc'], 0.0)
        song['ltm_gross'] = f"{ltm_val:,.2f}"
        song['_ltm_gross_raw'] = ltm_val
    top_songs.sort(key=lambda s: s.get('_ltm_gross_raw', 0.0), reverse=True)

    # --- YoY decay analysis ---
    # Calculate year-over-year growth rates for the catalog
    annual_totals = annual.set_index('year')['gross']
    yoy_decay = []
    years_list = sorted(annual_totals.index.tolist())
    for i in range(1, len(years_list)):
        prev_year = years_list[i - 1]
        curr_year = years_list[i]
        prev_val = annual_totals[prev_year]
        curr_val = annual_totals[curr_year]
        if prev_val > 0:
            pct_change = (curr_val - prev_val) / prev_val * 100
        else:
            pct_change = 0
        yoy_decay.append({
            'period': f"{prev_year} → {curr_year}",
            'prev_gross': f"{prev_val:,.2f}",
            'prev_gross_raw': round(float(prev_val), 2),
            'curr_gross': f"{curr_val:,.2f}",
            'curr_gross_raw': round(float(curr_val), 2),
            'change_pct': f"{pct_change:+.1f}%",
        })

    # --- Per-payor summary ---
    month_labels = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                     'July', 'August', 'September', 'October', 'November', 'December']
    payor_summaries = []
    for code, pr in payor_results.items():
        # Latest statement month
        payor_periods = sorted(pr.monthly['period'].unique().tolist())
        if payor_periods:
            latest_p = max(payor_periods)
            latest_yr = int(str(latest_p)[:4])
            latest_mo = int(str(latest_p)[4:6])
            latest_statement = f"{month_labels[latest_mo]} {latest_yr}"
            latest_period = int(latest_p)
        else:
            latest_statement = 'N/A'
            latest_period = 0

        # Missing months detection
        # Use expected_start/expected_end from config if provided, else use data min/max
        missing_months = []
        range_min = pr.config.expected_start if pr.config.expected_start else (min(payor_periods) if payor_periods else None)
        range_max = pr.config.expected_end if pr.config.expected_end else (max(payor_periods) if payor_periods else None)
        if range_min and range_max:
            min_yr, min_mo = int(str(range_min)[:4]), int(str(range_min)[4:6])
            max_yr, max_mo = int(str(range_max)[:4]), int(str(range_max)[4:6])
            actual_periods = set(int(p) for p in payor_periods)
            y, m = min_yr, min_mo
            while y * 100 + m <= max_yr * 100 + max_mo:
                if y * 100 + m not in actual_periods:
                    missing_months.append(f"{month_labels[m]} {y}")
                m += 1
                if m > 12:
                    m = 1
                    y += 1
        expected_range = f"{range_min} - {range_max}" if range_min and range_max else None
        range_source = 'manual' if pr.config.expected_start or pr.config.expected_end else 'data'

        # Per-payor annual breakdown & YoY
        pm = pr.monthly.copy()
        pm['year'] = pm['period'].astype(str).str[:4].astype(int)
        pa = pm.groupby('year').agg({'gross': 'sum', 'net': 'sum'}).reset_index().sort_values('year')
        annual_breakdown = []
        for _, r in pa.iterrows():
            annual_breakdown.append({
                'year': int(r['year']),
                'gross': round(float(r['gross']), 2),
                'gross_fmt': f"{r['gross']:,.2f}",
                'net': round(float(r['net']), 2),
                'net_fmt': f"{r['net']:,.2f}",
            })
        yoy_changes = []
        pa_years = pa['year'].tolist()
        for i in range(1, len(pa_years)):
            prev_g = float(pa[pa['year'] == pa_years[i-1]]['gross'].iloc[0])
            curr_g = float(pa[pa['year'] == pa_years[i]]['gross'].iloc[0])
            pct = ((curr_g - prev_g) / prev_g * 100) if prev_g > 0 else 0.0
            yoy_changes.append({
                'period': f"{pa_years[i-1]}\u2192{pa_years[i]}",
                'pct': round(pct, 1),
                'direction': 'up' if pct >= 0 else 'down',
            })

        summary = {
            'code': code,
            'name': pr.config.name,
            'files': pr.file_count,
            'isrcs': len(pr.isrc_meta),
            'total_gross': f"{pr.isrc_meta['total_gross'].sum():,.2f}",
            'total_gross_raw': round(float(pr.isrc_meta['total_gross'].sum()), 2),
            'currency_code': pr.config.source_currency,
            'fee': f"{pr.config.fee:.0%}",
            'fx': pr.config.source_currency,
            'currency_symbol': _CURRENCY_SYMBOLS.get(_resolve_currency(pr), '$'),
            'detected_currency': ', '.join(getattr(pr, 'detected_currencies', [])),
            'statement_type': STATEMENT_TYPES.get(pr.config.statement_type, pr.config.statement_type),
            'deal_type': getattr(pr.config, 'deal_type', 'artist'),
            'artist_split': pr.config.artist_split,
            'territory': pr.config.territory,
            'contract_summary': getattr(pr.config, 'contract_summary', None),
            'latest_statement': latest_statement,
            'latest_period': latest_period,
            'missing_months': missing_months,
            'missing_count': len(missing_months),
            'expected_range': expected_range,
            'range_source': range_source,
            'annual_breakdown': annual_breakdown,
            'yoy_changes': yoy_changes,
            'file_inventory': getattr(pr, 'file_inventory', []),
            'quality_warnings': getattr(pr, 'quality_warnings', []),
        }
        payor_summaries.append(summary)

    # --- Statement coverage matrix (payor × month grid) ---
    # Build full range of months across all payors
    all_payor_periods = {}
    for code, pr in payor_results.items():
        all_payor_periods[code] = set(int(p) for p in pr.monthly['period'].unique().tolist())

    global_min = min(periods)
    global_max = max(periods)
    g_min_yr, g_min_mo = int(str(global_min)[:4]), int(str(global_min)[4:6])
    g_max_yr, g_max_mo = int(str(global_max)[:4]), int(str(global_max)[4:6])
    coverage_months = []
    y, m = g_min_yr, g_min_mo
    short_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    while y * 100 + m <= g_max_yr * 100 + g_max_mo:
        coverage_months.append({
            'period': y * 100 + m,
            'label': f"{short_names[m]} {y}",
            'short': f"{short_names[m]} '{str(y)[2:]}",
        })
        m += 1
        if m > 12:
            m = 1
            y += 1

    coverage_rows = []
    for code, pr in payor_results.items():
        payor_set = all_payor_periods[code]
        cells = []
        missing = []
        for cm in coverage_months:
            has = cm['period'] in payor_set
            cells.append({'period': cm['period'], 'has': has})
            if not has:
                missing.append(cm['label'])
        coverage_rows.append({
            'code': code,
            'name': pr.config.name,
            'cells': cells,
            'total': len(payor_set),
            'expected': len(coverage_months),
            'missing_list': missing,
            'missing_count': len(missing),
        })

    # --- Monthly trend (all payors combined, for charts) ---
    monthly_agg = (
        monthly.groupby('period')
        .agg({'gross': 'sum', 'net': 'sum'})
        .reset_index()
        .sort_values('period')
    )
    monthly_trend = []
    for _, row in monthly_agg.iterrows():
        p = int(row['period'])
        yr = int(str(p)[:4])
        mo = int(str(p)[4:6])
        month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                       'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        label = f"{month_names[mo]} {yr}"
        monthly_trend.append({
            'period': p,
            'label': label,
            'gross': round(float(row['gross']), 2),
            'net': round(float(row['net']), 2),
        })

    # --- Monthly trend per payor (for stacked charts) ---
    monthly_by_payor = {}
    for code, pr in payor_results.items():
        payor_monthly = (
            pr.monthly.groupby('period')
            .agg({'gross': 'sum', 'net': 'sum'})
            .reset_index()
            .sort_values('period')
        )
        entries = []
        for _, row in payor_monthly.iterrows():
            entries.append({
                'period': int(row['period']),
                'gross': round(float(row['gross']), 2),
                'net': round(float(row['net']), 2),
            })
        monthly_by_payor[code] = entries

    # --- LTM by payor ---
    ltm_by_payor = []
    payor_csyms = _payor_currency_symbols(payor_results)
    for code, pr in payor_results.items():
        payor_ltm = pr.monthly[pr.monthly['period'] >= ltm_start_period]
        ltm_gross = float(payor_ltm['gross'].sum())
        ltm_net = float(payor_ltm['net'].sum())
        ltm_by_payor.append({
            'code': code,
            'name': pr.config.name,
            'ltm_gross': round(ltm_gross, 2),
            'ltm_gross_fmt': f"{ltm_gross:,.2f}",
            'ltm_net': round(ltm_net, 2),
            'ltm_net_fmt': f"{ltm_net:,.2f}",
            'currency_symbol': payor_csyms.get(code, '$'),
            'currency_code': pr.config.source_currency,
        })

    # --- Annual by payor (for stacked bar chart) ---
    annual_by_payor = {}
    for code, pr in payor_results.items():
        pm = pr.monthly.copy()
        pm['year'] = pm['period'].astype(str).str[:4].astype(int)
        pa = pm.groupby('year').agg({'gross': 'sum'}).reset_index().sort_values('year')
        annual_by_payor[code] = [
            {'year': int(r['year']), 'gross': round(float(r['gross']), 2)}
            for _, r in pa.iterrows()
        ]

    # --- Earnings matrix: rows=payors, columns=years ---
    all_years = sorted(set(y for entries in annual_by_payor.values() for y in [e['year'] for e in entries]))
    earnings_matrix = []
    for code, pr in payor_results.items():
        lookup = {e['year']: e['gross'] for e in annual_by_payor.get(code, [])}
        pm2 = pr.monthly.copy()
        pm2['year'] = pm2['period'].astype(str).str[:4].astype(int)
        pa2 = pm2.groupby('year').agg({'net': 'sum'}).reset_index()
        net_lookup = {int(r['year']): round(float(r['net']), 2) for _, r in pa2.iterrows()}
        row_total_gross = sum(lookup.get(y, 0) for y in all_years)
        row_total_net = sum(net_lookup.get(y, 0) for y in all_years)
        earnings_matrix.append({
            'code': code,
            'name': pr.config.name,
            'currency_symbol': payor_csyms.get(code, '$'),
            'currency_code': pr.config.source_currency,
            'years': {y: {'gross': round(lookup.get(y, 0), 2),
                          'gross_fmt': f"{lookup.get(y, 0):,.2f}",
                          'net': round(net_lookup.get(y, 0), 2),
                          'net_fmt': f"{net_lookup.get(y, 0):,.2f}"}
                      for y in all_years},
            'total_gross': round(row_total_gross, 2),
            'total_gross_fmt': f"{row_total_gross:,.2f}",
            'total_net': round(row_total_net, 2),
            'total_net_fmt': f"{row_total_net:,.2f}",
        })

    # Earnings year column totals
    earnings_year_totals = {}
    for y in all_years:
        col_gross = sum(row['years'][y]['gross'] for row in earnings_matrix)
        col_net = sum(row['years'][y]['net'] for row in earnings_matrix)
        earnings_year_totals[y] = {
            'gross': round(col_gross, 2),
            'gross_fmt': f"{col_gross:,.2f}",
            'net': round(col_net, 2),
            'net_fmt': f"{col_net:,.2f}",
        }
    earnings_grand_total = sum(row['total_gross'] for row in earnings_matrix)

    # --- Top stores across all payors (all-time) ---
    all_dist = []
    for code, pr in payor_results.items():
        if not hasattr(pr, 'by_store') or pr.by_store is None:
            continue
        d = pr.by_store.copy()
        d.columns = ['store', 'gross', 'net', 'sales']
        all_dist.append(d)
    if all_dist:
        combined_dist = pd.concat(all_dist, ignore_index=True)
        top_stores_df = (
            combined_dist.groupby('store')
            .agg({'gross': 'sum', 'net': 'sum', 'sales': 'sum'})
            .reset_index()
            .sort_values('gross', ascending=False)
        )
        dist_list = []
        for _, row in top_stores_df.head(15).iterrows():
            dist_list.append({
                'name': str(row['store']),
                'gross': round(float(row['gross']), 2),
                'gross_fmt': f"{row['gross']:,.2f}",
                'sales': int(row['sales']),
            })
    else:
        dist_list = []

    # --- LTM stores & media types ---
    all_ltm_detail = []
    for code, pr in payor_results.items():
        if not hasattr(pr, 'detail') or pr.detail is None:
            continue
        d = pr.detail.copy()
        d = d[d['period'] >= ltm_start_period]
        all_ltm_detail.append(d)
    ltm_dist_list = []
    ltm_media_types = []
    if all_ltm_detail:
        ltm_detail = pd.concat(all_ltm_detail, ignore_index=True)
        # LTM stores
        if 'store' in ltm_detail.columns:
            ltm_dist = (
                ltm_detail.groupby('store')
                .agg({'gross': 'sum', 'net': 'sum', 'sales': 'sum'})
                .reset_index()
                .sort_values('gross', ascending=False)
            )
            for _, row in ltm_dist.head(15).iterrows():
                name = str(row['store']).strip()
                if name:
                    ltm_dist_list.append({
                        'name': name,
                        'gross': round(float(row['gross']), 2),
                        'gross_fmt': f"{row['gross']:,.2f}",
                        'sales': int(row['sales']),
                    })
        # LTM media types
        if 'media_type' in ltm_detail.columns:
            ltm_types = (
                ltm_detail.groupby('media_type')
                .agg({'gross': 'sum', 'net': 'sum', 'sales': 'sum'})
                .reset_index()
                .sort_values('gross', ascending=False)
            )
            for _, row in ltm_types.iterrows():
                name = str(row['media_type']).strip()
                if name:
                    ltm_media_types.append({
                        'name': name,
                        'gross': round(float(row['gross']), 2),
                        'gross_fmt': f"{row['gross']:,.2f}",
                        'sales': int(row['sales']),
                    })

    # --- Earnings waterfall ---
    waterfall = _compute_waterfall(payor_results, formulas)

    # --- Weighted average age ---
    weighted_avg_age = _compute_weighted_average_age(payor_results, ltm_start_period)

    # --- Source breakdown ---
    source_breakdown = _compute_source_breakdown(enrichment_stats)

    # --- Phase 4: Cohort analysis, Revenue concentration, Age distribution ---
    cohort_analysis = _compute_cohort_analysis(payor_results, ltm_start_period, monthly, meta)
    revenue_concentration = _compute_revenue_concentration(ltm_monthly, ltm_gross_total)
    catalog_age_distribution = _compute_catalog_age_distribution(payor_results, ltm_monthly, ltm_gross_total)

    return {
        'total_files': total_files,
        'isrc_count': f"{len(all_isrcs):,}",
        'isrc_count_raw': len(all_isrcs),
        'total_gross': f"{total_gross:,.2f}",
        'total_gross_raw': round(float(total_gross), 2),
        'total_net': f"{total_net:,.2f}",
        'total_net_raw': round(float(total_net), 2),
        'period_range': period_range,
        'top_songs': top_songs,
        'annual_earnings': annual_earnings,
        'ltm_songs': ltm_songs,
        'yoy_decay': yoy_decay,
        'payor_summaries': payor_summaries,
        'monthly_trend': monthly_trend,
        'monthly_by_payor': monthly_by_payor,
        'ltm_by_payor': ltm_by_payor,
        'annual_by_payor': annual_by_payor,
        'top_stores': dist_list,
        'ltm_stores': ltm_dist_list,
        'ltm_media_types': ltm_media_types,
        # LTM totals
        'ltm_gross_total': round(ltm_gross_total, 2),
        'ltm_gross_total_fmt': f"{ltm_gross_total:,.2f}",
        'ltm_net_total': round(ltm_net_total, 2),
        'ltm_net_total_fmt': f"{ltm_net_total:,.2f}",
        'ltm_yoy_pct': round(ltm_yoy_pct, 1),
        'ltm_yoy_direction': ltm_yoy_direction,
        # Earnings matrix
        'earnings_matrix': earnings_matrix,
        'earnings_years': all_years,
        'earnings_year_totals': earnings_year_totals,
        'earnings_grand_total': round(earnings_grand_total, 2),
        'earnings_grand_total_fmt': f"{earnings_grand_total:,.2f}",
        # Statement coverage
        'coverage_months': coverage_months,
        'coverage_rows': coverage_rows,
        # Currency
        'currency_symbol': _currency_symbol(payor_results),
        'currency_code': _currency_code(payor_results),
        'payor_currencies': {code: pr.config.source_currency for code, pr in payor_results.items()},
        # Phase 3: Waterfall, WAA, Source Breakdown
        'waterfall': waterfall,
        'weighted_avg_age': weighted_avg_age,
        'source_breakdown': source_breakdown,
        # Phase 4: Cohort, Concentration, Age Distribution, LTM Comparison
        'cohort_analysis': cohort_analysis,
        'revenue_concentration': revenue_concentration,
        'catalog_age_distribution': catalog_age_distribution,
        'ltm_comparison': ltm_comparison,
    }


def _compute_waterfall(payor_results: Dict[str, 'PayorResult'],
                       formulas: Optional[Dict[str, str]] = None) -> dict:
    """Compute earnings waterfall: overall + per-payor breakdown."""
    overall = {'gross': 0, 'fees': 0, 'net_receipts': 0,
               'payable': 0, 'third_party': 0, 'net_earnings': 0}
    per_payor = {}

    for code, pr in payor_results.items():
        detail_23 = _build_detail_23col(pr, formulas=formulas)
        payor_wf = {
            'name': pr.config.name,
            'gross': round(float(detail_23['Gross Earnings'].sum()), 2),
            'fees': round(float(detail_23['Fees'].sum()), 2),
            'net_receipts': round(float(detail_23['Net Receipts'].sum()), 2),
            'payable': round(float(detail_23['Payable Share'].sum()), 2),
            'third_party': round(float(detail_23['Third Party Share'].sum()), 2),
            'net_earnings': round(float(detail_23['Net Earnings'].sum()), 2),
        }
        per_payor[code] = payor_wf
        for k in overall:
            overall[k] += payor_wf[k]

    # Round overall
    for k in overall:
        overall[k] = round(overall[k], 2)

    return {
        'overall': overall,
        'per_payor': per_payor,
    }


def _compute_weighted_average_age(payor_results: Dict[str, 'PayorResult'],
                                   ltm_start_period: int = 0) -> dict:
    """Compute weighted average age of catalog by LTM net earnings.

    WAA = sum(Age_i * LTM_NetEarnings_i) / sum(LTM_NetEarnings_i)
    Age = (today - release_date) in fractional years.
    """
    from datetime import date as _date

    today = _date.today()
    total_weighted = 0.0
    total_earnings = 0.0
    tracks_with_dates = 0
    tracks_without_dates = 0

    for code, pr in payor_results.items():
        detail = pr.detail
        if detail is None or detail.empty:
            continue

        # Filter to LTM period
        if ltm_start_period:
            ltm_detail = detail[detail['period'] >= ltm_start_period]
        else:
            ltm_detail = detail

        # Group by ISRC to get LTM net earnings + release date
        if 'identifier' not in ltm_detail.columns or 'net' not in ltm_detail.columns:
            continue

        isrc_ltm = ltm_detail.groupby('identifier').agg({
            'net': 'sum',
            'release_date': 'first',
        }).reset_index()

        for _, row in isrc_ltm.iterrows():
            net_val = float(row['net']) if pd.notna(row['net']) else 0
            if net_val <= 0:
                continue

            rd = str(row.get('release_date', '')).strip()
            if not rd or rd in ('', 'nan', 'None', 'NaT'):
                tracks_without_dates += 1
                continue

            # Parse release date
            try:
                # Handle YYYY-MM-DD, YYYY-MM, YYYY, MM/DD/YYYY
                rd_clean = rd.replace('/', '-')
                parts = rd_clean.split('-')
                if len(parts) >= 3:
                    year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                elif len(parts) == 2:
                    year, month, day = int(parts[0]), int(parts[1]), 15
                else:
                    year, month, day = int(parts[0]), 6, 15

                release = _date(year, month, day)
                age_years = (today - release).days / 365.25
                if age_years < 0:
                    age_years = 0

                total_weighted += age_years * net_val
                total_earnings += net_val
                tracks_with_dates += 1
            except (ValueError, IndexError):
                tracks_without_dates += 1

    if total_earnings > 0:
        waa_years = total_weighted / total_earnings
    else:
        waa_years = 0

    total_tracks = tracks_with_dates + tracks_without_dates
    pct_coverage = round(tracks_with_dates / total_tracks * 100, 1) if total_tracks > 0 else 0

    # Format display
    if waa_years >= 1:
        waa_display = f"{waa_years:.1f} years"
    else:
        waa_display = f"{waa_years * 12:.0f} months"

    return {
        'waa_years': round(waa_years, 2),
        'waa_display': waa_display,
        'tracks_with_dates': tracks_with_dates,
        'tracks_without_dates': tracks_without_dates,
        'pct_coverage': pct_coverage,
        'ltm_period': ltm_start_period,
    }


def _compute_source_breakdown(enrichment_stats: Optional[dict]) -> dict:
    """Build source breakdown from enrichment stats."""
    if not enrichment_stats:
        return {'rows': [], 'total': 0}

    total = enrichment_stats.get('total', 0)
    if total == 0:
        return {'rows': [], 'total': 0}

    rows = []
    for key, label in [('from_source', 'Source Data (SRC)'),
                        ('from_cache', 'Cache'),
                        ('mb_found', 'MusicBrainz (MB)'),
                        ('gn_found', 'Genius (GN)'),
                        ('gm_found', 'Gemini (GM)'),
                        ('not_found', 'Not Found')]:
        count = enrichment_stats.get(key, 0)
        pct = round(count / total * 100, 1) if total > 0 else 0
        rows.append({'label': label, 'key': key, 'count': count, 'pct': pct})

    return {'rows': rows, 'total': total}


def _compute_cohort_analysis(payor_results, ltm_start_period, monthly, meta):
    """A1: Group ISRCs by release year, compute revenue per calendar year per cohort."""
    try:
        # Gather release dates from all payors
        release_dates = {}  # isrc -> release_year
        for code, pr in payor_results.items():
            detail = pr.detail if hasattr(pr, 'detail') and pr.detail is not None else None
            if detail is not None and 'Release Date' in detail.columns and 'ISRC' in detail.columns:
                for _, row in detail[['ISRC', 'Release Date']].drop_duplicates('ISRC').iterrows():
                    isrc = str(row['ISRC']).strip()
                    rd = str(row['Release Date']).strip()
                    if isrc and rd and len(rd) >= 4 and rd[:4].isdigit() and isrc not in release_dates:
                        release_dates[isrc] = int(rd[:4])
            # Also try isrc_meta if it has release_date column
            if hasattr(pr, 'isrc_meta') and 'release_date' in pr.isrc_meta.columns:
                for _, row in pr.isrc_meta.iterrows():
                    isrc = str(row['identifier']).strip()
                    rd = str(row.get('release_date', '')).strip()
                    if isrc and rd and len(rd) >= 4 and rd[:4].isdigit() and isrc not in release_dates:
                        release_dates[isrc] = int(rd[:4])

        if not release_dates:
            return {'cohorts': [], 'years': []}

        # Map identifiers in monthly to release years
        monthly_copy = monthly.copy()
        monthly_copy['release_year'] = monthly_copy['identifier'].map(release_dates)
        monthly_copy = monthly_copy.dropna(subset=['release_year'])
        monthly_copy['release_year'] = monthly_copy['release_year'].astype(int)

        if 'year' not in monthly_copy.columns:
            monthly_copy['year'] = monthly_copy['period'].astype(str).str[:4].astype(int)

        # Group by cohort (release_year) and calendar year
        cohort_rev = (
            monthly_copy.groupby(['release_year', 'year'])
            .agg({'gross': 'sum'})
            .reset_index()
        )

        # Get unique ISRCs per cohort for track count
        cohort_tracks = monthly_copy.groupby('release_year')['identifier'].nunique().to_dict()

        years = sorted(cohort_rev['year'].unique().tolist())
        release_years = sorted(cohort_rev['release_year'].unique().tolist())

        cohorts = []
        for ry in release_years:
            subset = cohort_rev[cohort_rev['release_year'] == ry]
            rev_by_year = {}
            for _, r in subset.iterrows():
                rev_by_year[int(r['year'])] = round(float(r['gross']), 2)
            cohorts.append({
                'release_year': ry,
                'track_count': int(cohort_tracks.get(ry, 0)),
                'revenue_by_year': rev_by_year,
            })

        return {'cohorts': cohorts, 'years': [int(y) for y in years]}
    except Exception as e:
        log.warning("Cohort analysis failed: %s", e)
        return {'cohorts': [], 'years': []}


def _compute_revenue_concentration(ltm_monthly, ltm_gross_total):
    """A2: LTM gross by ISRC sorted descending, cumulative share + Herfindahl."""
    try:
        if len(ltm_monthly) == 0 or ltm_gross_total <= 0:
            return {'top_1_pct': 0, 'top_5_pct': 0, 'top_10_pct': 0,
                    'top_20_pct': 0, 'top_50_pct': 0, 'herfindahl_index': 0}

        by_isrc = (
            ltm_monthly.groupby('identifier')
            .agg({'gross': 'sum'})
            .reset_index()
            .sort_values('gross', ascending=False)
        )
        n_total = len(by_isrc)
        if n_total == 0:
            return {'top_1_pct': 0, 'top_5_pct': 0, 'top_10_pct': 0,
                    'top_20_pct': 0, 'top_50_pct': 0, 'herfindahl_index': 0}

        gross_values = by_isrc['gross'].values
        cumsum = gross_values.cumsum()

        def top_n_share(n_tracks):
            if n_tracks <= 0:
                return 0
            n_tracks = min(n_tracks, n_total)
            return round(float(cumsum[n_tracks - 1]) / ltm_gross_total * 100, 1)

        # Calculate share for top N tracks (not percentage of catalog)
        top_1 = top_n_share(max(1, int(round(n_total * 0.01))))
        top_5 = top_n_share(max(1, int(round(n_total * 0.05))))
        top_10 = top_n_share(max(1, int(round(n_total * 0.10))))
        top_20 = top_n_share(max(1, int(round(n_total * 0.20))))
        top_50 = top_n_share(max(1, int(round(n_total * 0.50))))

        # Herfindahl-Hirschman Index (sum of squared market shares)
        shares = gross_values / ltm_gross_total
        hhi = round(float((shares ** 2).sum()) * 10000, 1)  # Scale to 0-10000

        return {
            'top_1_pct': top_1,
            'top_5_pct': top_5,
            'top_10_pct': top_10,
            'top_20_pct': top_20,
            'top_50_pct': top_50,
            'herfindahl_index': hhi,
            'total_isrcs': n_total,
        }
    except Exception as e:
        log.warning("Revenue concentration failed: %s", e)
        return {'top_1_pct': 0, 'top_5_pct': 0, 'top_10_pct': 0,
                'top_20_pct': 0, 'top_50_pct': 0, 'herfindahl_index': 0}


def _compute_catalog_age_distribution(payor_results, ltm_monthly, ltm_gross_total):
    """A3: ISRCs grouped by release year bucket with track count + LTM revenue."""
    try:
        # Gather release dates from all payors
        release_dates = {}  # isrc -> release_year
        for code, pr in payor_results.items():
            detail = pr.detail if hasattr(pr, 'detail') and pr.detail is not None else None
            if detail is not None and 'Release Date' in detail.columns and 'ISRC' in detail.columns:
                for _, row in detail[['ISRC', 'Release Date']].drop_duplicates('ISRC').iterrows():
                    isrc = str(row['ISRC']).strip()
                    rd = str(row['Release Date']).strip()
                    if isrc and rd and len(rd) >= 4 and rd[:4].isdigit() and isrc not in release_dates:
                        release_dates[isrc] = int(rd[:4])
            if hasattr(pr, 'isrc_meta') and 'release_date' in pr.isrc_meta.columns:
                for _, row in pr.isrc_meta.iterrows():
                    isrc = str(row['identifier']).strip()
                    rd = str(row.get('release_date', '')).strip()
                    if isrc and rd and len(rd) >= 4 and rd[:4].isdigit() and isrc not in release_dates:
                        release_dates[isrc] = int(rd[:4])

        if not release_dates:
            return {'buckets': []}

        # LTM gross by ISRC
        ltm_by_isrc = {}
        if len(ltm_monthly) > 0:
            grp = ltm_monthly.groupby('identifier')['gross'].sum()
            ltm_by_isrc = grp.to_dict()

        # Build buckets by release year
        year_data = {}  # year -> {'track_count': set, 'ltm_gross': float}
        for isrc, ry in release_dates.items():
            if ry not in year_data:
                year_data[ry] = {'tracks': set(), 'ltm_gross': 0.0}
            year_data[ry]['tracks'].add(isrc)
            year_data[ry]['ltm_gross'] += ltm_by_isrc.get(isrc, 0.0)

        buckets = []
        for year in sorted(year_data.keys()):
            d = year_data[year]
            ltm_g = round(d['ltm_gross'], 2)
            buckets.append({
                'year': year,
                'track_count': len(d['tracks']),
                'ltm_gross': ltm_g,
                'ltm_gross_pct': round(ltm_g / ltm_gross_total * 100, 1) if ltm_gross_total > 0 else 0,
            })

        return {'buckets': buckets}
    except Exception as e:
        log.warning("Catalog age distribution failed: %s", e)
        return {'buckets': []}


_CURRENCY_SYMBOLS = {'USD': '$', 'EUR': '\u20ac', 'GBP': '\u00a3', 'CAD': 'C$', 'AUD': 'A$', 'JPY': '\u00a5'}


def _resolve_currency(pr):
    """Resolve a payor's effective currency code ('auto' -> detected or USD)."""
    cur = pr.config.source_currency
    if cur and cur.lower() != 'auto':
        return cur
    # Use first detected currency from file parsing
    detected = getattr(pr, 'detected_currencies', [])
    if detected:
        # Filter out 'auto' from detected list
        real = [c for c in detected if c.lower() != 'auto']
        if real:
            return real[0]
    return 'USD'


def _currency_symbol(payor_results):
    """Determine display currency symbol from payor configs."""
    currencies = [_resolve_currency(pr) for pr in payor_results.values()]
    if not currencies:
        return '$'
    from collections import Counter
    most_common = Counter(currencies).most_common(1)[0][0]
    return _CURRENCY_SYMBOLS.get(most_common, most_common + ' ')


def _currency_code(payor_results):
    """Determine the primary ISO currency code from payor configs."""
    currencies = [_resolve_currency(pr) for pr in payor_results.values()]
    if not currencies:
        return 'USD'
    from collections import Counter
    return Counter(currencies).most_common(1)[0][0]


def _payor_currency_symbols(payor_results):
    """Return a dict mapping payor code -> currency symbol."""
    return {
        code: _CURRENCY_SYMBOLS.get(_resolve_currency(pr), _resolve_currency(pr) + ' ')
        for code, pr in payor_results.items()
    }


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------

def print_summary(payor_results: Dict[str, PayorResult]):
    """Print a summary to the console."""
    analytics = compute_analytics(payor_results)

    print("\n" + "=" * 65)
    print("  CONSOLIDATION SUMMARY")
    print("=" * 65)
    print(f"  Total files processed:  {analytics['total_files']}")
    print(f"  Unique ISRCs (all):     {analytics['isrc_count']}")
    print(f"  Date range:             {analytics['period_range']}")
    print(f"  Total gross (USD):      ${analytics['total_gross']}")
    print(f"  Total net (USD):        ${analytics['total_net']}")

    print("\n  PAYORS:")
    print("  " + "-" * 61)
    for ps in analytics['payor_summaries']:
        print(f"  {ps['name']:<20s}  {ps['files']:>3} files  {ps['isrcs']:>5} ISRCs  "
              f"${ps['total_gross']:>14s}  fee={ps['fee']}  {ps['fx']}")

    print("\n  ANNUAL GROSS EARNINGS:")
    print("  " + "-" * 40)
    for ae in analytics['annual_earnings']:
        print(f"    {ae['year']}:  ${ae['gross']:>14s}")

    print("\n  YoY CATALOG DECAY:")
    print("  " + "-" * 50)
    for d in analytics['yoy_decay']:
        print(f"    {d['period']}:  {d['change_pct']:>8s}  (${d['prev_gross']} → ${d['curr_gross']})")

    print("\n  TOP 10 EARNING SONGS (all payors):")
    print("  " + "-" * 61)
    for idx, song in enumerate(analytics['top_songs']):
        print(f"  {idx+1:>3}. ${song['gross']:>14s}  {song['artist']:<20s} - {song['title']}")
    print("=" * 65)


# ---------------------------------------------------------------------------
# Main CLI
# ---------------------------------------------------------------------------

# Default payor configs (can be overridden via CLI or web app)
DEFAULT_PAYORS = [
    PayorConfig(
        code='B1', name='Believe 15%', fmt='auto', fee=0.15,
        source_currency='EUR',
        statements_dir=r'C:\Users\jacques\Downloads\Believe_15_extracted',
    ),
    PayorConfig(
        code='B2', name='Believe 20%', fmt='auto', fee=0.20,
        source_currency='EUR',
        statements_dir=r'C:\Users\jacques\Downloads\Believe_20_extracted',
    ),
    PayorConfig(
        code='RJ', name='RecordJet', fmt='auto', fee=0.07,
        source_currency='EUR',
        statements_dir=r'C:\Users\jacques\Downloads\RecordJet_extracted',
    ),
]


def main():
    parser = argparse.ArgumentParser(description='Multi-payor royalty statement consolidator.')
    parser.add_argument('--template', default=r'C:\Users\jacques\Documents\202601_PLYGRND Model_v3.xlsx',
                        help='Path to the financial model template (with payor tabs)')
    parser.add_argument('--output', default=None, help='Path for populated model output')
    parser.add_argument('--consolidated', default=None, help='Path for consolidated data Excel')
    parser.add_argument('--metadata', nargs='*', default=[], help='Supplemental metadata files')

    args = parser.parse_args()

    out_dir = os.path.dirname(args.template)
    if args.output is None:
        args.output = os.path.join(out_dir, 'PLYGRND_Model_populated.xlsx')
    if args.consolidated is None:
        args.consolidated = os.path.join(out_dir, 'Consolidated_All_Payors.xlsx')

    print("\n  ROYALTY STATEMENT CONSOLIDATOR (Multi-Payor)")
    print("  " + "=" * 50)
    for cfg in DEFAULT_PAYORS:
        print(f"  {cfg.code}: {cfg.name} ({cfg.fmt}) fee={cfg.fee:.0%} dir={cfg.statements_dir}")
    print()

    # Step 1: Load all payors
    print("  [1/4] Parsing statement files for all payors...")
    payor_results = load_all_payors(DEFAULT_PAYORS)

    if not payor_results:
        print("  ERROR: No payor data loaded.")
        sys.exit(1)

    # Step 2: Write consolidated Excel
    print("\n  [2/4] Writing consolidated Excel...")
    write_consolidated_excel(payor_results, args.consolidated)

    # Step 3: Load supplemental metadata
    supplemental = None
    if args.metadata:
        print("\n  [2.5] Loading supplemental metadata...")
        supplemental = load_supplemental_metadata(args.metadata)
        if supplemental is not None:
            print(f"    Loaded {len(supplemental)} metadata records.", flush=True)

    # Step 4: Populate template
    print("\n  [3/4] Populating financial model template...")
    populate_template(args.template, args.output, payor_results, supplemental)

    # Step 5: Summary
    print_summary(payor_results)

    print(f"\n  OUTPUT FILES:")
    print(f"    Consolidated: {args.consolidated}")
    print(f"    Model:        {args.output}")
    print()


if __name__ == '__main__':
    main()
