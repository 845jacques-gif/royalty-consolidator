import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\jacques\Documents\202601_PLYGRND Model_v3.xlsx"

# Use read_only=True for much faster loading of large files
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

print(f"Sheet names: {wb.sheetnames}")
print()

# Focus on Payor_Model tabs and Metadata
focus_sheets = ['Payor_Model', 'Payor_Model (2)', 'Metadata', 'Song Listing_By Payor', 'Summary - Earnings']

for sheet_name in focus_sheets:
    if sheet_name not in wb.sheetnames:
        matches = [s for s in wb.sheetnames if sheet_name.lower() in s.lower()]
        if matches:
            sheet_name = matches[0]
        else:
            print(f"MISSING: {sheet_name}")
            continue
    
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    # read_only mode doesn't support max_row reliably, so we count
    print(f"{'='*80}")
    
    if 'Payor_Model' in sheet_name:
        row_num = 0
        last_rows = []
        for row in ws.rows:
            row_num += 1
            if row_num <= 55:
                row_data = []
                for cell in row[:35]:
                    if cell.value is not None:
                        val = str(cell.value)
                        if len(val) > 60:
                            val = val[:60] + "..."
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        row_data.append(f"{col_letter}{row_num}: {val}")
                if row_data:
                    print(" | ".join(row_data[:12]))
            elif row_num <= 60:
                if row_num == 56:
                    print("\n--- Data rows 56-60 ---")
                row_data = []
                for cell in row[:100]:
                    if cell.value is not None:
                        val = str(cell.value)
                        if len(val) > 40:
                            val = val[:40] + "..."
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        row_data.append(f"{col_letter}{row_num}: {val}")
                if row_data:
                    print(" | ".join(row_data[:15]))
            else:
                # Keep track of last few rows with data
                row_data = []
                for cell in row[:100]:
                    if cell.value is not None:
                        val = str(cell.value)
                        if len(val) > 40:
                            val = val[:40] + "..."
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        row_data.append(f"{col_letter}{row_num}: {val}")
                if row_data:
                    last_rows.append(" | ".join(row_data[:15]))
                    if len(last_rows) > 4:
                        last_rows.pop(0)
        
        print(f"\nTotal rows scanned: {row_num}")
        if last_rows:
            print(f"\n--- Last few data rows ---")
            for lr in last_rows:
                print(lr)
    
    elif 'Metadata' in sheet_name:
        row_num = 0
        total = 0
        for row in ws.rows:
            row_num += 1
            total += 1
            if row_num <= 15:
                row_data = []
                for cell in row[:15]:
                    if cell.value is not None:
                        val = str(cell.value)
                        if len(val) > 50:
                            val = val[:50] + "..."
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        row_data.append(f"{col_letter}{row_num}: {val}")
                if row_data:
                    print(" | ".join(row_data))
        print(f"\n... total rows: {total}")
    
    else:
        row_num = 0
        for row in ws.rows:
            row_num += 1
            if row_num > 20:
                # Just count remaining
                continue
            row_data = []
            for cell in row[:30]:
                if cell.value is not None:
                    val = str(cell.value)
                    if len(val) > 50:
                        val = val[:50] + "..."
                    col_letter = openpyxl.utils.get_column_letter(cell.column)
                    row_data.append(f"{col_letter}{row_num}: {val}")
            if row_data:
                print(" | ".join(row_data[:12]))
        print(f"Total rows: {row_num}")

    sys.stdout.flush()

wb.close()
print("\nDone.")
